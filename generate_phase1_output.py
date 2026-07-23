"""
generate_phase1_output.py
==========================
Phase 1 Output Generator - Creates pre-deduction member records

PURPOSE:
--------
This script generates Phase 1 outputs (JSON + Excel) containing all member data
with ORIGINAL amounts from BOTH financial columns in the invoice:
- Charge Amount (varies per line)
- Adjustment/Debit Amount (total for member)

EXTRACTION APPROACH:
--------------------
Phase 1 uses a SEPARATE extraction specifically designed to capture BOTH
financial columns from the invoice PDF, independent of the main RPVE flow.

WHAT IT DOES:
-------------
- Extracts member data from PDF invoices using dedicated Phase 1 extraction
- Captures TWO financial columns:
  * Charge Amount → current_premium (varies: $471, $216, $4.20, etc.)
  * Adjustment/Debit → adjustment_amount (member total: $691, etc.)
- Allows duplicate member names (same person can have multiple plan lines)
- Creates two output files:
  * {filename}_phase1.json - JSON format with all member fields
  * {filename}_phase1.xlsx - Excel format with all member fields

FIELDS INCLUDED:
----------------
- full_name
- first_name
- middal_name (middle name)
- last_name
- coverage
- plan_name
- plan_type
- current_premium (CHARGE AMOUNT from invoice - varies per line)
- adjustment_amount (ADJUSTMENT/DEBIT AMOUNT - member total)
- birth_date
- gender
- home_zip_code
- billing_period

DIFFERENCE FROM PHASE 2:
-------------------------
Phase 1 captures BOTH financial columns with original amounts.
Phase 2 applies $250 deduction to consolidated amounts.

USAGE:
------
1. As a module import:
   from generate_phase1_output import generate_phase1_output
   generate_phase1_output(pdf_path, output_dir)

2. As a CLI tool:
   python generate_phase1_output.py <pdf_path> [output_dir]

Example:
   python generate_phase1_output.py invoice.pdf ./outputs/
"""

import sys
import json
import logging
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import extraction logic from RPVE_standalone
try:
    from RPVE_standalone import (
        extract_text,
        classify,
        OUTPUT_DIR,
        client as openai_client
    )
except ImportError as e:
    print(f"Error: Could not import RPVE_standalone.py. {e}")
    print("Make sure this script is in the same directory as RPVE_standalone.py")
    sys.exit(1)

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# CHUNKING UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def split_text_by_pages(text: str) -> List[str]:
    """
    Splits text using '--- Page N ---' markers.
    Returns list of page texts (each page includes its marker).
    """
    pattern = re.compile(r'(--- Page \d+ ---)', re.IGNORECASE)
    matches = list(pattern.finditer(text))
    
    if not matches:
        # No page markers found - treat entire text as one page
        return [text]
    
    pages = []
    for i, match in enumerate(matches):
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        page_text = text[start:end]
        if page_text.strip():
            pages.append(page_text)
    
    return pages


def create_page_chunks(pages: List[str], pages_per_chunk: int = 4, overlap_pages: int = 2) -> List[str]:
    """
    Creates overlapping chunks from pages.
    
    Args:
        pages: List of page texts
        pages_per_chunk: Number of pages per chunk (default: 4)
        overlap_pages: Number of pages to overlap between chunks (default: 2)
    
    Returns:
        List of chunk texts (each chunk contains multiple pages joined together)
    
    Example with 4-page chunks and 2-page overlap:
        Pages: [1, 2, 3, 4, 5, 6, 7, 8]
        Chunks: 
            Chunk 1: [1, 2, 3, 4]
            Chunk 2: [3, 4, 5, 6]  ← 2-page overlap
            Chunk 3: [5, 6, 7, 8]  ← 2-page overlap
    """
    if not pages:
        return []
    
    chunks = []
    step = pages_per_chunk - overlap_pages  # How many pages to advance each iteration
    
    for i in range(0, len(pages), step):
        chunk_pages = pages[i:i + pages_per_chunk]
        if chunk_pages:
            chunks.append("\n".join(chunk_pages))
        
        # Stop if we've reached the end
        if i + pages_per_chunk >= len(pages):
            break
    
    return chunks


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 FIELD SCHEMA
# ══════════════════════════════════════════════════════════════════════════════

PHASE1_FIELDS = [
    "full_name",
    "first_name",
    "middal_name",  # keeping original spelling from requirements
    "last_name",
    "coverage",
    "plan_name",
    "plan_type",
    "current_premium",      # ← CHARGE AMOUNT (varies per line)
    "adjustment_amount",    # ← ADJUSTMENT/DEBIT AMOUNT (member total)
    "birth_date",
    "gender",
    "home_zip_code",
    "billing_period",
]

# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 DEDICATED EXTRACTION PROMPT
# ══════════════════════════════════════════════════════════════════════════════

PHASE1_EXTRACTION_PROMPT = """
You are extracting data from a group health insurance invoice for Phase 1 baseline records.

CRITICAL REQUIREMENT - EXTRACT TWO SEPARATE FINANCIAL COLUMNS:
===============================================================

The invoice "Details" section has columns like:
- Policy No. / Member Name
- Plan ID / Plan Description
- Coverage Type / Status
- **Charge Amount** (varies per plan line)
- **Adjustment Detail Totals** (appears on FIRST line only per member)

IMPORTANT PATTERN - READ CAREFULLY:
------------------------------------
When a member has MULTIPLE plan lines in the invoice:
1. **FIRST line** shows: Charge Amount + Adjustment Total (e.g., $471.32 and $691.76)
2. **SUBSEQUENT lines** show: Charge Amount ONLY (e.g., $216.24, $4.20) - adjustment column is blank
3. You MUST **carry forward** the Adjustment Total from the first line to ALL subsequent lines for that member

REAL INVOICE EXAMPLE:
---------------------
```
ADNAN, MOHD  | Admin/Excess Loss | $471.32 | $691.76  ← First line: BOTH amounts appear
ADNAN, MOHD  | Max Claims        | $216.24 | [blank]  ← Second: only charge, REUSE $691.76!
ADNAN, MOHD  | Vision            | $4.20   | [blank]  ← Third: only charge, REUSE $691.76!
```

YOU MUST CREATE 3 SEPARATE RECORDS:
Record 1: full_name="ADNAN, MOHD", plan_name="Admin/Excess Loss", current_premium=471.32, adjustment_amount=691.76
Record 2: full_name="ADNAN, MOHD", plan_name="Max Claims", current_premium=216.24, adjustment_amount=691.76 ← SAME adjustment!
Record 3: full_name="ADNAN, MOHD", plan_name="Vision", current_premium=4.20, adjustment_amount=691.76 ← SAME adjustment!

EXTRACTION LOGIC STEPS:
========================
1. For each member's FIRST plan line:
   - Extract charge amount → current_premium
   - Extract adjustment amount → adjustment_amount
   - Remember this adjustment amount for this member

2. For each SUBSEQUENT plan line of same member:
   - Extract charge amount → current_premium (different value)
   - REUSE the adjustment amount from first line → adjustment_amount (same value)

3. Create a SEPARATE record for EVERY plan line (duplicates required!)

4. Verify your output:
   - current_premium MUST vary per line (471.32 → 216.24 → 4.20)
   - adjustment_amount MUST stay same for all lines of one member (691.76 → 691.76 → 691.76)

FIELD DEFINITIONS:
==================
- full_name: Complete member name as it appears
- first_name: First name extracted from full_name
- middal_name: Middle name if present, else ""
- last_name: Last name extracted from full_name
- coverage: Coverage tier (E=Employee, ES=Employee+Spouse, FAM=Family, EC=Employee+Child, etc.)
- plan_name: Complete plan/product name including any ID codes
- plan_type: Category (Medical, Dental, Vision, Life, LTD, Admin, Excess Loss, Other)
- current_premium: **Charge Amount** from invoice (varies per plan line)
- adjustment_amount: **Adjustment Detail Totals** from invoice (member total - carry forward to all their lines)
- birth_date: Date of birth if shown, else ""
- gender: M or F if shown, else ""
- home_zip_code: ZIP code if shown, else ""
- billing_period: Coverage period dates if shown, else ""

VALIDATION - YOUR OUTPUT MUST MATCH THIS PATTERN:
==================================================
✅ CORRECT: current_premium varies (471.32, 216.24, 4.20) AND adjustment_amount same (691.76, 691.76, 691.76)
❌ WRONG: Both fields have same value (691.76, 691.76) on every record
✅ CORRECT: Member "ADNAN, MOHD" has 3-4 separate records (multiple plans)
❌ WRONG: Each member has only 1 record (consolidated)

Return ONLY valid JSON (no markdown, no code blocks, no explanation):
{{
  "summary": {{"company_name": "", "total_amount_due": ""}},
  "employees": [
    {{
      "full_name": "",
      "first_name": "",
      "middal_name": "",
      "last_name": "",
      "coverage": "",
      "plan_name": "",
      "plan_type": "",
      "current_premium": 0.00,
      "adjustment_amount": 0.00,
      "birth_date": "",
      "gender": "",
      "home_zip_code": "",
      "billing_period": ""
    }}
  ]
}}

INVOICE TEXT TO EXTRACT FROM:
{text}
"""

# ══════════════════════════════════════════════════════════════════════════════
# EXTRACTION AND DATA PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

def extract_phase1_data_with_chunking(text: str, doc_type: str) -> Dict:
    """
    Extract member data using page-based chunking with overlap and parallel processing.
    
    Args:
        text: Extracted PDF text
        doc_type: Document type classification
        
    Returns:
        Dictionary with 'summary' and 'employees' keys containing extracted data
    """
    # Step 1: Split text into pages
    logger.info("[Phase 1] Splitting text into pages...")
    pages = split_text_by_pages(text)
    logger.info(f"[Phase 1] Found {len(pages)} pages")
    
    # Step 2: Create overlapping chunks
    PAGES_PER_CHUNK = 4
    OVERLAP_PAGES = 2
    chunks = create_page_chunks(pages, pages_per_chunk=PAGES_PER_CHUNK, overlap_pages=OVERLAP_PAGES)
    logger.info(f"[Phase 1] Created {len(chunks)} chunks (4-page chunks with 2-page overlap)")
    
    # Step 3: Define chunk processing function
    def process_single_chunk(chunk_text: str, chunk_index: int, depth: int = 0) -> tuple[List[Dict], Dict]:
        """
        Process a single chunk with recursive error recovery.
        
        Returns:
            (employees_list, summary_dict)
        """
        label = f"Chunk {chunk_index + 1}/{len(chunks)}"
        
        # Empty chunk check
        if not chunk_text.strip():
            return [], {}
        
        # Prepare prompt for this chunk
        prompt = PHASE1_EXTRACTION_PROMPT.format(text=chunk_text)
        
        try:
            # Call LLM for extraction
            response = openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a precise data extraction assistant."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0
            )
            
            # Parse response
            raw_content = response.choices[0].message.content
            data = json.loads(raw_content)
            
            employees = data.get("employees", [])
            summary = data.get("summary", {})
            
            logger.info(f"[Phase 1] {label} → {len(employees)} records extracted")
            return employees, summary
            
        except json.JSONDecodeError as e:
            logger.warning(f"[Phase 1] {label} → JSON parsing failed: {e}")
            # Try to extract JSON from response
            try:
                # Remove markdown code blocks if present
                cleaned = re.sub(r"```json|```", "", raw_content).strip()
                data = json.loads(cleaned)
                employees = data.get("employees", [])
                summary = data.get("summary", {})
                logger.info(f"[Phase 1] {label} → {len(employees)} records extracted (after cleanup)")
                return employees, summary
            except:
                pass  # Fall through to recursive splitting
        
        except Exception as e:
            logger.warning(f"[Phase 1] {label} → Extraction failed: {e}")
        
        # RECURSIVE ERROR RECOVERY: Split in half and retry
        MIN_CHUNK_SIZE = 4000
        MAX_RECURSION_DEPTH = 3
        
        if len(chunk_text) > MIN_CHUNK_SIZE and depth < MAX_RECURSION_DEPTH:
            logger.info(f"[Phase 1] {label} → Splitting into sub-chunks (depth {depth + 1})...")
            
            # Find midpoint at nearest newline to avoid splitting records
            mid = len(chunk_text) // 2
            split_at = chunk_text.rfind('\n', 0, mid)
            
            if split_at == -1 or split_at < mid * 0.5:
                # No good newline found, use exact midpoint
                split_at = mid
            
            half_a = chunk_text[:split_at]
            half_b = chunk_text[split_at:]
            
            # Recursively process sub-chunks
            emps_a, summ_a = process_single_chunk(half_a, chunk_index, depth + 1)
            emps_b, summ_b = process_single_chunk(half_b, chunk_index, depth + 1)
            
            # Merge results
            merged_employees = emps_a + emps_b
            merged_summary = summ_a or summ_b  # Use first non-empty summary
            
            logger.info(f"[Phase 1] {label} → Sub-chunks merged: {len(merged_employees)} records")
            return merged_employees, merged_summary
        
        # Give up if chunk is too small or max depth reached
        logger.warning(f"[Phase 1] {label} → Failed permanently (size: {len(chunk_text)}, depth: {depth})")
        return [], {}
    
    # Step 4: Process chunks in PARALLEL
    logger.info(f"[Phase 1] Processing {len(chunks)} chunks in parallel...")
    
    all_employees = []
    final_summary = {}
    
    with ThreadPoolExecutor(max_workers=min(len(chunks), 10)) as executor:
        # Submit all chunks for parallel processing
        futures = [
            executor.submit(process_single_chunk, chunk, i)
            for i, chunk in enumerate(chunks)
        ]
        
        # Collect results as they complete
        for i, future in enumerate(futures):
            try:
                employees, summary = future.result()
                
                if employees:
                    all_employees.extend(employees)
                
                if summary and not final_summary:
                    final_summary = summary  # Use first non-empty summary
                    
            except Exception as e:
                logger.error(f"[Phase 1] Chunk {i + 1} processing error: {e}")
    
    logger.info(f"[Phase 1] All chunks processed → Total: {len(all_employees)} records")
    
    return {
        "summary": final_summary,
        "employees": all_employees
    }


def extract_phase1_data(pdf_path: Path) -> Dict:
    """
    Extract member data from PDF invoice using Phase 1 dedicated extraction.
    Uses intelligent chunking strategy for large documents.
    
    Args:
        pdf_path: Path to the PDF invoice file
        
    Returns:
        Dictionary with 'summary' and 'employees' keys containing extracted data
    """
    logger.info(f"[Phase 1] Starting dedicated extraction from: {pdf_path.name}")
    
    # Step 1: Extract text from PDF
    logger.info("[Phase 1] Extracting text from PDF...")
    text = extract_text(pdf_path)
    
    if not text or len(text.strip()) < 100:
        raise ValueError(f"Failed to extract text from PDF: {pdf_path.name}")
    
    logger.info(f"[Phase 1] Extracted {len(text):,} characters of text")
    
    # Step 2: Classify document type
    logger.info("[Phase 1] Classifying document type...")
    doc_type = classify(text)
    logger.info(f"[Phase 1] Document classified as: {doc_type}")
    
    # Step 3: Decide whether to use chunking or single call based on PAGE COUNT
    # Single call is faster for small invoices (1-2 pages)
    # Chunking with overlap is essential for multi-page invoices (3+ pages)
    
    # Count pages in the text
    pages = split_text_by_pages(text)
    page_count = len(pages)
    logger.info(f"[Phase 1] Detected {page_count} pages in document")
    
    if page_count <= 2:
        # Small document (1-2 pages) - use single call (faster, sufficient)
        logger.info(f"[Phase 1] Using single LLM call (document has {page_count} pages)")
        
        try:
            prompt = PHASE1_EXTRACTION_PROMPT.format(text=text)
            
            response = openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a precise data extraction assistant."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0
            )
            
            extracted_data = json.loads(response.choices[0].message.content)
            employees = extracted_data.get("employees", [])
            logger.info(f"[Phase 1] Extracted {len(employees)} employee plan lines (duplicates allowed)")
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"[Phase 1] Single call failed: {e}, falling back to chunking...")
            # Fall through to chunking strategy
    
    # Multi-page document (3+ pages) or single call failed - use chunking strategy
    logger.info(f"[Phase 1] Using chunking strategy (document has {page_count} pages)")
    return extract_phase1_data_with_chunking(text, doc_type)


def normalize_phase1_record(raw_record: Dict) -> Dict:
    """
    Normalize a raw extracted record to Phase 1 schema.
    Maps all fields and ensures proper data types.
    
    Args:
        raw_record: Raw employee record from extraction
        
    Returns:
        Normalized record matching PHASE1_FIELDS schema
    """
    normalized = {}
    
    # Helper function to clean "None" strings and null values
    def clean_value(value):
        if value is None or value == "None" or value == "null":
            return ""
        return str(value).strip()
    
    # Name fields
    normalized["full_name"] = clean_value(raw_record.get("full_name"))
    normalized["first_name"] = clean_value(raw_record.get("first_name"))
    normalized["middal_name"] = clean_value(raw_record.get("middal_name"))
    normalized["last_name"] = clean_value(raw_record.get("last_name"))
    
    # If full_name is missing but first/last exist, construct it
    if not normalized["full_name"] and (normalized["first_name"] or normalized["last_name"]):
        parts = []
        if normalized["first_name"]:
            parts.append(normalized["first_name"])
        if normalized["middal_name"]:
            parts.append(normalized["middal_name"])
        if normalized["last_name"]:
            parts.append(normalized["last_name"])
        normalized["full_name"] = " ".join(parts)
    
    # Coverage and plan fields
    normalized["coverage"] = clean_value(raw_record.get("coverage"))
    normalized["plan_name"] = clean_value(raw_record.get("plan_name"))
    normalized["plan_type"] = clean_value(raw_record.get("plan_type"))
    
    # Financial fields - NO $250 DEDUCTION
    # These are the ORIGINAL amounts from the invoice
    current_premium = raw_record.get("current_premium", "")
    if current_premium and current_premium != "None" and current_premium != "null":
        # Clean currency formatting
        if isinstance(current_premium, str):
            current_premium = current_premium.replace("$", "").replace(",", "").strip()
        try:
            normalized["current_premium"] = float(current_premium)
        except (ValueError, TypeError):
            normalized["current_premium"] = ""
    else:
        normalized["current_premium"] = ""
    
    # CRITICAL: adjustment_amount should equal current_premium in Phase 1
    # (both represent the original charge amount - no deductions)
    # If adjustment_amount is missing or null, use current_premium value
    adjustment_amount = raw_record.get("adjustment_amount", "")
    if adjustment_amount and adjustment_amount != "None" and adjustment_amount != "null" and adjustment_amount != "":
        # Clean currency formatting
        if isinstance(adjustment_amount, str):
            adjustment_amount = adjustment_amount.replace("$", "").replace(",", "").strip()
        try:
            normalized["adjustment_amount"] = float(adjustment_amount)
        except (ValueError, TypeError):
            # If parsing fails, use current_premium value
            normalized["adjustment_amount"] = normalized["current_premium"]
    else:
        # If adjustment_amount is not provided, use current_premium
        # (In Phase 1, both should show the same original amount)
        normalized["adjustment_amount"] = normalized["current_premium"]
    
    # Personal information fields
    normalized["birth_date"] = clean_value(raw_record.get("birth_date"))
    normalized["gender"] = clean_value(raw_record.get("gender"))
    normalized["home_zip_code"] = clean_value(raw_record.get("home_zip_code"))
    normalized["billing_period"] = clean_value(raw_record.get("billing_period"))
    
    return normalized


# ══════════════════════════════════════════════════════════════════════════════
# JSON OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def generate_phase1_json(data: Dict, output_path: Path) -> None:
    """
    Generate Phase 1 JSON output file.
    
    Args:
        data: Extracted data dictionary with 'summary' and 'employees'
        output_path: Path where JSON file should be saved
    """
    logger.info(f"[Phase 1] Generating JSON output: {output_path.name}")
    
    output_data = {
        "metadata": {
            "phase": "phase1",
            "description": "Pre-deduction member records (original charge amounts)",
            "generated_at": datetime.now().isoformat(),
            "record_count": len(data.get("employees", []))
        },
        "summary": data.get("summary", {}),
        "employees": []
    }
    
    # Normalize all employee records
    for raw_record in data.get("employees", []):
        normalized = normalize_phase1_record(raw_record)
        output_data["employees"].append(normalized)
    
    # Write JSON file
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    
    logger.info(f"[Phase 1] JSON output saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def generate_phase1_excel(data: Dict, output_path: Path) -> None:
    """
    Generate Phase 1 Excel output file with professional formatting.
    
    Args:
        data: Extracted data dictionary with 'summary' and 'employees'
        output_path: Path where Excel file should be saved
    """
    logger.info(f"[Phase 1] Generating Excel output: {output_path.name}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Phase 1 - Member Data"
    ws.sheet_view.showGridLines = False
    
    # Styling constants
    HEADER_COLOR = "4472C4"  # Blue
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(size=10, name="Calibri")
    cell_align = Alignment(vertical="center")
    
    thin_border = Side(style="thin", color="DDDDDD")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    alt_row_fill = PatternFill("solid", fgColor="F7F7F7")
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PHASE1_FIELDS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Phase 1 - Member Data (Original Amounts - No Deductions)"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Header row
    for col_idx, field in enumerate(PHASE1_FIELDS, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = field.replace("_", " ").title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        
        # Set column widths
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    ws.row_dimensions[2].height = 25
    
    # Data rows
    employees = data.get("employees", [])
    for row_idx, raw_record in enumerate(employees, 3):
        normalized = normalize_phase1_record(raw_record)
        ws.row_dimensions[row_idx].height = 20
        
        for col_idx, field in enumerate(PHASE1_FIELDS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = normalized.get(field, "")
            
            # Format financial fields
            if field in ["current_premium", "adjustment_amount"] and value:
                try:
                    cell.value = float(value)
                    cell.number_format = '$#,##0.00'
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = value
            
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = border
            
            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill
    
    # Summary row (totals for financial columns)
    if employees:
        total_row = len(employees) + 3
        ws.row_dimensions[total_row].height = 22
        
        # "TOTAL" label
        total_label_cell = ws.cell(row=total_row, column=1)
        total_label_cell.value = "TOTAL"
        total_label_cell.font = Font(bold=True, size=11, name="Calibri")
        total_label_cell.fill = PatternFill("solid", fgColor="F0F0F0")
        total_label_cell.border = border
        
        # Calculate totals for premium and adjustment columns
        premium_col_idx = PHASE1_FIELDS.index("current_premium") + 1
        adjustment_col_idx = PHASE1_FIELDS.index("adjustment_amount") + 1
        
        # Premium total
        premium_cell = ws.cell(row=total_row, column=premium_col_idx)
        premium_cell.value = f"=SUM({get_column_letter(premium_col_idx)}3:{get_column_letter(premium_col_idx)}{total_row-1})"
        premium_cell.number_format = '$#,##0.00'
        premium_cell.font = Font(bold=True, size=11, name="Calibri")
        premium_cell.fill = PatternFill("solid", fgColor="F0F0F0")
        premium_cell.border = border
        
        # Adjustment total
        adjustment_cell = ws.cell(row=total_row, column=adjustment_col_idx)
        adjustment_cell.value = f"=SUM({get_column_letter(adjustment_col_idx)}3:{get_column_letter(adjustment_col_idx)}{total_row-1})"
        adjustment_cell.number_format = '$#,##0.00'
        adjustment_cell.font = Font(bold=True, size=11, name="Calibri")
        adjustment_cell.fill = PatternFill("solid", fgColor="F0F0F0")
        adjustment_cell.border = border
        
        # Fill empty cells in total row with styling
        for col_idx in range(2, len(PHASE1_FIELDS) + 1):
            if col_idx not in [premium_col_idx, adjustment_col_idx]:
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = PatternFill("solid", fgColor="F0F0F0")
                cell.border = border
    
    # Freeze panes (header rows)
    ws.freeze_panes = "A3"
    
    # Save workbook
    wb.save(output_path)
    logger.info(f"[Phase 1] Excel output saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATION
# ══════════════════════════════════════════════════════════════════════════════

def generate_phase1_output_from_text(
    text: str,
    output_dir: Path,
    base_filename: str
) -> Dict[str, Path]:
    """
    Generate Phase 1 outputs from already-extracted PDF text (faster, no re-extraction).
    Uses intelligent chunking strategy for large documents.
    
    Args:
        text: Pre-extracted PDF text
        output_dir: Directory where outputs should be saved
        base_filename: Base filename for output files (without extension)
        
    Returns:
        Dictionary with paths to generated files
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate output filenames
    json_path = output_dir / f"{base_filename}_phase1.json"
    excel_path = output_dir / f"{base_filename}_phase1.xlsx"
    
    logger.info("="*80)
    logger.info("[Phase 1] Starting Phase 1 Output Generation (from cached text)")
    logger.info(f"[Phase 1] Text length: {len(text):,} characters")
    logger.info(f"[Phase 1] Output Directory: {output_dir}")
    logger.info("="*80)
    
    try:
        # Classify document
        logger.info("[Phase 1] Classifying document type...")
        doc_type = classify(text)
        logger.info(f"[Phase 1] Document classified as: {doc_type}")
        
        # Decide whether to use chunking or single call based on PAGE COUNT
        # Single call is faster for small invoices (1-2 pages)
        # Chunking with overlap is essential for multi-page invoices (3+ pages)
        
        # Count pages in the text
        pages = split_text_by_pages(text)
        page_count = len(pages)
        logger.info(f"[Phase 1] Detected {page_count} pages in document")
        
        if page_count <= 2:
            # Small document (1-2 pages) - use single call (faster, sufficient)
            logger.info(f"[Phase 1] Using single LLM call (document has {page_count} pages)")
            
            prompt = PHASE1_EXTRACTION_PROMPT.format(text=text)
            
            response = openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a precise data extraction assistant."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0
            )
            
            extracted_data = json.loads(response.choices[0].message.content)
        else:
            # Multi-page document (3+ pages) - use chunking strategy
            logger.info(f"[Phase 1] Using chunking strategy (document has {page_count} pages)")
            extracted_data = extract_phase1_data_with_chunking(text, doc_type)
        
        employees = extracted_data.get("employees", [])
        logger.info(f"[Phase 1] Extracted {len(employees)} employee plan lines (duplicates allowed)")
        
        # Generate JSON output
        generate_phase1_json(extracted_data, json_path)
        
        # Generate Excel output
        generate_phase1_excel(extracted_data, excel_path)
        
        logger.info("="*80)
        logger.info("[Phase 1] ✓ Phase 1 Output Generation Complete!")
        logger.info(f"[Phase 1] Generated Files:")
        logger.info(f"[Phase 1]   • JSON:  {json_path.name}")
        logger.info(f"[Phase 1]   • Excel: {excel_path.name}")
        logger.info(f"[Phase 1] Record Count: {len(employees)}")
        logger.info("="*80)
        
        return {
            "json_path": json_path,
            "excel_path": excel_path,
            "record_count": len(employees)
        }
        
    except Exception as e:
        logger.error(f"[Phase 1] ✗ Phase 1 generation failed: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise


def generate_phase1_output(
    pdf_path: Path,
    output_dir: Optional[Path] = None
) -> Dict[str, Path]:
    """
    Generate Phase 1 outputs (JSON + Excel) from a PDF invoice.
    
    Args:
        pdf_path: Path to the PDF invoice file
        output_dir: Directory where outputs should be saved (defaults to OUTPUT_DIR)
        
    Returns:
        Dictionary with paths to generated files:
        {
            "json_path": Path to JSON file,
            "excel_path": Path to Excel file
        }
    """
    pdf_path = Path(pdf_path)
    
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    if output_dir is None:
        output_dir = OUTPUT_DIR
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate output filenames
    base_name = pdf_path.stem
    json_path = output_dir / f"{base_name}_phase1.json"
    excel_path = output_dir / f"{base_name}_phase1.xlsx"
    
    logger.info("="*80)
    logger.info("[Phase 1] Starting Phase 1 Output Generation")
    logger.info(f"[Phase 1] Input PDF: {pdf_path.name}")
    logger.info(f"[Phase 1] Output Directory: {output_dir}")
    logger.info("="*80)
    
    try:
        # Extract data from PDF
        extracted_data = extract_phase1_data(pdf_path)
        
        # Generate JSON output
        generate_phase1_json(extracted_data, json_path)
        
        # Generate Excel output
        generate_phase1_excel(extracted_data, excel_path)
        
        logger.info("="*80)
        logger.info("[Phase 1] ✓ Phase 1 Output Generation Complete!")
        logger.info(f"[Phase 1] Generated Files:")
        logger.info(f"[Phase 1]   • JSON:  {json_path.name}")
        logger.info(f"[Phase 1]   • Excel: {excel_path.name}")
        logger.info(f"[Phase 1] Record Count: {len(extracted_data.get('employees', []))}")
        logger.info("="*80)
        
        return {
            "json_path": json_path,
            "excel_path": excel_path,
            "record_count": len(extracted_data.get("employees", []))
        }
        
    except Exception as e:
        logger.error(f"[Phase 1] ✗ Phase 1 generation failed: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise


# ══════════════════════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    """CLI entry point for Phase 1 output generation."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Generate Phase 1 outputs (pre-deduction member records)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_phase1_output.py invoice.pdf
  python generate_phase1_output.py invoice.pdf ./outputs/
  python generate_phase1_output.py "C:/Documents/Invoice_Jan2026.pdf" "C:/Outputs/"
        """
    )
    
    parser.add_argument(
        "pdf_path",
        help="Path to the PDF invoice file"
    )
    
    parser.add_argument(
        "output_dir",
        nargs="?",
        default=None,
        help="Output directory (optional, defaults to rpve_outputs/)"
    )
    
    args = parser.parse_args()
    
    try:
        result = generate_phase1_output(
            pdf_path=Path(args.pdf_path),
            output_dir=Path(args.output_dir) if args.output_dir else None
        )
        
        print("\n" + "="*80)
        print("SUCCESS! Phase 1 outputs generated:")
        print(f"  JSON:  {result['json_path']}")
        print(f"  Excel: {result['excel_path']}")
        print(f"  Records: {result['record_count']}")
        print("="*80 + "\n")
        
        sys.exit(0)
        
    except Exception as e:
        print("\n" + "="*80)
        print(f"ERROR: {e}")
        print("="*80 + "\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
