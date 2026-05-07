"""
RPVE_standalone.py
==================
Standalone FastAPI server for the RPVE Benefit Invoice Extractor.

RPVE = Resourcing · Prestige · Velocity · Engage

USAGE
-----
1. pip install -r requirements_RPVE.txt
2. Add OPENAI_API_KEY=sk-... to .env
3. python RPVE_standalone.py
4. Open http://localhost:8009

ENDPOINTS
---------
POST /extract          Upload PDF -> JSON + Excel download link
GET  /download/{file}  Download generated Excel
GET  /health           Health check
GET  /                 Serves RPVE_ui.html
"""

import os, re, json, shutil, uuid, time
import pdfplumber
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
POPPLER_PATH = os.getenv("POPPLER_PATH")
if POPPLER_PATH and os.path.exists(POPPLER_PATH):
    if POPPLER_PATH not in os.environ["PATH"]:
        os.environ["PATH"] += os.pathsep + POPPLER_PATH
    print(f"[RPVE] Poppler path added to PATH: {POPPLER_PATH}")

BASE_DIR   = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "rpve_uploads"
OUTPUT_DIR = BASE_DIR / "rpve_outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# ══════════════════════════════════════════════════════════════════════════════
# UNIFIED FIELD SCHEMA
# ══════════════════════════════════════════════════════════════════════════════

# A single, unified schema for all extractions as per user requirements.
# The specialized "engage" prompt is also designed to return these fields.
UNIFIED_FIELDS = [
    "full_name",
    "first_name",
    "middal_name",
    "last_name",
    "coverage",
    "plan_name",
    "plan_type",
    "current_premium",
    "adjustment_amount",
    "birth_date",
    "gender",
    "home_zip_code",
    "billing_period",
]

# The employee fields dictionary now only distinguishes between ADP and generic.
EMPLOYEE_FIELDS = {
    "engage":       UNIFIED_FIELDS,
    "generic":      UNIFIED_FIELDS,
    "datalink_emi": UNIFIED_FIELDS,   # Data Link EMI uses the same unified schema
}

# Simplified summary fields. The generic prompt will attempt to find these.
SUMMARY_FIELDS = {
    "engage":  ["COMPANY_NAME", "INVOICE_NUMBER", "BILLING_DATE", "DUE_DATE", "REFERENCE_NUMBER"],
    "generic": ["COMPANY_NAME", "INVOICE_NUMBER", "BILLING_DATE", "DUE_DATE"],
}

SUB_TYPE_LABELS = {
    "engage":  "Engage (ADP TotalSource)",
    "generic": "Generic Document",
}

HEADER_COLOURS = {
    "engage":  "3C3489",
    "generic": "666666",
}

# ══════════════════════════════════════════════════════════════════════════════
# KEYWORD CLASSIFIER
# ══════════════════════════════════════════════════════════════════════════════

# Simplified to identify ADP ("engage") and ACSA documents.
KEYWORDS = {
    "engage": ["TOTALSOURCE", "TOTALSOURCE BENEFITS INVOICE", "TOTALSOURCE® BENEFITS INVOICE", "NCT3-EPO", "ADP", "ADP, INC", 
               "ASSOCIATION OF COMMUNITY SERVICE", "ACSA", "ACSA GROUP INSURANCE", "HEALTHNET", "HEALTH NET"],
    # Data Link EMI carrier
    "datalink_emi": ["DATA LINK EMI", "DATALINK EMI", "DATALINKEMI"],
}

# ══════════════════════════════════════════════════════════════════════════════
# LLM PROMPTS - output keys match EMPLOYEE_FIELDS exactly
# ══════════════════════════════════════════════════════════════════════════════

PROMPTS = {

    "deduction_roster": """
You are extracting data from a BENEFIT DEDUCTION ROSTER.

Extract a SUMMARY and EMPLOYEES array.

SUMMARY: company_name, invoice_number, billing_date, total_amount_due

EMPLOYEE RECORDS:
For each employee plan line, you MUST extract:
  first_name: member first name
  last_name: member last name  
  coverage: coverage level (e.g. EE, ES, FAM, EC)
  plan_name: the "Plan Description" column value
  current_premium: YOU MUST TAKE THE VALUE FROM THE "Monthly Premium Total" COLUMN.
  
🔹 CRITICAL RULE:
There are two "Total" columns: "Monthly Premium Total" and "Pay Period Amount Total".
YOU MUST USE THE "Monthly Premium Total".
DO NOT USE THE "Pay Period Amount Total".

Example:
If a row shows "525.00" under Monthly Premium Total and "262.50" under Pay Period Amount Total, current_premium MUST BE "525.00".

Use "" for missing values. Return ONLY valid JSON.

{
  "summary": {"company_name":"","invoice_number":"","billing_date":"","total_amount_due":""},
  "employees": [{"first_name":"","last_name":"","coverage":"","plan_name":"","current_premium":""}]
}

PDF TEXT: {text}
""",

    "engage": """
You are extracting data from a group health insurance invoice (ADP TotalSource, ACSA, or similar).

Extract a SUMMARY and EMPLOYEES array.

SUMMARY: company_name, invoice_number, billing_date, due_date, reference_number, total_amount_due

FOR SUMMARY-ONLY INVOICES (like ACSA):
If the invoice shows only summary financial data (no individual employee roster), create ONE consolidated record:
  first_name: "SUMMARY"
  last_name: "TOTAL" 
  coverage: "EMPLOYER"
  plan_name: "HEALTH PLAN"
  coverage_option: billing period (e.g. "03/01/2026 through 03/31/2026")
  current_premium: total amount due

FOR EMPLOYEE ROSTER INVOICES (like ADP):
Create individual records for each employee plan line:
  first_name: member first name
  last_name: member last name  
  coverage: EXACT coverage tier/code (e.g. Employee, Family, EE+1, E, ES, ESC, EC, E1D, etc.)
  plan_name: insurance category/type (e.g. Medical, Dental, Vision)
  coverage_option: specific insurance product name
  current_premium: dollar amount for that plan line

EXTRACTION RULES:
1. DETECT FORMAT: Look for individual employee names vs summary-only financial data
2. SUMMARY-ONLY: Extract employer name, billing period, and total financial summary
3. ROSTER FORMAT: Extract each employee plan line with individual costs
4. FINANCIAL DATA: Prior balance, adjustments, current cost, admin fees, total due
5. DATES: Billing date, due date, billing period
6. BENEFIT DEDUCTION ROSTER: If the document is titled "BENEFIT DEDUCTION ROSTER", you MUST extract the value from the "Monthly Premium" -> "Total" column as the `current_premium`. STRICTLY FORBIDDEN: Do NOT use the "Pay Period Amount" values for the premium.

Use "" for missing values. Return ONLY valid JSON.

{
  "summary": {"company_name":"","invoice_number":"","billing_date":"","due_date":"","reference_number":"","total_amount_due":""},
  "employees": [{"first_name":"","last_name":"","coverage":"","plan_name":"","coverage_option":"","current_premium":""}]
}

PDF TEXT: {text}
""",

}

# ══════════════════════════════════════════════════════════════════════════════
# FALLBACK EXTRACTION FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def extract_engage_fallback(text: str) -> dict:
    """
    Fallback regex-based extractor for engage subtype when LLM extraction fails.
    Handles both ADP roster format and ACSA summary format.
    """
    import re
    
    # Initialize result structure
    result = {
        "summary": {
            "company_name": "",
            "invoice_number": "",
            "billing_date": "",
            "due_date": "",
            "reference_number": "",
            "total_amount_due": ""
        },
        "employees": []
    }
    
    try:
        text_upper = text.upper()
        
        # Check if this is ACSA format
        if "ASSOCIATION OF COMMUNITY SERVICE" in text_upper or "ACSA" in text_upper:
            return extract_acsa_summary_fallback(text)
        
        # For ADP/TotalSource format - extract basic summary info
        # Extract company name (look for common patterns)
        company_patterns = [
            r'Company\s*:\s*(.+?)(?:\n|\r)',
            r'Employer\s*Name\s*:\s*(.+?)(?:\n|\r)',
            r'Client\s*:\s*(.+?)(?:\n|\r)'
        ]
        
        for pattern in company_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                result["summary"]["company_name"] = match.group(1).strip()
                break
        
        # Extract billing date
        date_match = re.search(r'Billing\s*Date\s*:\s*(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if date_match:
            result["summary"]["billing_date"] = date_match.group(1)
        
        # Extract total amount (look for final totals)
        total_patterns = [
            r'Total\s*Amount\s*Due\s*[:\s]*(\d{1,3}(?:,\d{3})*\.\d{2})',
            r'Grand\s*Total\s*[:\s]*(\d{1,3}(?:,\d{3})*\.\d{2})',
            r'Amount\s*Due\s*[:\s]*(\d{1,3}(?:,\d{3})*\.\d{2})'
        ]
        
        for pattern in total_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                result["summary"]["total_amount_due"] = match.group(1)
                break
        
        print(f"[RPVE] Engage fallback extraction: Found {len(result['employees'])} employee records")
        return result
        
    except Exception as e:
        print(f"[RPVE] Engage fallback extraction failed: {e}")
        return {"summary": {}, "employees": []}


def extract_acsa_summary_fallback(text: str) -> dict:
    """
    Fallback regex-based extractor for ACSA summary format when LLM extraction fails.
    """
    import re
    
    result = {
        "summary": {
            "company_name": "",
            "invoice_number": "",
            "billing_date": "",
            "due_date": "",
            "reference_number": "",
            "total_amount_due": ""
        },
        "employees": []
    }
    
    try:
        # Extract company name for ACSA
        company_match = re.search(r'Association\s+of\s+Community\s+Service\s+Agencies?[\s\-]*(.+?)(?:\n|\r)', text, re.IGNORECASE)
        if company_match:
            result["summary"]["company_name"] = company_match.group(1).strip()
        
        # Extract billing period
        period_match = re.search(r'Billing\s+Period\s*[:\-\s]*(\d{2}/\d{2}/\d{4})[\s\-]*(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if period_match:
            result["summary"]["billing_date"] = period_match.group(1)
            result["summary"]["due_date"] = period_match.group(2)
        
        # Extract total amount
        total_match = re.search(r'Total\s+Premium\s*[:\-\s]*\$(\d{1,3}(?:,\d{3})*\.\d{2})', text, re.IGNORECASE)
        if total_match:
            result["summary"]["total_amount_due"] = total_match.group(1)
        
        print(f"[RPVE] ACSA fallback extraction: Found {len(result['employees'])} employee records")
        return result
        
    except Exception as e:
        print(f"[RPVE] ACSA fallback extraction failed: {e}")
        return {"summary": {}, "employees": []}


# ══════════════════════════════════════════════════════════════════════════════
# CORE FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def extract_text(pdf_path: Path, max_pages: int = 1000) -> str:
    """
    Robustly extracts text from a PDF by iterating through pages and using 
    multiple engines (pdfplumber, fitz, and OCR) to ensure full capture.
    Handles rotated and reversed text mapping issues.
    """
    text = ""
    print(f"[RPVE] Extracting text from {pdf_path.name}...")

    # Keywords we EXPECT to find in a valid RPVE document page
    VALID_KEYWORDS = [
        "TOTALSOURCE", "PAYCHEX", "AETNA", "KAISER", "UNITEDHEALTHCARE", "INVOICE", "BILLING", 
        "PREMIUM", "AMOUNT DUE", "PAGE", "EMPLOYEE", "MEMBERS", "CURRENT DETAIL", 
        "RETRO DETAIL", "ADJUSTMENT DETAIL", "MEDICA", "ADP", "BLUE CROSS", "CIGNA", "GUARDIAN"
    ]

    try:
        import pdfplumber
        import fitz
        import pytesseract
        from PIL import Image
        
        with pdfplumber.open(str(pdf_path)) as pdf:
            with fitz.open(pdf_path) as doc:
                pages_to_extract = min(max_pages, len(pdf.pages))
                for i in range(pages_to_extract):
                    page_text = ""
                    
                    # 1. Try pdfplumber (best for layout preservation)
                    try:
                        plumber_page = pdf.pages[i]
                        p_text = plumber_page.extract_text(layout=True) or ""
                        if len(p_text.strip()) > 100 and any(kw in p_text.upper() for kw in VALID_KEYWORDS):
                            page_text = p_text
                    except Exception as e:
                        print(f"  [PAGE {i+1}] pdfplumber error: {e}")

                    # 2. Fallback to fitz (PyMuPDF) if pdfplumber is empty or fails keywords
                    if not page_text.strip():
                        try:
                            fitz_page = doc[i]
                            f_text = fitz_page.get_text() or ""
                            if len(f_text.strip()) > 50 and any(kw in f_text.upper() for kw in VALID_KEYWORDS):
                                page_text = f_text
                        except Exception as e:
                            print(f"  [PAGE {i+1}] fitz failed: {e}")

                    # 3. Last Resort: Robust High-Accuracy OCR (handles scanned/rotated text)
                    if not page_text.strip():
                        print(f"  [PAGE {i+1}] Running Robust OCR Fallback...")
                        try:
                            # Render at 3x zoom (216 dpi) for high accuracy
                            pix = doc[i].get_pixmap(matrix=fitz.Matrix(3, 3))
                            img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
                            
                            # A. Try Standard OSD detection first
                            try:
                                osd = pytesseract.image_to_osd(img)
                                rotation = re.search(r'Rotate: (\d+)', osd)
                                if rotation:
                                    angle = int(rotation.group(1))
                                    if angle != 0:
                                        print(f"    [OSD] Correcting {angle}° rotation...")
                                        img = img.rotate(-angle, expand=True)
                            except:
                                pass

                            # B. Try OCR at current orientation
                            page_text = pytesseract.image_to_string(img, config='--psm 6')

                            # C. Brute-Force Rotation Fallback (if keywords missing)
                            if not any(kw in page_text.upper() for kw in VALID_KEYWORDS):
                                print(f"    [PAGE {i+1}] Keywords not found. Retrying 90/180/270 degree rotations...")
                                ori_img = img.copy()
                                for rot in [90, 180, 270]:
                                    img_rot = ori_img.rotate(rot, expand=True)
                                    test_text = pytesseract.image_to_string(img_rot, config='--psm 6')
                                    if any(kw in test_text.upper() for kw in VALID_KEYWORDS):
                                        page_text = test_text
                                        print(f"    [PAGE {i+1}] Success at {rot}° rotation.")
                                        break
                            
                            # D. Final check: if still empty, use psm 3 (standard)
                            if not page_text.strip():
                                page_text = pytesseract.image_to_string(img, config='--psm 3')

                        except Exception as e:
                            print(f"  [PAGE {i+1}] OCR failed: {e}")

                    text += page_text + "\n"
    except Exception as e:
        print(f"[RPVE] Global extraction error: {e}")

    extracted_upper = text.upper()
    text_quality_score = assess_text_quality(text)
    
    # ── CARRIER-SPECIFIC EXTRACTION UPGRADE ───────────────────────────
    # If this is a wide-layout carrier like BlueCross or IBX, standard 
    # vertical extraction often shreds rows. We force the high-accuracy 
    # layout-preserving OCR fallback for these specifically.
    WIDE_KEYWORDS = ["BLUE CROSS", "INDEPENDENCE", "BCBS", "CAPITAL BLUE", "IBX", "BLUE SHIELD"]
    is_wide_carrier = any(kw in extracted_upper for kw in WIDE_KEYWORDS)

    print(f"[RPVE] Extraction check: quality={text_quality_score:.2f}, is_wide={is_wide_carrier}")

    if (not text.strip() 
        or not any(kw in extracted_upper for kw in VALID_KEYWORDS)
        or text_quality_score < 0.55
        or is_wide_carrier):
        
        reason = ""
        if is_wide_carrier:
            reason = "wide carrier layout detected (forcing high-accuracy pass)"
        elif not text.strip():
            reason = "text is empty"
        elif not any(kw in extracted_upper for kw in VALID_KEYWORDS):
            reason = f"keywords not found (quality: {text_quality_score:.2f})"
        else:
            reason = f"text quality too low ({text_quality_score:.2f} < 0.55 threshold)"
        
        print(f"[RPVE] Triggering high-accuracy fallback: {reason}")
        rostaing_text = extract_text_with_rostaing(pdf_path)
        
        # ── INTELLIGENT FALLBACK VALIDATION ───────────────────────────
        # Only use Rostaing text if it's actually "better" or at least 
        # comparable in volume. If Rostaing fails on rotated pages 
        # (producing less text than standard), we stick to standard.
        if rostaing_text and rostaing_text.strip():
            if len(rostaing_text) >= (len(text) * 0.8):
                print(f"[RPVE] Using Rostaing OCR result ({len(rostaing_text)} chars).")
                return rostaing_text
            else:
                print(f"[RPVE] Rostaing result suspiciously short ({len(rostaing_text)} vs {len(text)}). Sticking to standard extraction.")

    return text


def assess_text_quality(text: str) -> float:
    """
    Score text quality (0.0 to 1.0) to detect OCR corruption.
    Low score = high corruption (garbled chars, broken tables, fragmentation).
    """
    if not text or len(text) < 50:
        return 0.0
    
    lines = text.split('\n')
    line_count = len(lines)
    
    # 1. Check for garbled/non-ASCII characters (corruption indicator)
    garbled_count = 0
    for char in text:
        if ord(char) > 127 and char not in 'àáâãäåèéêëìíîïòóôõöùúûüýÿªºñ—–':
            garbled_count += 1
    
    garbled_ratio = garbled_count / len(text) if len(text) > 0 else 0
    
    # 2. Check for repeated fragmented lines (broken table markers)
    fragment_pattern = r'\s*\|\s*|\s+\[.{1,3}\]\s+'
    fragment_count = len(re.findall(fragment_pattern, text))
    fragment_ratio = fragment_count / max(line_count, 1)
    
    # 3. Check for very short lines (fragmentation sign)
    short_lines = sum(1 for line in lines if len(line.strip()) < 5 and line.strip())
    short_line_ratio = short_lines / max(line_count, 1) if line_count > 0 else 0
    
    # 4. Check for repeated consecutive lines (OCR duplication artifact)
    duplicate_lines = 0
    for i in range(1, len(lines)):
        if lines[i].strip() and lines[i-1].strip() and lines[i].strip() == lines[i-1].strip():
            duplicate_lines += 1
    
    duplicate_ratio = duplicate_lines / max(line_count - 1, 1)
    
    # Weighted quality score
    quality = 1.0
    quality -= min(0.3, garbled_ratio * 3)        # Up to 30% penalty for non-ASCII
    quality -= min(0.4, fragment_ratio * 3)       # Up to 40% penalty for fragments (increased weight)
    quality -= min(0.25, short_line_ratio * 2)    # Up to 25% penalty for short lines
    quality -= min(0.25, duplicate_ratio * 2)     # Up to 25% penalty for duplicates
    
    quality = max(0.0, min(1.0, quality))
    print(f"[RPVE] Text quality assessment: {quality:.2f} (garbled: {garbled_ratio:.2%}, fragments: {fragment_ratio:.2%}, short_lines: {short_line_ratio:.2%}, duplicates: {duplicate_ratio:.2%})")
    
    return quality


def extract_text_with_rostaing(pdf_path: Path) -> str:
    """Fallback PDF text extraction using rostaing-ocr when the standard path is noisy."""
    try:
        from schema_ocr import SchemaOCRExtractor
    except Exception as e:
        print(f"[RPVE] Could not import schema_ocr for rostaing fallback: {e}")
        return ""

    if shutil.which("tesseract") is None:
        print("[RPVE] Tesseract not found in PATH. rostaing-ocr may still work, but OCR accuracy could be reduced.")

    try:
        extractor = SchemaOCRExtractor(pdf_path)
        text = extractor.extract_layout_text(save_debug_output=True)
        if text and text.strip():
            print(f"[RPVE] Rostaing OCR fallback produced {len(text.splitlines())} lines of text.")
            return text
        print("[RPVE] Rostaing OCR fallback returned empty text.")
    except Exception as e:
        print(f"[RPVE] Rostaing OCR fallback failed: {e}")

    return ""


def clean_invoice_text(text: str) -> str:
    """
    Remove common noise like page footers, headers, and copyright notices.
    Also handles 'orphaned' Total rows that can confuse the LLM at page boundaries.
    """
    if not text: return ""
    lines = text.split('\n')
    cleaned_lines = []
    
    # Regex to detect page footers, headers, copyright notices, and orphaned Totals
    # We remove "Total" rows that have NO provider/plan data if they appear near headers
    header_footer_pattern = re.compile(r"""
        ^\s*page\s+\d+\s+of\s+\d+\s*$|
        copyright\s+©\s+.*adp,\s+inc|
        ^\s*Name\s+Provider\s+Plan\s+Coverage\s+Type\s+Month\s+Cost\s*$|
        ^\s*Total\s+\$[\d,]+\.\d{2}\s*$
    """, re.IGNORECASE | re.VERBOSE)

    for line in lines:
        # If the line doesn't match the noise pattern, keep it
        if not header_footer_pattern.search(line):
            cleaned_lines.append(line)
            
    return '\n'.join(cleaned_lines)

def group_indented_lines(text: str) -> str:
    """
    DYNAMICAL SOLUTION: 
    Detects lines starting with significant whitespace OR 'EE ID:' and 
    stitches them to the preceding 'Name' row. 
    """
    if not text: return ""
    lines = text.split('\n')
    output = []
    current_block = ""

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
            
        # DYNAMICAL RULES:
        # 1. Any line starting with "EE ID:" belongs to the employee above.
        # 2. Any line starting with 10+ spaces (indentation > name) belongs to employee above.
        # 3. Any line starting with "Total" (at sub-indentation) belongs to employee above.
        
        is_sub_row = (
            stripped.startswith("EE ID:") or 
            line.startswith("          ") or 
            (line.startswith("     ") and stripped.startswith("Total"))
        )
        
        if is_sub_row and current_block:
            current_block += " [SUB-ROW] " + stripped
        else:
            if current_block:
                output.append(current_block)
            current_block = line
            
    if current_block:
        output.append(current_block)
        
    return '\n'.join(output)


def classify(text: str) -> str:
    """Classifies the document for extraction type."""
    t = text.upper()
    
    # Priority 0: Check for Benefit Deduction Roster
    if "BENEFIT DEDUCTION ROSTER" in t:
        print("[RPVE] Detected Benefit Deduction Roster format.")
        return "deduction_roster"

    # Priority 1: Check for ADP/TotalSource (original engage)
    if "TOTALSOURCE" in t or "ADP" in t:
        print("[RPVE] Detected ADP TotalSource format.")
        return "engage"
    
    # Priority 2: Check for ACSA/Community Service Agencies (treat as engage)
    if "ASSOCIATION OF COMMUNITY SERVICE" in t or "ACSA" in t:
        print("[RPVE] Detected ACSA Health Plan format, treating as engage type.")
        return "engage"
    
    # Priority 3: Check for other specific formats
    for sub_type, kwlist in KEYWORDS.items():
        if any(kw in t for kw in kwlist):
            print(f"[RPVE] Classified as {sub_type.upper()} format.")
            return sub_type
    
    # Check for summary-only invoice characteristics (treat as engage)
    if "EMPLOYER NAME" in t and "BILLING DATE" in t and "TOTAL AMOUNT" in t:
        print("[RPVE] Detected summary-only invoice format, treating as engage type.")
        return "engage"
    
    print("[RPVE] No specific keywords matched. Using GENERIC extractor.")
    return "generic"


def clean_invoice_text(text: str) -> str:
    """
    Cleans the extracted invoice text by removing headers, footers, and other noise
    that can disrupt the LLM's parsing of continuous employee data.
    """
    lines = text.split('\n')
    cleaned_lines = []
    
    # Regex to detect page footers, headers, and copyright notices
    header_footer_pattern = re.compile(r"""
        ^\s*page\s+\d+\s+of\s+\d+\s*$|
        copyright\s+©\s+.*adp,\s+inc|
        ^\s*Name\s+Provider\s+Plan\s+Coverage\s+Type\s+Month\s+Cost\s*$
    """, re.IGNORECASE | re.VERBOSE)

    for line in lines:
        # If the line doesn't match the pattern, keep it
        if not header_footer_pattern.search(line):
            cleaned_lines.append(line)
            
    # Rejoin the lines
    cleaned_text = '\n'.join(cleaned_lines)
    
    return cleaned_text


def extract_with_llm(sub_type: str, text: str, ev_mode: bool = False) -> dict:
    """
    Calls the LLM to extract structured summary and employee data.
    Uses carrier-specific prompts if available, otherwise falls back to a standard prompt.
    """
    # Clean the text to handle multi-page table fragmentation
    text = clean_invoice_text(text)
    
    # DYNAMICAL STITCHING: Group indented rows before sending to LLM
    if sub_type in ["engage", "prestige", "velocity"]:
        text = group_indented_lines(text)

    # 1. Determine which prompt to use
    prompt_template = PROMPTS.get(sub_type)
    
    if not prompt_template:
        # Fallback to the Standard / Generic Prompt
        prompt_template = """
You are a data extraction engine processing a group insurance invoice.

🔹 CAPTURE ALL MEMBERS & ADJUSTMENTS (CRITICAL)
This invoice may list members in a ROSTER format where each row is a separate individual (Subscriber, Spouse, Dependent).
You MUST extract EVERY person listed in ANY section:
  - CURRENT DETAIL section
  - RETRO DETAIL section
  - ADJUSTMENT DETAIL section
  - Any other detail section in the document

🔹 NEGATIVE VALUES (CRITICAL):
  - If a value is negative (e.g. $-100.00), you MUST preserve the minus sign in the current_premium or adjustment_amount field.

🔹 ADP FORMAT SPECIFIC RULES (APPLY ONLY IF "TOTALSOURCE", "ADP", OR "NCT3-EPO" IS PRESENT)
If the document is EXPLICITLY identified as an ADP invoice (e.g. ADP TotalSource format), you MUST apply these strict rules. If it is NOT an ADP file, ignore these specific constraints and extract EVERY record regardless of amount:

1. Plan Name Extraction (CRITICAL for ADP):
Extract ONLY the exact, valid ADP plan name.
Do NOT extract random text near plan sections, headers, footers, or unrelated labels.
✅ Plan name must belong to a defined benefits section, be consistent across employee entries, and appear as a clear plan title.

🔹 DO NOT ABBREVIATE OR TRUNCATE PLAN NAMES (CRITICAL):
The plan name MUST match the FULL string found in the "Plan" column of the PDF.

🔹 ADP LAYOUT STRUCTURE:
Data is typically organized as a Header Row (Name + First Plan) followed by Indented Rows (Additional Plans). 
The pre-processor has marked these as '[SUB-ROW]'. You MUST split these [SUB-ROW] markers into separate individual plan records for the same employee.

🔹 PAGE BREAK CONTINUATIONS (CRITICAL):
If you see an employee's data interrupted or continued, associate all subsequent plan lines with that specific employee until a new Name is encountered. 
Do NOT create unnamed records for orphaned plan lines.

🔹 NAME FORMATTING (CRITICAL):
Names are often printed as "LastName, FirstName" or "LastName, FirstName Middle" (e.g. "Smith, John Adam"). Properly identify and split the `last_name` and `first_name` without inverting them.

🔹 OUTPUT FIELDS (13 fields per person)
- full_name
- first_name
- middal_name (Middle Name)
- last_name
- coverage (e.g. ES / EC / FAM / EE / SP / CH)
- plan_name (FULL plan/product description — do NOT truncate). IMPORTANT: Plan names often wrap across multiple lines. You MUST concatenate these into a single string (e.g., "Non-Contributory 25K Flat Basic Life EE Only").
- plan_type (insurance category: Medical, Dental, Vision, etc.)
- current_premium: The individual plan line cost for that specific plan row. (**CRITICAL FOR UHC/UHC NA**: Use the "Totals -> Total" column for single-line rows, but use "Charge Amount" if multiple distinct plan rows exist for the same member.)
- adjustment_amount: Any adjustment amount listed.
- birth_date
- gender (M / F — infer if not present)
- home_zip_code
- billing_period: The start and end date of the billing cycle for the line item (e.g., "01/01/2024 - 01/31/2024").

🔹 MAXIMUM RECALL (CRITICAL):
- **Completeness is the HIGHEST priority.** You MUST extract EVERY individual enrollment row found in the document.
- Do NOT skip any rows. Even if a row has partial data, extract what is available.
- If the document contains a roster (like IBX or BCBS), expect dozens of members. You must continue until the very end of the list.

🔹 KEY RULES:
- One row per individual member.
- Strictly adhere to JSON format.
- Do not hallucinate, but do not omit valid rows.

🔹 INSPERITY / MANIFEST MEDEX:
- If column headers include "Coverage Type" and "Coverage Option", map "Coverage Type" -> `plan_type` and "Coverage Option" -> `plan_name`. Do not mix them up.

🔹 WARWICK / DEDUCTION REGISTER:
- If column headers include "Ded Code" or "Benefit Plan", map "Benefit Plan" to `plan_name` and "Ded Code" to `plan_type`.

🔹 KARPEN_STEEL_PRODUCTS:
- For this carrier, you MUST map the value from the "Total Premium" column to `adjustment_amount`.

🔹"DATA LINK EMI", "DATALINK EMI", "DATALINKEMI" :
- **CRITICAL: EXTRACT ONLY THE "Medical" COLUMN VALUE. ALL OTHER COLUMNS ARE FORBIDDEN.**
- FORBIDDEN columns (never extract these): "Total Due", "Dental", "Vision", "Garner HRA". Ignore them completely.
- For REGULAR rows (current billing):
    - `current_premium` = value from "Medical" column ONLY (e.g. $334.78, $0.00).
    - `adjustment_amount` = null.
- For RETRO ACTIVE ADJUSTMENT rows (negative values / retroactive section):
    - `current_premium` = null.
    - `adjustment_amount` = value from "Medical" column ONLY (e.g. $-334.78).
- If a member has Medical=$0.00, then current_premium MUST be "$0.00" — do NOT substitute Dental or Vision or Total Due.
- Example: Medical=$0.00, Dental=$58.10, Vision=$0.00, Total Due=$58.10 → current_premium="$0.00" ✅ NOT "$58.10" ❌
- Example: Medical=$334.78, Dental=$27.90, Vision=$7.70, Total Due=$412.38 → current_premium="$334.78" ✅ NOT "$412.38" ❌
- **If you return Total Due, Dental, Vision, or Garner HRA in any field, you have failed the task.**

🔹 BLUECROSS BLUE SHIELD (BCBS) SPECIAL RULE (CRITICAL):
- **ALWAYS extract the value from the "Total Premium" column as the `current_premium`.**
- **STRICTLY FORBIDDEN:** Do NOT use the "Employee Medical", "Dependent(s) Medical", or "Total Medical" columns for the premium value on these forms. 
- You MUST ignore the internal sub-columns and jump to the far-right "Total Premium" column for every member.

🔹 COVERAGE FALLBACK (e.g. BLUECROSS):
- If the document lacks an explicit 'Coverage' column or it is blank, YOU MUST INFER the coverage tier from the relationship or enrollee type (e.g. 'EE', 'Subscriber' -> EE, 'SP', 'Spouse' -> ES).

🔹# CORRECTED UHC PREMIUM SELECTION RULE

# CASE 1: Employee has PAIRED lines (Max Claims Liability + Admin/Excess Loss 
# under the SAME plan group/policy)
#   → current_premium = the shared "Total" column value (e.g., $742.11)
#   → Apply the SAME Total to BOTH rows for that employee
#   → Do NOT use individual Charge Amounts ($399.09 / $343.02)

# CASE 2: Employee has a SINGLE line (no paired Admin/Excess Loss line)
#   → current_premium = the "Total" column value (same as before)

# CASE 3: Adjustment rows (negative values like -$742.11 for TRM status)
#   → Map to adjustment_amount, NOT current_premium
#   → current_premium = NULL for terminated rows

# FORBIDDEN: Never use raw Charge Amount as current_premium
#            when a Total column value is present for that employee group
🔹 Coverage Recovery  : Map coverage type codes using the following legend for ALL UHC documents:
    - `E` or "Employee Only" → **EE**
    - `ES` or "Employee and Spouse" → **ES**
    - `ESC` or "Employee and Family" → **FAM**
    - `EC` or "Employee and Child(ren)" → **EC**
    - `E1D` or "Employee and One Dependent" → **EC**
    - `E2D` or "Employee and Two Dependents" → **EC**
    - `E3D` or "Employee and Three Dependents" → **EC**
    - `E4D` or "Employee and Four Dependents" → **EC**
    - `E5D` or "Employee & One or More Dependent" → **EC**
    - `E6D` or "Employee & Two or More Dependents" → **EC**
    - `E7D` or "Employee & Three or More Dependents" → **EC**
    - `E8D` or "Employee & Four or More Dependents" → **EC**
    - `E9D` or "Employee & Five or More Dependents" → **EC**
    - Single-letter codes only (when alone): `E` → **EE**, `S` → **ES**, `F` → **FAM**, `C` → **EC**, `E E` → **EE**

{
  "summary": {"company_name": "", "total_amount_due": ""},
  "employees": [{"full_name": null, "first_name": null, "middal_name": null, "last_name": null, "coverage": null, "plan_name": null, "plan_type": null, "current_premium": null, "adjustment_amount": null, "birth_date": null, "gender": null, "home_zip_code": null, "billing_period": null}]
}

🔹 SPECIAL CASE - PAYROLL / DEDUCTION REGISTERS (e.g. WARWICK / BENEFIT DEDUCTION ROSTER):
- If column headers include "Pay Date", "Deduction Date", or "Check Date", you MUST extract this into the `billing_period` field for EVERY row.
- **BENEFIT DEDUCTION ROSTER SPECIFIC (CRITICAL):** If the document is titled "BENEFIT DEDUCTION ROSTER", you MUST extract the value from the "Monthly Premium" -> "Total" column as the `current_premium`. 
- **STRICTLY FORBIDDEN:** Do NOT use the "Pay Period Amount" column values for the premium. Always use the monthly total.

PDF TEXT: {text}
"""

    lines = text.split('\n')
    chunks = []
    current_chunk = []
    current_len = 0

    # Use 7,500 char chunk size to stay within GPT-4o output token limits (4k tokens).
    # A 15,000-40,000 char chunk could contain ~150-350 rows, which exceeds output limits.
    CHUNK_MAX = 40000
    OVERLAP   = 4000

    for line in lines:
        if current_len + len(line) > CHUNK_MAX and current_chunk:
            chunks.append('\n'.join(current_chunk))
            # Carry the last OVERLAP chars of this chunk into the next one
            overlap_text = '\n'.join(current_chunk)[-OVERLAP:]
            current_chunk = [overlap_text]
            current_len   = len(overlap_text)
        current_chunk.append(line)
        current_len += len(line) + 1
    if current_chunk:
        chunks.append('\n'.join(current_chunk))

    all_employees = []
    final_summary = {}

    print(f"[RPVE] LLM extraction ({sub_type}) split into {len(chunks)} chunks...")
    for i, chunk in enumerate(chunks):
        # Use simple string replacement for the placeholder
        prompt = prompt_template.replace("{text}", chunk)
        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a precise insurance billing data extraction assistant. Return valid JSON only. No markdown. No extra text."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0,
                response_format={"type": "json_object"},
            )
            raw = response.choices[0].message.content
            try:
                # Try multiple JSON parsing strategies
                data = None
                
                # Strategy 1: Direct parsing
                try:
                    data = json.loads(raw)
                except json.JSONDecodeError:
                    pass
                
                # Strategy 2: Remove markdown code blocks
                if not data:
                    try:
                        cleaned = re.sub(r"```json|```", "", raw).strip()
                        data = json.loads(cleaned)
                    except json.JSONDecodeError:
                        pass
                
                # Strategy 3: Extract JSON from text
                if not data:
                    try:
                        # Look for JSON object pattern
                        json_match = re.search(r'\{[\s\S]*\}', raw)
                        if json_match:
                            data = json.loads(json_match.group())
                    except json.JSONDecodeError:
                        pass
                
                # Strategy 4: Try to fix common JSON issues
                if not data:
                    try:
                        # Fix unterminated strings and missing quotes
                        fixed = re.sub(r'(?<!\\)"(?!\\)', '"', raw)  # Fix quote escaping
                        fixed = re.sub(r'\n\s*"', '"', fixed)     # Remove newlines before quotes
                        fixed = re.sub(r'"\s*\n', '"', fixed)     # Remove newlines after quotes
                        data = json.loads(fixed)
                    except json.JSONDecodeError:
                        pass
                
                if data:
                    if not final_summary and data.get("summary"):
                        # Always grab summary from the first successful chunk
                        final_summary = data.get("summary")
                        
                    emps = data.get("employees", [])
                    all_employees.extend(emps)
                    print(f"  [RPVE] Chunk {i+1}/{len(chunks)} processed -> found {len(emps)} records")
                else:
                    print(f"  [RPVE] Chunk {i+1}/{len(chunks)} failed: Could not parse JSON from response")
                    
            except Exception as parse_error:
                print(f"  [RPVE] Chunk {i+1}/{len(chunks)} failed: {parse_error}")
        except Exception as e:
            print(f"  [RPVE] Chunk {i+1}/{len(chunks)} failed: {e}")

    result = {
        "summary": final_summary,
        "employees": all_employees
    }
    
    # Special fallback for engage type when LLM extraction fails or returns empty
    if sub_type == "engage" and (not all_employees or len(all_employees) == 0):
        print(f"[RPVE] LLM extraction returned empty for engage, trying fallback extractors...")
        
        # Check if this is an ACSA-style summary invoice
        if "ASSOCIATION OF COMMUNITY SERVICE" in text.upper() or "ACSA" in text.upper():
            print(f"[RPVE] Detected ACSA format, using summary fallback...")
            fallback_result = extract_acsa_summary_fallback(text)
            if fallback_result and fallback_result.get("employees"):
                print(f"[RPVE] ACSA fallback successful, returning {len(fallback_result['employees'])} records")
                return fallback_result
        
        # Try generic regex extraction for other engage formats
        print(f"[RPVE] Trying generic engage fallback...")
        generic_fallback = extract_engage_fallback(text)
        if generic_fallback and generic_fallback.get("employees"):
            print(f"[RPVE] Generic engage fallback successful, returning {len(generic_fallback['employees'])} records")
            return generic_fallback
    
    return result


def deduplicate_employees(employees: list[dict]) -> list[dict]:
    """
    Removes duplicate employee records from a list.

    This function identifies duplicates based on a combination of the employee's
    name and their plan name. It prioritizes a full name field but will fall
    back to combining first and last names. It preserves the first occurrence
    of each unique record and discards subsequent duplicates.

    Args:
        employees: A list of employee data dictionaries, where each dictionary
                   represents an extracted row from the invoice.

    Returns:
        A new list of employee data dictionaries with duplicates removed.
    """
    original_count = len(employees)
    seen = set()
    deduplicated_list = []
    for employee in employees:
        # To identify a unique record, we use a combination of the employee's
        # name and their plan name. The generic extractor returns lowercase keys.
        plan_name = (employee.get("plan_name") or "").strip()
        full_name = (employee.get("full_name") or "").strip()

        if not full_name:
            first_name = (employee.get("first_name") or "").strip()
            last_name = (employee.get("last_name") or "").strip()
            if first_name and last_name:
                full_name = f"{first_name} {last_name}"

        # We only consider records for deduplication if they have both a name
        # and a plan. If either is missing, we keep the record to avoid data loss.
        if not full_name or not plan_name:
            deduplicated_list.append(employee)
            continue

        premium = str(employee.get("current_premium") or "").strip()
        adjustment = str(employee.get("adjustment_amount") or "").strip()
        billing_period = str(employee.get("billing_period") or "").strip()
        unique_key = (full_name.upper(), plan_name.upper(), premium, adjustment, billing_period)
        if unique_key not in seen:
            seen.add(unique_key)
            deduplicated_list.append(employee)

    deduplicated_count = len(deduplicated_list)
    removed_count = original_count - deduplicated_count
    if removed_count > 0:
        print(f"[RPVE] Deduplication: {original_count} rows -> {deduplicated_count} rows ({removed_count} duplicates removed)")

    return deduplicated_list

def build_excel(data: dict, sub_type: str, stem: str, active_employee_fields: list[str] | None = None, out_dir: Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out = out_dir if out_dir is not None else OUTPUT_DIR
    wb        = Workbook()
    hex_col   = HEADER_COLOURS.get(sub_type, "1A1A2E")
    hdr_fill  = PatternFill("solid", fgColor=hex_col)
    hdr_font  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="DDDDDD")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    da        = Alignment(vertical="center")
    tf        = Font(bold=True, size=11, name="Calibri")
    tfill     = PatternFill("solid", fgColor="F0F0F0")

    summary   = data.get("summary", {})
    employees = data.get("employees", [])

    # ── Sheet 1: Employee Details ─────────────────────────────────────────────
    we = wb.active
    we.title = "Employee Details"
    we.sheet_view.showGridLines = False

    # Use passed fields or fallback to global mapping
    all_cols = active_employee_fields if active_employee_fields is not None else EMPLOYEE_FIELDS.get(sub_type, [])

    we.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(all_cols),1))
    t2 = we.cell(row=1, column=1, value=f"Employee Details - {len(employees)} records")
    t2.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t2.fill      = PatternFill("solid", fgColor=hex_col)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    we.row_dimensions[1].height = 28

    for ci, col in enumerate(all_cols, 1):
        c = we.cell(row=2, column=ci, value=col.replace("_", " "))
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, hdr_align, bdr
        we.column_dimensions[get_column_letter(ci)].width = 22
    we.row_dimensions[2].height = 22

    for ri, emp in enumerate(employees, 3):
        we.row_dimensions[ri].height = 18
        for ci, col in enumerate(all_cols, 1):
            c = we.cell(row=ri, column=ci, value=emp.get(col.lower(), ""))
            c.border, c.alignment = bdr, da
            c.font = Font(size=10, name="Calibri")
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F7F7F7")

    fin_cols = {"CURRENT_PREMIUM", "MONTHLY_TOTAL_PREMIUM", "GRAND_TOTAL", "TOTAL_COST", "ADJUSTMENT_AMOUNT"}
    fin_present = [c for c in all_cols if c.upper() in fin_cols]
    if fin_present and employees:
        tr = len(employees) + 3
        we.row_dimensions[tr].height = 20
        lc = we.cell(row=tr, column=1, value="TOTAL")
        lc.font, lc.fill, lc.border = tf, tfill, bdr
        for ci, col in enumerate(all_cols, 1):
            c = we.cell(row=tr, column=ci)
            c.fill, c.border = tfill, bdr
            if col.upper() in fin_cols:
                total = 0.0
                for emp in employees:
                    # Skip subtotal rows ("TOTAL" or empty plan_name) to avoid double counting
                    p_opt = emp.get("coverage_option", "")
                    pname = str(p_opt if p_opt is not None else "").strip().upper()
                    
                    if pname == "TOTAL" or (sub_type in ["engage", "velocity"] and (not pname or pname == "NONE")):
                        continue
                    
                    val_str = str(emp.get(col.lower(), "")).replace("$", "").replace(",", "")
                    try:
                        clean_val = re.sub(r'[^\d.-]', '', val_str)
                        if clean_val:
                            total += float(clean_val)
                    except:
                        pass
                c.value = f"${total:,.2f}"
                c.font  = tf

    we.freeze_panes = "A3"
    wb.active = we

    xlsx_path = out / f"{stem}_RPVE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(xlsx_path))
    print(f"[RPVE] Excel -> {xlsx_path.name}")
    return xlsx_path


def build_json_file(data: dict, sub_type: str, stem: str, active_employee_fields: list[str] | None = None, out_dir: Path | None = None) -> Path:
    out = out_dir if out_dir is not None else OUTPUT_DIR
    summary   = data.get("summary", {})
    employees = data.get("employees", [])
    required  = active_employee_fields if active_employee_fields is not None else EMPLOYEE_FIELDS.get(sub_type, [])

    rows = []
    for emp in employees:
        row = {k.upper(): v for k, v in summary.items()}
        # Only include required fields - strip everything else
        for col in required:
            row[col] = emp.get(col.lower(), "")
        row["RPVE_SUB_TYPE"] = sub_type.upper()
        rows.append(row)

    json_path = out / f"{stem}_RPVE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(str(json_path), "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    print(f"[RPVE] JSON  -> {json_path.name}")
    return json_path


# ══════════════════════════════════════════════════════════════════════════════
# FASTAPI APP
# ══════════════════════════════════════════════════════════════════════════════

app = FastAPI(
    title="RPVE - Benefit Invoice Extractor",
    description="Resourcing · Prestige · Velocity · Engage",
    version="1.0.0",
)

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

_cache: dict[str, str] = {}

FRONTEND_DIST_DIR = BASE_DIR / "frontend" / "dist"

if (FRONTEND_DIST_DIR / "assets").exists():
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIST_DIR / "assets"), name="assets")

@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def serve_ui():
    ui = FRONTEND_DIST_DIR / "index.html"
    if ui.exists():
        return HTMLResponse(content=ui.read_text(encoding="utf-8"))
    
    ui_fallback = BASE_DIR / "RPVE_ui.html"
    if ui_fallback.exists():
        return HTMLResponse(content=ui_fallback.read_text(encoding="utf-8"))
        
    return HTMLResponse("<h2>RPVE running</h2><p>Build the frontend first.</p><a href='/docs'>Swagger -></a>")


@app.get("/api/health")
async def health():
    return {"status": "ok", "service": "RPVE", "sub_types": list(KEYWORDS.keys())}


async def process_invoice_data(file_path: Path, original_filename: str):
    print(f"\n[RPVE] Processing -> {original_filename}")
    ext = Path(original_filename).suffix.lower()
    
    # ── UNIQUE OUTPUT PATH ────────────────────────────────────────────
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    # Clean and truncate stem for the output folder (cosmetic naming)
    safe = re.sub(r'[\\/:*?"<>|]', "_", original_filename)
    stem = Path(safe).stem.replace(" ", "_").strip()[:50]
    
    # Create specific output directory
    run_out_dir = OUTPUT_DIR / f"{stem}_{timestamp_str}"
    run_out_dir.mkdir(parents=True, exist_ok=True)

    try:
        # Using local extract_text function instead of missing identification module
        text = extract_text(file_path)
        
        # Consistent Text Output: Save the extracted text for ALL file types
        txt_path = run_out_dir / f"{stem}.txt"
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"[RPVE] Saved structured text to {txt_path.name}")
        
    except Exception as read_err:
        print(f"[RPVE] Read error: {read_err}")
        raise Exception(f"Failed to extract text from {ext} file: {read_err}")

    if not text.strip():
        raise Exception("No text extracted. File may be empty or an unreadable image.")

    # ---- Filename-based fallback for Data Link EMI ----
    # Normalise filename: replace underscores/hyphens with spaces so that
    # "Data_Link_EMI_Invoice..." and "Data-Link-EMI..." both match correctly.
    safe_normalised = safe.upper().replace("_", " ").replace("-", " ")
    if "DATA LINK EMI" in safe_normalised:
        sub_type = "datalink_emi"
        print(f"[RPVE] Filename -> Data Link EMI detected, forcing sub_type. (key: {safe_normalised[:40]})")
    else:
        # Use classification to determine carrier sub-type
        sub_type = classify(text)
    print(f"[RPVE] Classified as -> {sub_type.upper()}")

    try:
        data = extract_with_llm(sub_type, text)
    except Exception as e:
        raise Exception(f"LLM extraction failed: {str(e)}")

    # Clean up results: Remove rows without names (Center for Human Development fix)
    all_emps = data.get("employees", [])
    valid_emps = []
    for e in all_emps:
        # Require at least one name field to be present
        fname = str(e.get("first_name") or "").strip()
        lname = str(e.get("last_name") or "").strip()
        fulln = str(e.get("full_name") or "").strip()
        if (fname and lname) or fulln:
            valid_emps.append(e)
    data["employees"] = valid_emps
    emp_count = len(valid_emps)
    print(f"[RPVE] Extracted  -> {emp_count} rows")

    # ── Global FULL_NAME construction (applies to ALL file types) ─────────────
    # If the LLM did not return a full_name, build it from first/middle/last.
    for emp in data["employees"]:
        if not str(emp.get("full_name") or "").strip():
            parts = [
                str(emp.get("first_name") or "").strip(),
                str(emp.get("middal_name") or "").strip(),
                str(emp.get("last_name") or "").strip(),
            ]
            emp["full_name"] = " ".join(p for p in parts if p)
            
        # ── Global Plan Name Fallback ─────────────────────────────────────────
        # If the document lacks an explicit plan name and the LLM returns null
        # (to strictly avoid hallucination), fall back to using the plan type.
        if not str(emp.get("plan_name") or "").strip() and emp.get("plan_type"):
            emp["plan_name"] = emp.get("plan_type")

    # ── Post-LLM clean-up for Data Link EMI ─────────────────────────────────
    if sub_type == "datalink_emi":
        medical_lookup: dict[str, str] = {}
        retro_lookup:   dict[str, str] = {}
        try:
            import pdfplumber
            with pdfplumber.open(str(file_path)) as _pdf:
                for _page in _pdf.pages:
                    _tables = _page.extract_tables()
                    for _tbl in _tables:
                        for _row in _tbl:
                            if not _row or len(_row) < 5: continue
                            _name_cell = str(_row[1] or "").strip()
                            if not _name_cell or _name_cell.lower() in ("name", ""): continue
                            _is_retro = False
                            _medical_idx = 3
                            if len(_row) >= 9:
                                _date_cell = str(_row[2] or "")
                                if "/" in _date_cell or "-" in _date_cell:
                                    _is_retro = True
                                    _medical_idx = 4
                            _medical_val = str(_row[_medical_idx] or "").strip()
                            if not re.match(r'^\$?-?[\d,]+\.\d{2}$', _medical_val.replace("(", "").replace(")", "")): continue
                            _key = re.sub(r'\s+', ' ', _name_cell.upper().replace(",", "")).strip()
                            if _is_retro: retro_lookup[_key] = _medical_val
                            else: medical_lookup[_key] = _medical_val
            print(f"[RPVE] Data Link EMI: built lookup tables.")
        except Exception as _lookup_err:
            print(f"[RPVE] Data Link EMI: pdfplumber lookup failed ({_lookup_err})")

        def _emi_key(emp: dict) -> str:
            fn = str(emp.get("first_name") or "").strip().upper()
            ln = str(emp.get("last_name") or "").strip().upper()
            return f"{ln} {fn}" if fn and ln else ""

        for emp in data.get("employees", []):
            _k = _emi_key(emp)
            _is_retro_emp = emp.get("adjustment_amount") and not emp.get("current_premium")
            if _is_retro_emp:
                if _k and _k in retro_lookup: emp["adjustment_amount"] = retro_lookup[_k]
                emp["current_premium"] = None
            else:
                if _k and _k in medical_lookup: emp["current_premium"] = medical_lookup[_k]
                emp["adjustment_amount"] = None

    data["employees"] = deduplicate_employees(data["employees"])
    analysis_file_name = None

    try:
        active_fields = EMPLOYEE_FIELDS.get(sub_type, UNIFIED_FIELDS)
        extracted_text_upper = text.upper()
        is_strict_adp = ("TOTALSOURCE" in extracted_text_upper or "ADP" in extracted_text_upper or "NCT3-EPO" in extracted_text_upper)
        is_peo = is_strict_adp or "INSPERITY" in extracted_text_upper
        
        analysis_data = []
        if is_peo:
            from collections import defaultdict
            grouped2: dict = defaultdict(list)
            for emp in data.get("employees", []):
                pname = str(emp.get("plan_name") or "").strip().upper()
                ptype = str(emp.get("plan_type") or "").strip().upper()
                copt  = str(emp.get("coverage_option") or "").strip().upper()
                if any(x in pname or x in ptype or x in copt for x in ("TOTAL", "SUBTOTAL", "GRAND TOTAL")): continue
                key = (str(emp.get("first_name", "")).strip().upper(), str(emp.get("last_name", "")).strip().upper())
                grouped2[key].append(emp)

            collapsed = []
            for (fname, lname), rows in grouped2.items():
                if not rows: continue
                parsed_rows = []
                for r in rows:
                    val_str = str(r.get("current_premium") or "").replace("$", "").replace(",", "")
                    try: v = round(float(re.sub(r'[^\d.-]', '', val_str)), 2)
                    except: v = 0.0
                    parsed_rows.append((v, r))
                parsed_rows.sort(key=lambda x: x[0], reverse=True)
                valid_benefit_rows = []
                if len(parsed_rows) > 1:
                    top_val, top_row = parsed_rows[0]
                    remaining_sum = sum(v for v, _ in parsed_rows[1:])
                    if abs(top_val - remaining_sum) < 0.1: valid_benefit_rows = parsed_rows[1:]
                    else: valid_benefit_rows = parsed_rows
                else: valid_benefit_rows = parsed_rows
                if valid_benefit_rows:
                    valid_benefit_rows.sort(key=lambda x: x[0], reverse=True)
                    best_val, best_row = valid_benefit_rows[0]
                    pname = str(best_row.get("plan_name") or "").strip().upper()
                    if pname in ("TOTAL", "SUBTOTAL"): continue
                    cov_opt = str(best_row.get("coverage_option") or "").strip()
                    cat_name = str(best_row.get("plan_name") or "").strip()
                    if cov_opt and cov_opt.upper() not in ("TOTAL", "NONE", ""):
                        best_row = dict(best_row)
                        best_row["plan_name"] = cov_opt
                        if not best_row.get("plan_type"): best_row["plan_type"] = cat_name
                    if not str(best_row.get("full_name") or "").strip():
                        best_row = dict(best_row)
                        parts = [str(best_row.get("first_name") or "").strip(), str(best_row.get("middal_name") or "").strip(), str(best_row.get("last_name") or "").strip()]
                        best_row["full_name"] = " ".join(p for p in parts if p)
                    collapsed.append(best_row)
            data["employees"] = collapsed

        is_uhc = any(x in extracted_text_upper for x in ["UNITEDHEALTHCARE", "UNITED HEALTHCARE", "UHC"])
        is_excel = ext in [".xlsx", ".xls"]
        final_employees = []
        for emp in data.get("employees", []):
            val_str = str(emp.get("current_premium") or "").replace("$", "").replace(",", "")
            try: premium_val = round(float(re.sub(r'[^\d.-]', '', val_str)), 2)
            except: premium_val = 0.0
            # apply the $250 rule to ALL document types (including UHC, Excel, and Data Link)
            if premium_val < 250:
                analysis_data.append(emp)
            else:
                final_employees.append(emp)
        data["employees"] = final_employees

        xlsx_path = build_excel(data, sub_type, stem, active_employee_fields=active_fields, out_dir=run_out_dir)
        json_path = build_json_file(data, sub_type, stem, active_employee_fields=active_fields, out_dir=run_out_dir)
        if analysis_data:
            analysis_path = run_out_dir / f"{stem}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(str(analysis_path), "w", encoding="utf-8") as af:
                json.dump(analysis_data, af, indent=2, ensure_ascii=False)
            analysis_file_name = analysis_path.name
            _cache[analysis_file_name] = str(analysis_path)

    except Exception as build_err:
        import traceback
        print(f"[RPVE] Output building error:\n{traceback.format_exc()}")
        raise Exception(f"Failed to generate output files: {str(build_err)}")

    _cache[xlsx_path.name] = str(xlsx_path)
    _cache[json_path.name] = str(json_path)

    summary_dict = data.get("summary", {})
    total_val_str = "0"
    for tk in ["total_cost", "grand_total", "total_amount_due", "total_balance_due", "amount_due", "total_amount"]:
        val = summary_dict.get(tk) or summary_dict.get(tk.upper())
        if val:
            total_val_str = val
            break
    try: numeric_total = float(re.sub(r'[^0-9\.]', '', str(total_val_str)))
    except: numeric_total = 0.0

    return {
        "status": "success", "type": "INVOICE", "sub_type": sub_type, "sub_type_label": "Standard Mode",
        "employee_count": emp_count, "fields_in_excel": active_fields, "summary": summary_dict,
        "excel_file": xlsx_path.name, "json_file": json_path.name, 
        "excel_path": str(xlsx_path.absolute()), "json_path": str(json_path.absolute()),
        "output_file": xlsx_path.name,
        "output_json": json_path.name, "total_value": numeric_total,
        "excel_url": f"/api/download/{xlsx_path.name}", "json_url": f"/api/download/{json_path.name}",
        "analysis_file": analysis_file_name, "analysis_url": f"/api/download/{analysis_file_name}" if analysis_file_name else None,
        "employees": [{col: emp.get(col.lower(), "") for col in active_fields} for emp in data.get("employees", [])],
    }

@app.post("/api/extract")
async def extract(file: UploadFile = File(...)):
    print(f"\n[RPVE] Extraction Mode -> Standard")
    if not file.filename:
        raise HTTPException(400, "No filename provided")
    ext = Path(file.filename).suffix.lower()
    if ext not in [".pdf", ".csv", ".xlsx", ".xls"]:
        raise HTTPException(400, f"Supported formats: PDF, CSV, XLSX, XLS. Got: {ext}")

    unique_id = uuid.uuid4().hex[:8]
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_filename = re.sub(r'[\\/:*?\"<>|]', '_', file.filename)
    filename_unique = f"{unique_id}_{timestamp_str}_{safe_filename}"
    file_path = UPLOAD_DIR / filename_unique
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    try:
        return await process_invoice_data(file_path, file.filename)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/process-flow")
async def process_flow(files: list[UploadFile] = File(...)):
    try:
        print(f"\n[RPVE] Processing Flow with {len(files)} files")
        import sys
        if str(BASE_DIR) not in sys.path:
            sys.path.append(str(BASE_DIR))
        
        try:
            import flow_orchestrator
        except ImportError as e:
            raise HTTPException(500, f"Error importing flow orchestrator: {e}")

        pdf_file = None
        excel_files = []
        
        for file in files:
            if not file.filename: continue
            ext = Path(file.filename).suffix.lower()
            
            unique_id = uuid.uuid4().hex[:8]
            timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_filename = re.sub(r'[\\/:*?\"<>|]', '_', file.filename)
            filename_unique = f"{unique_id}_{timestamp_str}_{safe_filename}"
            file_path = UPLOAD_DIR / filename_unique
            
            with open(file_path, "wb") as f:
                shutil.copyfileobj(file.file, f)
                
            if ext == ".pdf":
                pdf_file = file_path
            elif ext in [".xlsx", ".xls"]:
                excel_files.append(file_path)

        if not pdf_file or not excel_files:
            raise HTTPException(400, "Please upload at least one PDF and one Excel template.")
            
        try:
            ref_census = excel_files[1] if len(excel_files) > 1 else None
            result = await flow_orchestrator.run_flow(str(pdf_file), str(excel_files[0]), str(ref_census) if ref_census else None)
            return result
        except Exception as e:
            import traceback
            traceback.print_exc()
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(500, str(e))
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/download/{filename}", include_in_schema=False)
async def download(filename: str):
    # 1. Try to get from in-memory cache
    path_str = _cache.get(filename)
    if path_str:
        fp = Path(path_str)
        if fp.exists():
            mt = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                  if filename.endswith(".xlsx") else "application/json")
            return FileResponse(path=fp, filename=filename, media_type=mt)

    # 2. If not in cache (e.g. server restarted), search subdirectories of OUTPUT_DIR
    # This ensures files are still downloadable even after a reload.
    matches = list(OUTPUT_DIR.rglob(filename))
    if matches:
        fp = matches[0]
        mt = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
              if filename.endswith(".xlsx") else "application/json")
        return FileResponse(path=fp, filename=filename, media_type=mt)

    # 3. Fallback to root of OUTPUT_DIR
    fp = OUTPUT_DIR / filename
    if fp.exists():
        mt = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
              if filename.endswith(".xlsx") else "application/json")
        return FileResponse(path=fp, filename=filename, media_type=mt)
        
    raise HTTPException(404, f"File not found: {filename}")


@app.get("/{filename}", include_in_schema=False)
async def serve_root_files(filename: str):
    if filename.startswith("api") or filename in ["docs", "openapi.json"]:
        raise HTTPException(404)
    file_path = FRONTEND_DIST_DIR / filename
    if file_path.exists() and file_path.is_file():
        return FileResponse(file_path)
    raise HTTPException(404)


if __name__ == "__main__":
    import uvicorn, sys
    port = 8009
    if "--port" in sys.argv:
        try: port = int(sys.argv[sys.argv.index("--port") + 1])
        except: pass

    print("\n" + "="*50)
    print("  RPVE - Benefit Invoice Extractor")
    print("="*50)
    print(f"  UI      ->  http://localhost:{port}")
    print(f"  Swagger ->  http://localhost:{port}/docs")
    print("="*50 + "\n")
    uvicorn.run("RPVE_standalone:app", host="0.0.0.0", port=port, reload=True, reload_excludes=["rpve_uploads/*", "rpve_outputs/*", "*.log", "*.txt"])