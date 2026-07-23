"""
flow_orchestrator.py
====================
Synchronous RPVE pipeline orchestrator.

run_job() is called from a worker thread (job_worker.py).
Because it runs in a thread (not on the asyncio event loop),
all subprocess.run() calls here are safe — they block only
the current thread, not the entire server.

Business logic is 100% preserved from the original run_flow():
  - .xls → .xlsx conversion
  - 3-way swap intelligence (name / volume / format)
  - Phase 2 template classification (type1 / type2 / type3)
  - Phase 3 data validation (non-fatal fallback)
  - Phase 4 LLM resolution (non-fatal fallback)
"""

import logging
import sys
import os
import threading
import importlib.util
from pathlib import Path
from typing import Callable, Optional

import pandas as pd

# Global lock for thread-safe module loading
_load_lock = threading.Lock()

# Cache for loaded modules to avoid redundant exec_module
_module_cache = {}

def load_phase_module(name: str, path: Path):
    """Safely loads a module from a path, ensuring its parent is in sys.path for local imports."""
    with _load_lock:
        if name in _module_cache:
            return _module_cache[name]
        
        spec = importlib.util.spec_from_file_location(name, str(path))
        module = importlib.util.module_from_spec(spec)
        
        # Add parent directory to sys.path temporarily to resolve local imports in the script
        parent_dir = str(path.parent)
        sys.path.insert(0, parent_dir)
        try:
            spec.loader.exec_module(module)
        finally:
            sys.path.pop(0)
            
        _module_cache[name] = module
        return module

# ─────────────────────────────────────────────────────────────────────────────
# Helpers (preserved from original)
# ─────────────────────────────────────────────────────────────────────────────

# Set up paths so that RPVE_standalone can be imported regardless of cwd
CURRENT_DIR = Path(__file__).parent.absolute()
sys.path.append(str(CURRENT_DIR))

try:
    import RPVE_standalone
    from RPVE_standalone import process_invoice_data_sync, OUTPUT_DIR
except ImportError as e:
    print(f"Error: Could not import RPVE_standalone.py. {e}")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers (preserved from original)
# ─────────────────────────────────────────────────────────────────────────────

def ensure_xlsx(file_path: Optional[str]) -> Optional[str]:
    """If file is .xls, convert to .xlsx using pandas and return new path."""
    if not file_path or not file_path.lower().endswith(".xls"):
        return file_path

    print(f"    -> Converting old .xls format to .xlsx: {Path(file_path).name}")
    try:
        new_path = str(Path(file_path).with_suffix(".xlsx"))
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            xl = pd.ExcelFile(file_path)
            with pd.ExcelWriter(new_path, engine='openpyxl') as writer:
                for sheet_name in xl.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        return new_path
    except Exception as e:
        print(f"    -> Warning: Failed to convert .xls: {e}")
        return file_path


def classify_excel_template(excel_path: Path) -> str:
    """Scans the first 25 rows of an Excel file to fingerprint its Type."""
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(excel_path, nrows=25, header=None)
        all_text = " ".join([str(val).lower() for val in df.values.flatten() if pd.notna(val)])
        
        # Type 1 Detection (Engage/Kaiser)
        if "ee row" in all_text or "relation-ship to employee" in all_text or "kaiser networks" in all_text:
            return "type1"
            
        # Type 3 Detection (RAPT Blue Headers)
        if "data row" in all_text or "cobra participant" in all_text or "discrepancies" in all_text:
            return "type3"
            
        # Type 2 (Basic Titan Intake / Generic Census)
        elif "first name" in all_text or "last name" in all_text or "name" in all_text:
            return "type2"

        # Default fallback to type2 (dynamic column header mapping)
        return "type2"
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return "type2"


def is_likely_source_invoice(excel_path: Path) -> bool:
    """
    Returns True if the Excel file looks more like an invoice/billing report than a census.
    Heuristics:
    1. Filename contains 'report', 'billing', 'invoice', 'premium', 'invoice'.
    2. Content contains 'monthly premium', 'billed', 'arrears', 'amount due'.
    """
    name = excel_path.name.lower()
    if any(k in name for k in ['report', 'billing', 'invoice', 'premium']):
        if 'census' not in name:
            return True
    
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(excel_path, nrows=20, header=None)
        all_text = " ".join([str(val).lower() for val in df.values.flatten() if pd.notna(val)])
        
        invoice_kws = ['premium', 'billed', 'amount due', 'arrears', 'total cost', 'billing period']
        census_kws  = ['census', 'home zip', 'cobra', 'hire date', 'birth date']
        
        invoice_score = sum(1 for k in invoice_kws if k in all_text)
        census_score  = sum(1 for k in census_kws if k in all_text)
        
        return invoice_score > census_score
    except:
        return False


def post_process_clear_wo_rows(file_path: Path, logger: logging.Logger):
    """
    [ignoring loop detection]
    Post-processing layer: If the plan and premium/amount are present on a row with coverage 'WO', clear them.
    Also clears discrepancies/status column.
    """
    if not file_path or not file_path.exists():
        return

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path))
        modified = False
        
        # Process all sheets or at least the active sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find the header row
            header_row = None
            for r in range(1, 40):
                row_vals = [str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, min(ws.max_column + 1, 20))]
                if any(kw in row_vals for kw in ('ee row', 'first name', 'last name', 'relationship', 'coverage')):
                    header_row = r
                    break
            
            if not header_row:
                continue
                
            # Scan columns in the header row
            cov_cols = []
            plan_cols = []
            prem_cols = []
            disc_cols = []
            
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
                if not val:
                    continue
                if 'tier' in val or 'level' in val or 'coverage type' in val:
                    # 'coverage tier' / 'coverage level' / 'tier' / 'coverage type' is highly specific
                    cov_cols.insert(0, c) # prioritised
                elif 'coverage' in val and 'note' not in val and 'desc' not in val:
                    cov_cols.append(c)
                elif 'plan' in val or 'product' in val or 'desc' in val:
                    plan_cols.append(c)
                elif 'premium' in val or 'amount' in val or 'rate' in val or 'cost' in val:
                    prem_cols.append(c)
                elif 'discrep' in val or 'notes' in val or 'status' in val:
                    disc_cols.append(c)
            
            # Fallback for coverage if none found
            if not cov_cols:
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
                    if 'coverage' in val:
                        cov_cols.append(c)
            
            if not cov_cols:
                continue
                
            cov_col = cov_cols[0]
            logger.info(f"Post-processing clear WO: Sheet '{sheet_name}', Header Row {header_row}, Coverage Col {cov_col}")
            
            cleared_count = 0
            for r in range(header_row + 1, ws.max_row + 1):
                cov_val = str(ws.cell(row=r, column=cov_col).value or '').strip().upper()
                if cov_val == 'WO':
                    # Clear plan columns
                    for pc in plan_cols:
                        if ws.cell(row=r, column=pc).value is not None:
                            ws.cell(row=r, column=pc).value = None
                            cleared_count += 1
                    # Clear premium columns
                    for prc in prem_cols:
                        if ws.cell(row=r, column=prc).value is not None:
                            ws.cell(row=r, column=prc).value = None
                            cleared_count += 1
                    # Clear discrepancy columns
                    for dc in disc_cols:
                        if ws.cell(row=r, column=dc).value is not None:
                            ws.cell(row=r, column=dc).value = None
            
            if cleared_count > 0:
                logger.info(f"Post-processing clear WO: Cleared {cleared_count} fields in sheet '{sheet_name}'")
                modified = True
                
        if modified:
            wb.save(str(file_path))
    except Exception as e:
        logger.error(f"Post-processing clear WO failed for {file_path}: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point — called from job_worker._execute_job()
# ─────────────────────────────────────────────────────────────────────────────


def run_job(
    job_id: str,
    pdf_path: Path,
    template_path: str,
    ref_census_path: Optional[str],
    job_dir: Path,
    status_callback: Callable[[str, Optional[str]], None],
    logger: logging.Logger,
) -> dict:
    """
    Run the full RPVE pipeline for a single job.

    Parameters
    ----------
    job_id          : Unique job identifier (backend-internal only)
    pdf_path        : Absolute Path to the uploaded invoice PDF/XLSX/CSV
    template_path   : Absolute path string to the census template Excel
    ref_census_path : Optional absolute path string to a reference census Excel
    job_dir         : Per-job workspace root (jobs/{job_id}/)
    status_callback : Called at each phase transition → (phase_name, template_type?)
    logger          : Per-job file logger

    Returns
    -------
    dict — the merged result dict (same shape as the old run_flow() return value)
    """

    work_dir   = job_dir / "work"
    output_dir = job_dir / "output"
    work_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    logger.info("--- Starting End-to-End RPVE Flow (job_id=%s) ---", job_id)

    # Initialize Phase 1 baseline variables at function scope
    phase1_baseline_json = None
    phase1_baseline_excel = None

    # ── PHASE 1: Extraction (or Bypass if Excel provided) ─────────────────────
    is_direct_excel = pdf_path.suffix.lower() in ['.xlsx', '.xls']

    if is_direct_excel:
        logger.info("[1] Skipping Phase 1 - Direct Excel input detected: %s", pdf_path.name)
        
        # If it's an old .xls, we must convert it first
        if pdf_path.suffix.lower() == '.xls':
            converted_path = ensure_xlsx(str(pdf_path))
            phase1_output_excel = converted_path
        else:
            phase1_output_excel = str(pdf_path)
            
        # Mock result data for subsequent phases (no RPVE JSON in direct Excel mode)
        result_data = {
            "excel_path": phase1_output_excel,
            "json_path":  None,
            "text_path":  None,
            "output_json": None,
            "json_file":   None,
        }
        phase2_report_path    = work_dir / f"FINAL_AUDIT_REPORT_{job_id}.xlsx"
        validated_report_path = work_dir / f"VALIDATED_AUDIT_REPORT_{job_id}.xlsx"
        status_callback("classifying") # Skip Phase 1 callback
    else:
        status_callback("phase1")
        logger.info("[1] Running Phase 1 (Extraction) on: %s", pdf_path.name)
        try:
            result_data = process_invoice_data_sync(pdf_path, pdf_path.name, out_dir=work_dir)
            phase1_output_excel = result_data["excel_path"]
            logger.info("    -> Phase 1 Success! Extracted: %s", Path(phase1_output_excel).name)

            phase2_report_path    = work_dir / f"FINAL_AUDIT_REPORT_{job_id}.xlsx"
            validated_report_path = work_dir / f"VALIDATED_AUDIT_REPORT_{job_id}.xlsx"
            
            # ── PHASE 0.5: Generate Phase 1 Baseline using CACHED text ──────────
            logger.info("[0.5] Generating Phase 1 Baseline with dedicated extraction...")
            try:
                # Check if text file was already created by RPVE extraction
                text_path = result_data.get("text_path")
                if not text_path:
                    # Look for text file in work directory
                    text_files = list(work_dir.glob("*.txt"))
                    if text_files:
                        text_path = str(text_files[0])
                
                extracted_text = None
                if text_path and Path(text_path).exists():
                    logger.info("    -> ✓ OPTIMIZATION: Reusing cached text from: %s", Path(text_path).name)
                    with open(text_path, 'r', encoding='utf-8') as f:
                        extracted_text = f.read()
                    logger.info("    -> Loaded %d characters (skipping PDF re-extraction)", len(extracted_text))
                
                # Use dedicated Phase 1 extraction with cached text
                from generate_phase1_output import generate_phase1_output_from_text
                
                if extracted_text:
                    phase1_result = generate_phase1_output_from_text(
                        text=extracted_text,
                        output_dir=work_dir,
                        base_filename=pdf_path.stem
                    )
                else:
                    # Fallback: extract from PDF (slower)
                    logger.warning("    -> ⚠ Text file not found, falling back to PDF extraction (slower)...")
                    from generate_phase1_output import generate_phase1_output
                    phase1_result = generate_phase1_output(
                        pdf_path=pdf_path,
                        output_dir=work_dir
                    )
                
                phase1_baseline_json = phase1_result['json_path']
                phase1_baseline_excel = phase1_result['excel_path']
                
                logger.info("    -> Phase 1 Baseline Success!")
                logger.info("    -> JSON: %s", phase1_baseline_json.name)
                logger.info("    -> Excel: %s", phase1_baseline_excel.name)
                logger.info("    -> Records: %s (duplicates allowed)", phase1_result['record_count'])
                    
            except Exception as e:
                import traceback
                logger.warning("    [WARN] Phase 1 Baseline generation failed (non-fatal): %s", e)
                logger.debug(traceback.format_exc())

        except Exception as e:
            import traceback
            logger.error("    -> Phase 1 Failed: %s\n%s", e, traceback.format_exc())
            raise

    # ── PHASE 1.5: Pre-process Templates (.xls -> .xlsx) ────────────────────
    template_path   = ensure_xlsx(template_path)
    ref_census_path = ensure_xlsx(ref_census_path)

    # ── PHASE 2: Template Classification + Fill ──────────────────────────────
    status_callback("classifying")
    logger.info("[2] Analyzing Template: %s", template_path)

    template_type = classify_excel_template(Path(template_path))

    if ref_census_path:
        ref_type = classify_excel_template(Path(ref_census_path))

        # ── ROBUST SWAP INTELLIGENCE (preserved exactly from original) ───────
        t_name = Path(template_path).name.lower()
        r_name = Path(ref_census_path).name.lower()

        logger.info("    -> Swap Check: Template=%s, Ref=%s", t_name, r_name)
        logger.info("    -> Type Check: Template=%s, Ref=%s", template_type, ref_type)

        swapped = False

        # 1. Name-based hint
        if template_type not in ["type1", "type3"] and (
            ("rapt" in r_name and "rapt" not in t_name)
            or ("template" in r_name and "template" not in t_name)
        ):
            logger.info("    -> Name-based Swap: User put RAPT/Template file in Reference slot.")
            template_path, ref_census_path = ref_census_path, template_path
            template_type = classify_excel_template(Path(template_path))
            ref_type      = classify_excel_template(Path(ref_census_path))
            swapped = True
            logger.info("    -> Post-Swap Types: Template=%s, Ref=%s", template_type, ref_type)

        # 2. Volume-based check
        if not swapped:
            try:
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    t_df = pd.read_excel(template_path, nrows=100)
                    r_df = pd.read_excel(ref_census_path, nrows=100)
                    t_val = t_df.iloc[:, :3].count().sum()
                    r_val = r_df.iloc[:, :3].count().sum()
                    logger.info("    -> Content Check: Template=%s values, Ref=%s values", t_val, r_val)
                    if t_val > (r_val + 10):
                        logger.info("    -> Volume-based Swap Triggered!")
                        template_path, ref_census_path = ref_census_path, template_path
                        template_type = classify_excel_template(Path(template_path))
                        ref_type      = classify_excel_template(Path(ref_census_path))
                        swapped = True
            except Exception as e:
                logger.warning("    -> Volume check failed: %s", e)

        # 3. Format-based fallback
        if not swapped and ref_type == "type3" and template_type != "type3":
            logger.info("    -> Format-based Swap: User put RAPT file in Reference slot.")
            template_path, ref_census_path = ref_census_path, template_path
            template_type = "type3"
        elif ref_census_path:
            template_type = "type3"

        logger.info("    -> Final Decision: Template Type=%s", template_type.upper())
    else:
        logger.info("    -> Detected Template Type: %s", template_type.upper())

    # Record template_type in DB
    status_callback("phase2", template_type)

    phase2_base = CURRENT_DIR / "phase_2" / "template_fill"

    logger.info("[3] Executing Phase 2 (Template Fill)...")
    try:
        if template_type == "type1":
            script = phase2_base / "type1" / "fill_template_v2.py"
            mod = load_phase_module("filler_type1", script)
            filler = mod.DynamicCensusFiller()
            if filler.load_source(phase1_output_excel):
                if not filler.fill_template(template_path, str(phase2_report_path)):
                    raise RuntimeError("Phase 2 Type 1 filler failed")
            else:
                raise RuntimeError("Phase 2 Type 1 could not load source")

        elif template_type == "type2":
            script = phase2_base / "type2" / "fill_template.py"
            mod = load_phase_module("filler_type2", script)
            filler = mod.DynamicCensusFiller()
            if filler.load_source(phase1_output_excel):
                if not filler.fill_template(template_path, str(phase2_report_path)):
                    raise RuntimeError("Phase 2 Type 2 filler failed")
            else:
                raise RuntimeError("Phase 2 Type 2 could not load source")

        elif template_type == "type3":
            if not ref_census_path:
                default_ref = phase2_base / "type3" / "reference_census" / "TEPCensus.xlsx"
                if default_ref.exists():
                    ref_census_path = str(default_ref)
                else:
                    raise ValueError("Type 3 requires a Reference Census file!")
            script = phase2_base / "type3" / "fill_template.py"
            mod = load_phase_module("filler_type3", script)
            mod.fill_rapt_template(phase1_output_excel, ref_census_path, template_path, str(phase2_report_path))

        else:
            raise ValueError(f"Could not identify Excel template type: {template_type}")

        logger.info("    -> Phase 2 (Census Fill) Success!")
    except Exception as e:
        import traceback
        logger.error("--- Flow Failed in Phase 2 ---\n%s\n%s", e, traceback.format_exc())
        raise

    # ── PHASE 3: Data Validation ─────────────────────────────────────────────
    status_callback("phase3")
    logger.info("[4] Executing Phase 3 (Data Validation)...")
    validation_script = phase2_base / "data_validation.py"

    if not validation_script.exists():
        logger.warning("    [WARN] data_validation.py not found. Skipping Phase 3.")
        validated_report_path = phase2_report_path
    else:
        try:
            mod = load_phase_module("data_validation", validation_script)
            mod.run_validation(
                filled_path=str(phase2_report_path),
                invoice_path=phase1_output_excel,
                output_path=str(validated_report_path),
                threshold=85,
                template_type=template_type
            )
            if validated_report_path.exists():
                logger.info("    -> Phase 3 (Data Validation) Success!")
            else:
                logger.warning("    [WARN] Phase 3 did not create the validated report file. Using Phase 2 output.")
                validated_report_path = phase2_report_path
        except Exception as e:
            import traceback
            logger.warning("    [WARN] Phase 3 failed (non-fatal). Using Phase 2 output.\n%s\n%s", e, traceback.format_exc())
            validated_report_path = phase2_report_path

    # ── PHASE 4: LLM Resolution ───────────────────────────────────────────────
    status_callback("phase4")
    logger.info("[5] Executing Phase 4 (LLM Resolution)...")
    llm_script     = phase2_base / "llm_resolution.py"
    audit_json     = validated_report_path.with_suffix(".audit.json")
    llm_report_path = validated_report_path.with_name(
        validated_report_path.name.replace("VALIDATED_", "LLM_RESOLVED_")
    )

    if not llm_script.exists() or not audit_json.exists():
        logger.warning("    [WARN] LLM resolution requirements not met. Skipping Phase 4.")
        final_report_path = validated_report_path
    else:
        try:
            mod = load_phase_module("llm_resolution", llm_script)
            mod.run_llm_resolution(
                validated_excel=validated_report_path,
                audit_json=audit_json,
                output_excel=llm_report_path,
                template_type=template_type
            )
            if llm_report_path.exists():
                logger.info("    -> Phase 4 (LLM Resolution) Success!")
                final_report_path = llm_report_path
            else:
                raise RuntimeError("LLM report file not created")
        except Exception as e:
            import traceback
            logger.warning(
                "    [WARN] Phase 4 failed or skipped. Using Phase 3 output.\n%s\n%s",
                e, traceback.format_exc()
            )
            final_report_path = validated_report_path

    # ── Clean 'WO' waiver rows in all output files ─────────────────────────
    logger.info("[Post-process] Cleaning WO rows in report files...")
    report_paths_to_clean = {phase2_report_path, validated_report_path, final_report_path}
    for rpath in report_paths_to_clean:
        if rpath:
            post_process_clear_wo_rows(Path(rpath), logger)

    # ── Build merged result (same shape as original run_flow return value) ───
    merged_result = result_data.copy()

    final_report_name     = final_report_path.name
    phase2_report_name    = phase2_report_path.name
    validated_report_name = validated_report_path.name
    
    # Preserve the original RPVE JSON path from Phase 1 for top JSON download button
    phase1_rpve_json = result_data.get("json_path")
    phase1_rpve_json_name = result_data.get("output_json") or result_data.get("json_file")
    
    # Ensure mandatory fields for UI
    merged_result["status"]      = "success"
    merged_result["type"]        = "INVOICE"
    merged_result["sub_type"]    = "standard"
    merged_result["output_file"] = final_report_name
    merged_result["excel_file"]  = final_report_name
    merged_result["excel_url"]   = f"/api/download/{final_report_name}"
    merged_result["final_report_path"] = str(final_report_path.absolute())
    merged_result["excel_path"]        = str(final_report_path.absolute())

    # Top JSON button should download the Phase 1 RPVE JSON (Oxford_Invoice_*_RPVE_*.json)
    if phase1_rpve_json and Path(phase1_rpve_json).exists():
        merged_result["output_json"] = phase1_rpve_json_name
        merged_result["json_file"]   = phase1_rpve_json_name
        merged_result["json_url"]    = f"/api/download/{phase1_rpve_json_name}"
    elif not is_direct_excel:
        # If Phase 1 ran but JSON is missing, log a warning
        logger.warning("    [WARN] Phase 1 RPVE JSON not found. Top JSON download may not work.")
    
    # Keep audit JSON for UI table display (separate key)
    if audit_json.exists():
        merged_result["audit_json"] = audit_json.name
        merged_result["audit_json_url"] = f"/api/download/{audit_json.name}"

    merged_result["phase1_invoice_excel"]         = Path(phase1_output_excel).name
    merged_result["phase1_invoice_excel_url"]     = f"/api/download/{Path(phase1_output_excel).name}"
    merged_result["phase2_filled_census"]         = phase2_report_name
    merged_result["phase2_filled_census_url"]     = f"/api/download/{phase2_report_name}"
    merged_result["phase3_validated_census"]      = validated_report_name
    merged_result["phase3_validated_census_url"]  = f"/api/download/{validated_report_name}"

    if final_report_path != validated_report_path:
        merged_result["phase4_llm_census"]      = final_report_name
        merged_result["phase4_llm_census_url"]  = f"/api/download/{final_report_name}"
    
    # Add Phase 1 Baseline outputs to result if generated
    if phase1_baseline_json and phase1_baseline_json.exists():
        merged_result["phase1_baseline_json"] = phase1_baseline_json.name
        merged_result["phase1_baseline_json_url"] = f"/api/download/{phase1_baseline_json.name}"
    if phase1_baseline_excel and phase1_baseline_excel.exists():
        merged_result["phase1_baseline_excel"] = phase1_baseline_excel.name
        merged_result["phase1_baseline_excel_url"] = f"/api/download/{phase1_baseline_excel.name}"

    logger.info("--- Flow Completed Successfully! ---")
    logger.info("  Excel 1 — Invoice Extraction : %s", Path(phase1_output_excel).name)
    logger.info("  Excel 2 — Filled Census      : %s", phase2_report_name)
    logger.info("  Excel 3 — Validated Census   : %s", validated_report_name)
    if final_report_path != validated_report_path:
        logger.info("  Excel 4 — LLM Resolved       : %s", final_report_name)

    return merged_result


# ─────────────────────────────────────────────────────────────────────────────
# Legacy CLI entry point (for direct invocation testing)
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uuid
    if len(sys.argv) < 3:
        print("Usage: python flow_orchestrator.py <pdf_path> <template_excel> [ref_census_excel]")
        sys.exit(1)

    _pdf      = Path(sys.argv[1])
    _template = sys.argv[2]
    _ref      = sys.argv[3] if len(sys.argv) > 3 else None
    _job_id   = f"cli_{uuid.uuid4().hex[:8]}"
    _job_dir  = Path("jobs") / _job_id
    _job_dir.mkdir(parents=True, exist_ok=True)

    _logger = logging.getLogger("rpve.cli")
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")

    result = run_job(
        job_id=_job_id,
        pdf_path=_pdf,
        template_path=_template,
        ref_census_path=_ref,
        job_dir=_job_dir,
        status_callback=lambda phase, ttype=None: print(f"[STATUS] {phase}"),
        logger=_logger,
    )
    import json
    print(json.dumps(result, indent=2, default=str))