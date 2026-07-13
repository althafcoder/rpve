from openpyxl import load_workbook

wb = load_workbook('jobs/fbded347c33943ec9880088b02aac1cd/work/VALIDATED_AUDIT_REPORT_fbded347c33943ec9880088b02aac1cd.xlsx')

ws = None
for s in wb.sheetnames:
    if any(k in s.lower() for k in ('census', 'employee', 'table', 'sheet')):
        ws = wb[s]
        break
if ws is None:
    ws = wb.active

# Reproduce the FIXED _find_columns logic (with "first match wins" guards)
best_cols = {}
best_score = -1
header_row = 1

for r in range(1, 40):
    row_vals = {
        c: str(ws.cell(row=r, column=c).value or '').strip().lower()
        for c in range(1, min(ws.max_column + 1, 50))
    }
    
    score = 0
    current_cols = {
        'name': None, 'first': None, 'last': None,
        'plan': None, 'premium': None, 'disc': None,
        'relation': None,
        'coverage': None,
    }
    
    for c, v in row_vals.items():
        if   ('employee' in v and 'name' in v) or ('full' in v and 'name' in v):
            current_cols['name'] = c; score += 2
        elif 'first' in v and 'name' in v:
            current_cols['first'] = c; score += 2
        elif 'last' in v and 'name' in v:
            current_cols['last'] = c; score += 2
        elif 'premium' in v and not current_cols['premium']:
            current_cols['premium'] = c; score += 1
        elif 'plan' in v and not current_cols['plan']:
            current_cols['plan'] = c; score += 1
        elif ('discrep' in v or v == 'notes') and not current_cols['disc']:
            current_cols['disc'] = c; score += 3
        elif 'relation' in v and 'discrep' not in v and not current_cols['relation']:
            current_cols['relation'] = c; score += 1
        elif ('coverage' in v or 'tier' in v) and not current_cols['coverage']:
            current_cols['coverage'] = c; score += 1
    
    if score > best_score and (current_cols['first'] or current_cols['name']):
        best_score = score
        best_cols = current_cols
        header_row = r
        print(f"  ** New best at row {r}, score={score}, cols={current_cols}")

best_cols['header_row'] = header_row
print(f"\nFINAL col_positions = {best_cols}")
print(f"Coverage column: {best_cols.get('coverage')}")

# Now check what value is at coverage col for a few data rows
cov_col = best_cols.get('coverage')
if cov_col:
    for r in [23, 26, 31, 32, 42, 46, 47, 54]:
        val = ws.cell(row=r, column=cov_col).value
        print(f"  Row {r}, col {cov_col}: value = {repr(val)}")
else:
    print("  ** Coverage column is None!")
