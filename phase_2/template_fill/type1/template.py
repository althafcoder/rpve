#!/usr/bin/env python3
"""
Excel Form Filler - Dynamic POC
Takes 2 Excel files (source data + template form) and fills the form with matched data
Usage: python excel_form_filler.py <source_file> <template_file> <output_file> [--config config.json]
"""

import sys
import json
import argparse
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelFormFiller:
    """
    Fills Excel form template with data from source Excel file.
    Supports dynamic field mapping via configuration.
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize with optional configuration dictionary."""
        self.config = config or self._get_default_config()
        self.source_data = None
        self.template_wb = None
        self.match_results = []
        
    def _get_default_config(self) -> Dict[str, Any]:
        """Return default configuration for Employee Details -> Engage PEO form."""
        return {
            "source_sheet": "Employee Details",
            "form_sheet": "Census",
            "match_field": "full name",
            "form_name_column": "A",
            "form_start_row": 22,
            "fuzzy_match": True,
            "fuzzy_threshold": 0.8,
            "fields": [
                {"source": "first name", "form_column": "B", "description": "First Name"},
                {"source": "last name", "form_column": "C", "description": "Last Name"},
                {"source": "plan type", "form_column": "M", "description": "Plan Type"},
                {"source": "plan name", "form_column": "N", "description": "Current Plan"},
                {"source": "current premium", "form_column": "O", "description": "Current Premium"},
            ]
        }
    
    def load_source_data(self, source_file: str) -> pd.DataFrame:
        """Load source Excel file."""
        try:
            sheet_name = self.config.get('source_sheet', 'Employee Details')
            df = pd.read_excel(source_file, sheet_name=sheet_name, skiprows=1)
            self.source_data = df
            logger.info(f"Loaded source data: {len(df)} rows, {len(df.columns)} columns")
            logger.info(f"Source columns: {list(df.columns)}")
            return df
        except Exception as e:
            logger.error(f"Error loading source file: {e}")
            raise
    
    def load_template(self, template_file: str):
        """Load Excel template file."""
        try:
            self.template_wb = load_workbook(template_file)
            logger.info(f"Loaded template with sheets: {self.template_wb.sheetnames}")
        except Exception as e:
            logger.error(f"Error loading template file: {e}")
            raise
    
    def _normalize_name(self, name: str) -> str:
        """Normalize name for matching (lowercase, strip whitespace)."""
        if pd.isna(name):
            return ""
        return str(name).strip().lower()
    
    def _fuzzy_match(self, target: str, candidates: List[str], threshold: float = 0.8) -> Optional[str]:
        """Find best fuzzy match for target in candidates list."""
        target_norm = self._normalize_name(target)
        if not target_norm:
            return None
        
        best_match = None
        best_score = threshold
        
        for candidate in candidates:
            candidate_norm = self._normalize_name(candidate)
            if not candidate_norm:
                continue
            
            score = SequenceMatcher(None, target_norm, candidate_norm).ratio()
            if score > best_score:
                best_score = score
                best_match = candidate
        
        return best_match
    
    def _find_matching_record(self, form_name: str) -> Optional[Dict[str, Any]]:
        """Find matching record in source data by name."""
        if self.source_data is None:
            return None
        
        match_field = self.config.get('match_field', 'full_name')
        
        # Exact match first
        exact_matches = self.source_data[
            self.source_data[match_field].apply(lambda x: self._normalize_name(x) == self._normalize_name(form_name))
        ]
        
        if not exact_matches.empty:
            record = exact_matches.iloc[0].to_dict()
            self.match_results.append({
                'form_name': form_name,
                'matched_name': exact_matches.iloc[0][match_field],
                'match_type': 'exact'
            })
            return record
        
        # Fuzzy match if enabled
        if self.config.get('fuzzy_match', True):
            threshold = self.config.get('fuzzy_threshold', 0.8)
            candidates = self.source_data[match_field].tolist()
            best_match = self._fuzzy_match(form_name, candidates, threshold)
            
            if best_match:
                record = self.source_data[
                    self.source_data[match_field].apply(lambda x: self._normalize_name(x) == self._normalize_name(best_match))
                ].iloc[0].to_dict()
                self.match_results.append({
                    'form_name': form_name,
                    'matched_name': best_match,
                    'match_type': 'fuzzy'
                })
                return record
        
        self.match_results.append({
            'form_name': form_name,
            'matched_name': None,
            'match_type': 'no_match'
        })
        return None
    
    def _get_value_from_record(self, record: Dict[str, Any], field_key: str) -> Any:
        """Safely extract value from record."""
        try:
            value = record.get(field_key)
            # Handle NaN values
            if pd.isna(value):
                return None
            return value
        except Exception as e:
            logger.warning(f"Error extracting field {field_key}: {e}")
            return None
    
    def fill_form(self, output_file: str):
        """Fill template form with matched source data."""
        if self.source_data is None or self.template_wb is None:
            raise ValueError("Source data and template must be loaded first")
        
        sheet_name = self.config.get('form_sheet', 'Census')
        if sheet_name not in self.template_wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in template")
        
        ws = self.template_wb[sheet_name]
        form_name_col = self.config.get('form_name_column', 'A')
        start_row = self.config.get('form_start_row', 22)
        fields = self.config.get('fields', [])
        
        filled_count = 0
        current_row = start_row
        
        # Iterate through form rows
        while current_row <= ws.max_row:
            # Get employee name from form
            form_name_cell = ws[f'{form_name_col}{current_row}']
            form_name = form_name_cell.value
            
            # Stop if empty cell (end of data)
            if form_name is None or str(form_name).strip() == "":
                break
            
            # Find matching record
            matched_record = self._find_matching_record(str(form_name))
            
            if matched_record:
                # Fill each field
                for field_config in fields:
                    source_field = field_config.get('source')
                    form_column = field_config.get('form_column')
                    
                    if source_field and form_column:
                        value = self._get_value_from_record(matched_record, source_field)
                        if value is not None:
                            cell = ws[f'{form_column}{current_row}']
                            cell.value = value
                            logger.debug(f"Row {current_row}: {form_column} = {value}")
                
                filled_count += 1
                logger.info(f"Filled row {current_row}: {form_name}")
            else:
                logger.warning(f"Row {current_row}: No match found for '{form_name}'")
            
            current_row += 1
        
        # Save output
        self.template_wb.save(output_file)
        logger.info(f"Form filled and saved to: {output_file}")
        logger.info(f"Total rows filled: {filled_count}")
        
        return filled_count
    
    def print_match_report(self):
        """Print summary of matching results."""
        print("\n" + "="*60)
        print("MATCHING REPORT")
        print("="*60)
        
        exact = sum(1 for r in self.match_results if r['match_type'] == 'exact')
        fuzzy = sum(1 for r in self.match_results if r['match_type'] == 'fuzzy')
        no_match = sum(1 for r in self.match_results if r['match_type'] == 'no_match')
        
        print(f"Total Records: {len(self.match_results)}")
        print(f"  - Exact Matches: {exact}")
        print(f"  - Fuzzy Matches: {fuzzy}")
        print(f"  - No Match: {no_match}")
        print()
        
        if no_match > 0:
            print("Unmatched records:")
            for result in self.match_results:
                if result['match_type'] == 'no_match':
                    print(f"  - {result['form_name']}")
        
        print("="*60 + "\n")


def load_config_file(config_path: str) -> Dict[str, Any]:
    """Load configuration from JSON file."""
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading config file: {e}")
        raise


def main():
    parser = argparse.ArgumentParser(
        description='Fill Excel form template with data from source Excel file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_form_filler.py source.xlsx template.xlsx output.xlsx
  python excel_form_filler.py source.xlsx template.xlsx output.xlsx --config custom_config.json
        """
    )
    
    parser.add_argument('source_file', help='Source Excel file with data')
    parser.add_argument('template_file', help='Template Excel file (form to fill)')
    parser.add_argument('output_file', help='Output Excel file path')
    parser.add_argument('--config', help='Optional JSON config file for field mapping')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose logging')
    
    args = parser.parse_args()
    
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    
    # Validate input files
    for file in [args.source_file, args.template_file]:
        if not Path(file).exists():
            logger.error(f"File not found: {file}")
            sys.exit(1)
    
    # Load configuration
    config = None
    if args.config:
        config = load_config_file(args.config)
        logger.info(f"Loaded custom config from: {args.config}")
    
    # Execute
    try:
        filler = ExcelFormFiller(config)
        logger.info("Starting form fill process...")
        
        filler.load_source_data(args.source_file)
        filler.load_template(args.template_file)
        filled_count = filler.fill_form(args.output_file)
        filler.print_match_report()
        
        logger.info(f"✓ Successfully filled {filled_count} records")
        print(f"\n✓ Output file created: {args.output_file}")
        
    except Exception as e:
        logger.error(f"✗ Process failed: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()