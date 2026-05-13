import pandas as pd
from openpyxl import load_workbook
import logging
import re
import argparse
import sys
from pathlib import Path
from validation import discrepancy_status, NOT_ON_CENSUS_STATUS

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class DynamicCensusFiller:
    def __init__(self):
        self.source_lookup = {}
        self.target_columns = {
            'plan': 11,    # Column K: Current Plan Description
            'premium': 12  # Column L: Monthly Total Premium
        }
        self.discrepancy_column = None
        self.census_coverage_column = None
    
    def normalize_name(self, name):
        """Intelligent name normalization for robust matching."""
        if not name or pd.isna(name): return ""
        
        s = str(name).lower().strip()
        
        # Handle "Last, First" format
        if ',' in s:
            parts = [p.strip() for p in s.split(',')]
            if len(parts) >= 2:
                s = f"{parts[1]} {parts[0]}"
                
        # Remove punctuation
        s = re.sub(r"[^a-z\s]", " ", s)
        s = re.sub(r"\s+", " ", s)
        
        parts = s.split()
        
        # Strip known suffixes
        strip_tokens = {'jr', 'sr', 'ii', 'iii', 'iv', 'v', 'esq', 'phd', 'md', 'dds', 'mr', 'mrs', 'ms', 'dr', 'prof'}
        
        cleaned_parts = []
        for p in parts:
            if p in strip_tokens:
                continue
            # Drop single-letter token (middle initial) only if name has 3+ tokens
            if len(p) == 1 and len(parts) >= 3:
                continue
            cleaned_parts.append(p)
            
        # Return sorted tokens so that FIRST LAST and LAST FIRST match exactly
        return " ".join(sorted(cleaned_parts))

    def is_valid_employee_name(self, raw_full_name):
        """Skip summary/total rows from source files."""
        text = str(raw_full_name or "").strip().lower()
        if not text:
            return False
        blocked_terms = ("total", "summary", "record", "employee details", "employer health plan")
        return not any(term in text for term in blocked_terms)

    def find_source_columns(self, df):
        """Dynamically identify columns in the source data based on keywords."""
        cols = {
            'first': None,
            'last': None,
            'full': None,
            'coverage': None,
            'plan': None,
            'premium': None
        }
        
        for col in df.columns:
            c = str(col).lower()
            if 'first' in c and 'name' in c: cols['first'] = col
            elif 'last' in c and 'name' in c: cols['last'] = col
            elif 'full' in c and 'name' in c: cols['full'] = col
            elif 'coverage' in c: cols['coverage'] = col
            elif 'plan' in c and ('name' in c or 'desc' in c): cols['plan'] = col
            elif 'premium' in c or 'total' in c or 'amt' in c: cols['premium'] = col
            
        return cols

    def load_source(self, source_path):
        """Loads source Excel and organizes data into a searchable lookup."""
        try:
            # Try to find the correct sheet
            xl = pd.ExcelFile(source_path)
            sheet_name = None
            for s in xl.sheet_names:
                if 'employee' in s.lower() or 'detail' in s.lower() or 'data' in s.lower():
                    sheet_name = s
                    break
            
            if not sheet_name:
                sheet_name = xl.sheet_names[0]
            
            # Read data, trying to find the header row automatically
            # We look for a row that contains 'name' or 'plan'
            df_probe = pd.read_excel(source_path, sheet_name=sheet_name, nrows=10, header=None)
            header_row = 0
            for i, row in df_probe.iterrows():
                row_str = " ".join([str(x).lower() for x in row if pd.notna(x)])
                if 'name' in row_str or 'plan' in row_str:
                    header_row = i
                    break
            
            df = pd.read_excel(source_path, sheet_name=sheet_name, skiprows=header_row)
            logger.info(f"Loaded source sheet '{sheet_name}' with {len(df)} rows.")
            
            cols = self.find_source_columns(df)
            
            for _, row in df.iterrows():
                # Get name
                name_key = ""
                if cols['full'] and pd.notna(row[cols['full']]):
                    name_key = self.normalize_name(row[cols['full']])
                elif cols['first'] and cols['last']:
                    name_key = self.normalize_name(f"{row[cols['first']]} {row[cols['last']]}")
                
                if not name_key: continue

                # Get plan and premium
                plan = row[cols['plan']] if cols['plan'] else None
                premium = row[cols['premium']] if cols['premium'] else None
                
                # Clean premium
                if isinstance(premium, str):
                    premium = re.sub(r'[^\d.]', '', premium)
                    try: premium = float(premium)
                    except: pass
                
                raw_full_name = ""
                if cols['full'] and pd.notna(row[cols['full']]):
                    raw_full_name = str(row[cols['full']]).strip()
                elif cols['first'] and cols['last']:
                    raw_full_name = f"{row[cols['first']]} {row[cols['last']]}".strip()

                if not self.is_valid_employee_name(raw_full_name):
                    continue

                self.source_lookup[name_key] = {
                    'plan': plan,
                    'premium': premium,
                    'raw_name': raw_full_name,
                    'coverage': row[cols['coverage']] if cols['coverage'] else None
                }
            
            return True
        except Exception as e:
            logger.error(f"Failed to load source: {e}")
            return False

    def find_discrepancy_column(self, ws):
        """Find the Discrepancies column based on header text."""
        for r in range(1, 40):
            for c in range(1, ws.max_column + 1):
                value = ws.cell(row=r, column=c).value
                if value and 'discrep' in str(value).strip().lower():
                    return c
        return None

    def find_coverage_column(self, ws):
        """Find the census coverage column based on header text."""
        for r in range(1, 40):
            for c in range(1, ws.max_column + 1):
                value = ws.cell(row=r, column=c).value
                if not value:
                    continue
                header = str(value).strip().lower()
                if 'coverage' in header:
                    return c
        return 7

    def fill_template(self, template_path, output_path):
        """Fills the template form dynamically."""
        try:
            wb = load_workbook(template_path)
            # Find the Census sheet
            ws = None
            for s in wb.sheetnames:
                if 'census' in s.lower():
                    ws = wb[s]
                    break
            
            if not ws:
                logger.error("Could not find a 'Census' sheet in the template.")
                return False

            # Find data start (looking for headers like 'EE Row' or 'First Name')
            start_row = 1
            for r in range(1, 40):
                row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, 5)]
                if 'ee row' in row_vals or 'first name' in row_vals:
                    start_row = r + 1
                    break
            
            if start_row == 1: start_row = 22 # Default fallback
            
            self.discrepancy_column = self.find_discrepancy_column(ws)
            self.census_coverage_column = self.find_coverage_column(ws)
            if self.discrepancy_column:
                logger.info(f"Using Discrepancies column at index {self.discrepancy_column}.")
            else:
                # Append a new Discrepancies column after the last used column
                self.discrepancy_column = ws.max_column + 1
                from openpyxl.styles import Font, PatternFill, Alignment
                hdr_cell = ws.cell(row=start_row - 1, column=self.discrepancy_column)
                hdr_cell.value = "Discrepancies"
                hdr_cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
                hdr_cell.fill = PatternFill('solid', start_color='4472C4')
                hdr_cell.alignment = Alignment(horizontal='center', vertical='center')
                letter = hdr_cell.column_letter
                ws.column_dimensions[letter].width = 30
                logger.info(f"Appended new Discrepancies column at index {self.discrepancy_column}.")
            logger.info(f"Using census coverage column at index {self.census_coverage_column}.")


            filled_count = 0
            validated_count = 0
            appended_count = 0
            seen_census_names = set()
            last_data_row = start_row - 1
            for row_idx in range(start_row, ws.max_row + 1):
                # Target names are always in columns B (2) and C (3)
                first = ws.cell(row=row_idx, column=2).value
                last = ws.cell(row=row_idx, column=3).value
                
                if not first and not last:
                    if ws.cell(row=row_idx, column=1).value is None: break
                    continue
                
                norm_name = self.normalize_name(f"{first} {last}")
                seen_census_names.add(norm_name)
                last_data_row = row_idx
                
                if norm_name in self.source_lookup:
                    data = self.source_lookup[norm_name]
                    extracted_full_name = f"{first} {last}".strip()
                    extracted_coverage_tier = ws.cell(
                        row=row_idx, column=self.census_coverage_column
                    ).value
                    status = discrepancy_status(
                        extracted_name=extracted_full_name,
                        invoice_name=data.get('raw_name', extracted_full_name),
                        extracted_coverage_tier=extracted_coverage_tier,
                        invoice_coverage_tier=data.get('coverage'),
                        name_is_matched=True
                    )

                    ws.cell(row=row_idx, column=self.target_columns['plan']).value = data['plan']
                    ws.cell(row=row_idx, column=self.target_columns['premium']).value = data['premium']

                    if self.discrepancy_column:
                        ws.cell(row=row_idx, column=self.discrepancy_column).value = status
                        validated_count += 1

                    filled_count += 1
                    logger.info(f"Matched & Filled: {first} {last}")
                else:
                    if self.discrepancy_column:
                        ws.cell(row=row_idx, column=self.discrepancy_column).value = 'not available on invoice'

            append_row = last_data_row
            for source_name_key, data in self.source_lookup.items():
                if source_name_key in seen_census_names:
                    continue

                append_row += 1
                raw_name = str(data.get('raw_name') or "").strip()
                name_parts = raw_name.split()
                first_name = name_parts[0] if name_parts else ""
                last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""

                ws.cell(row=append_row, column=2).value = first_name
                ws.cell(row=append_row, column=3).value = last_name

                if self.census_coverage_column:
                    ws.cell(row=append_row, column=self.census_coverage_column).value = data.get('coverage')

                ws.cell(row=append_row, column=self.target_columns['plan']).value = data.get('plan')
                ws.cell(row=append_row, column=self.target_columns['premium']).value = data.get('premium')

                if self.discrepancy_column:
                    ws.cell(row=append_row, column=self.discrepancy_column).value = NOT_ON_CENSUS_STATUS

                appended_count += 1
                
            wb.save(output_path)
            logger.info(
                f"Done! Saved to {output_path}. Total records filled: {filled_count}. "
                f"Total records validated: {validated_count}. "
                f"Appended source-only records: {appended_count}"
            )
            return True
            
        except Exception as e:
            logger.error(f"Failed to fill template: {e}")
            return False

def main():
    parser = argparse.ArgumentParser(description='Fully Dynamic Insurance Census Filler')
    parser.add_argument('source', help='Path to source Excel data')
    parser.add_argument('template', help='Path to Census template Excel')
    parser.add_argument('output', nargs='?', default='filled_output.xlsx', help='Output filename')
    
    args = parser.parse_args()
    
    filler = DynamicCensusFiller()
    if filler.load_source(args.source):
        if not filler.fill_template(args.template, args.output):
            sys.exit(1)
    else:
        sys.exit(1)

if __name__ == "__main__":
    main()
