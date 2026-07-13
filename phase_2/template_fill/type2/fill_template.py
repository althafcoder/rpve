import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
import re
import argparse
import sys
import os
import json
from pathlib import Path
from openai import OpenAI
from dotenv import load_dotenv
from validation import discrepancy_status, NOT_ON_CENSUS_STATUS

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
_HEADER_FONT      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
_HEADER_FILL      = PatternFill('solid', start_color='4472C4')
_CELL_FONT        = Font(name='Arial', size=10)
_CENTER           = Alignment(horizontal='center', vertical='center')
_LEFT             = Alignment(horizontal='left',   vertical='center')
_THIN             = Side(style='thin', color='D9D9D9')
_BORDER           = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_CORRECT     = PatternFill('solid', start_color='C6EFCE')   # green
_FILL_NOT_MATCH   = PatternFill('solid', start_color='FFC7CE')   # red
_FILL_NOT_INVOICE = PatternFill('solid', start_color='FFEB9C')   # yellow


class DynamicCensusFiller:
    def __init__(self):
        self.source_lookup = {}

        # Detected at runtime
        self.census_name_mode    = None   # 'split' | 'full'
        self.census_name_col     = None   # col index for full-name mode
        self.census_first_col    = None   # col index for split mode
        self.census_last_col     = None   # col index for split mode
        self.census_coverage_col = None
        self.plan_col            = None
        self.premium_col         = None
        self.discrepancy_col     = None
        self.header_row          = None
        self.data_start_row      = None

    # ------------------------------------------------------------------
    # Name normalisation
    # ------------------------------------------------------------------
    def normalize_name(self, name):
        """Intelligent name normalization for robust matching using sorted tokens."""
        if not name or (isinstance(name, float) and pd.isna(name)): return ""

        s = str(name).lower().strip()

        # Handle "Last, First" format
        if ',' in s:
            parts = [p.strip() for p in s.split(',')]
            if len(parts) >= 2:
                s = f"{parts[1]} {parts[0]}"

        # Remove punctuation
        s = re.sub(r"[^a-z\s]", " ", s)
        s = re.sub(r"\s+", " ", s)

        parts = s.split()

        # Strip known suffixes
        strip_tokens = {'jr', 'sr', 'ii', 'iii', 'iv', 'v', 'esq', 'phd', 'md', 'dds', 'mr', 'mrs', 'ms', 'dr', 'prof'}

        cleaned_parts = []
        for p in parts:
            if p in strip_tokens:
                continue
            # Drop single-letter token (middle initial) only if name has 3+ tokens
            if len(p) == 1 and len(parts) >= 3:
                continue
            cleaned_parts.append(p)

        # Return sorted tokens so that FIRST LAST and LAST FIRST match exactly
        return " ".join(sorted(cleaned_parts))

    def is_valid_employee_name(self, raw_full_name):
        """Skip summary/total rows from source files."""
        text = str(raw_full_name or "").strip().lower()
        if not text:
            return False
        blocked_terms = ("total", "summary", "record", "employee details", "employer health plan")
        return not any(term in text for term in blocked_terms)

    # ------------------------------------------------------------------
    # LLM Dynamic Mapping
    # ------------------------------------------------------------------
    def _detect_columns_llm(self, df: pd.DataFrame, context: str = "census") -> dict:
        """Uses LLM to dynamically map unpredictable column headers to standard fields."""
        load_dotenv()
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            logger.warning("OPENAI_API_KEY missing. Skipping LLM column detection.")
            return {}

        client = OpenAI(api_key=api_key)
        headers = list(df.columns)
        sample_data = df.head(3).to_dict(orient='records')

        if context == "source":
            target_fields = (
                "- full_name: The employee or subscriber's full name.\n"
                "- first_name: First name (if separate).\n"
                "- last_name: Last name (if separate).\n"
                "- coverage: Coverage tier or level (e.g., EE, ES, FAM, EC).\n"
                "- plan_name: The detailed plan/product name (e.g., 'UnitedHealthcare Choice Plus HDHP 5000'). Look for columns like 'Coverage Option', 'Plan Name', 'Plan Description', 'Product Name'.\n"
                "- plan_type: The plan CATEGORY/TYPE (e.g., 'Medical', 'Dental', 'Vision', 'Life', 'Choice Plus'). Look for columns like 'Coverage Type', 'Benefit Type', 'Plan Type', 'Product Type', 'Insurance Type'.\n"
                "- premium: The billed premium amount."
            )
        else:
            target_fields = (
                "- full_name: The employee's full name.\n"
                "- first_name: First name (if separate).\n"
                "- last_name: Last name (if separate).\n"
                "- coverage: Coverage tier (EE, ES, FAM).\n"
                "- plan_name: Existing plan column (if any).\n"
                "- premium: Existing premium column (if any).\n"
                "- discrepancy: Existing discrepancies column (if any)."
            )

        system_prompt = (
            f"You are a highly accurate data mapping expert. Map the provided column headers from a {context} file "
            "to our standard internal fields. Use the sample data to understand the content.\n\n"
            f"Standard Fields:\n{target_fields}\n\n"
            "CRITICAL RULES:\n"
            "1. Return ONLY valid JSON mapping our Standard Fields to EXACT column headers.\n"
            "2. IGNORE emergency contact or payroll metadata headers."
        )

        user_prompt = f"Headers:\n{headers}\n\nSample:\n{json.dumps(sample_data, indent=2, default=str)}"

        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                response_format={"type": "json_object"},
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                temperature=0.0
            )
            mapping = json.loads(response.choices[0].message.content)
            valid = {k: v for k, v in mapping.items() if v in headers}
            logger.info(f"LLM Dynamic Mapping ({context}): {valid}")
            return valid
        except Exception as e:
            logger.error(f"LLM column detection failed ({context}): {e}")
            return {}

    def _classify_plan_types_batch(self, plan_names: list) -> dict:
        """
        Uses LLM to classify plan names into categories (Medical, Dental, Vision, Life, LTD, Other).
        Only called when plan_type column doesn't exist in source file.
        
        Args:
            plan_names: List of unique plan names to classify
            
        Returns:
            Dictionary mapping plan name to category, e.g. {"Kaiser HMO": "Medical"}
        """
        if not plan_names:
            return {}
            
        load_dotenv()
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            logger.warning("OPENAI_API_KEY missing. Cannot classify plan types.")
            return {}

        try:
            client = OpenAI(api_key=api_key)
            
            # Create numbered list for LLM
            plan_list = "\n".join([f"{i+1}. {name}" for i, name in enumerate(plan_names)])
            
            system_prompt = (
                "You are a health insurance plan classification expert. "
                "Classify each plan name into EXACTLY ONE category: Medical, Dental, Vision, Life, LTD, or Other.\n\n"
                "Rules:\n"
                "- Medical: Health/medical plans including HMO, PPO, EPO, POS, HDHP, HSA plans, Choice Plus, Select, Open Access, etc.\n"
                "- Dental: Dental plans (even if they contain 'PPO' or 'HMO')\n"
                "- Vision: Vision/eye care plans\n"
                "- Life: Life insurance and AD&D plans\n"
                "- LTD: Long-term disability plans\n"
                "- Other: Everything else (FSA, HSA, commuter, adoption assistance, etc.)\n\n"
                "Return ONLY valid JSON with plan names as keys and categories as values."
            )
            
            user_prompt = (
                f"Classify these insurance plans:\n\n{plan_list}\n\n"
                "Return JSON format: {\"plan name\": \"category\", ...}"
            )
            
            response = client.chat.completions.create(
                model="gpt-4o",
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.0
            )
            
            classifications = json.loads(response.choices[0].message.content)
            logger.info(f"LLM classified {len(classifications)} plan types")
            return classifications
            
        except Exception as e:
            logger.error(f"LLM plan classification failed: {e}")
            return {}

    # ------------------------------------------------------------------
    # Source loading
    # ------------------------------------------------------------------
    def load_source(self, source_path):
        """Load invoice/source Excel and build a name-keyed lookup."""
        try:
            xl = pd.ExcelFile(source_path)
            sheet_name = next(
                (s for s in xl.sheet_names
                 if any(k in s.lower() for k in ('employee', 'detail', 'data'))),
                xl.sheet_names[0]
            )

            # Auto-detect header row
            df_probe = pd.read_excel(source_path, sheet_name=sheet_name, nrows=10, header=None)
            header_row = 0
            for i, row in df_probe.iterrows():
                row_str = " ".join(str(x).lower() for x in row if pd.notna(x))
                if any(kw in row_str for kw in ('name', 'plan', 'premium', 'total cost', 'coverage')):
                    header_row = i
                    break

            df = pd.read_excel(source_path, sheet_name=sheet_name, skiprows=header_row)
            df.columns = df.columns.astype(str)
            logger.info(f"Loaded source sheet '{sheet_name}' with {len(df)} rows.")

            col_map = self._detect_columns_llm(df, "source")

            # Static fallback
            if not col_map:
                for col in df.columns:
                    c = str(col).lower()
                    if   'full' in c and 'name' in c:      col_map['full_name'] = col
                    elif 'employee' in c and 'name' in c: col_map['full_name'] = col
                    elif 'coverage' in c:                 col_map['coverage']  = col
                    elif 'plan' in c:                     col_map['plan_name'] = col
                    elif 'premium' in c or 'cost' in c:   col_map['premium']   = col

            # ═══════════════════════════════════════════════════════════════════
            # HYBRID PLAN TYPE DETECTION
            # ═══════════════════════════════════════════════════════════════════
            plan_type_col = col_map.get('plan_type')
            plan_name_col = col_map.get('plan_name')
            
            # Check if we need LLM classification (when plan_type column missing)
            plan_classifications = {}
            if not plan_type_col and plan_name_col:
                # Extract unique plan names for batch classification
                unique_plans = df[plan_name_col].dropna().unique().tolist()
                if unique_plans:
                    logger.info(f"No plan_type column found. Classifying {len(unique_plans)} unique plans via LLM...")
                    plan_classifications = self._classify_plan_types_batch(unique_plans)
            elif plan_type_col:
                logger.info(f"Found plan_type column: '{plan_type_col}'. Will use direct values.")

            # Clean whitespace and replace empty/whitespace strings with NaN so ffill works (non-destructive add-on)
            for col_key in ['full_name', 'first_name', 'last_name']:
                col_name = col_map.get(col_key)
                if col_name and col_name in df.columns:
                    cleaned_series = df[col_name].astype(str).str.strip().replace({'': None, 'nan': None, 'None': None, '<NA>': None})
                    df[col_name] = cleaned_series.ffill()

            for _, row in df.iterrows():
                f_col = col_map.get('full_name')
                first_col = col_map.get('first_name')
                last_col = col_map.get('last_name')

                raw_full_name = ""
                if f_col and pd.notna(row.get(f_col)):
                    raw_full_name = str(row[f_col]).strip()
                else:
                    name_parts = []
                    if first_col and pd.notna(row.get(first_col)):
                        name_parts.append(str(row[first_col]).strip())
                    if last_col and pd.notna(row.get(last_col)):
                        name_parts.append(str(row[last_col]).strip())
                    raw_full_name = " ".join(name_parts).strip()
                if not raw_full_name or not self.is_valid_employee_name(raw_full_name):
                    continue

                name_key = self.normalize_name(raw_full_name)
                
                premium = row.get(col_map.get('premium')) if col_map.get('premium') else 0.0
                if isinstance(premium, str):
                    premium = re.sub(r'[^\d.]', '', premium)
                    try: premium = float(premium)
                    except: premium = 0.0
                elif pd.isna(premium):
                    premium = 0.0

                # ═══════════════════════════════════════════════════════════════
                # DETERMINE IF THIS IS A MEDICAL PLAN (Hybrid Approach)
                # ═══════════════════════════════════════════════════════════════
                is_medical = False
                
                # Priority 1: Use explicit plan_type column if available
                if plan_type_col and pd.notna(row.get(plan_type_col)):
                    plan_category = str(row[plan_type_col]).lower().strip()
                    
                    # Check for medical indicators in category
                    is_medical = any(kw in plan_category for kw in [
                        'medical', 'health', 'choice plus', 'choice', 'hmo', 'ppo', 
                        'epo', 'pos', 'hdhp', 'hsa', 'kaiser', 'anthem', 'aetna', 
                        'cigna', 'united', 'blue cross', 'bcbs'
                    ])
                    
                # Priority 2: Use LLM classification (when plan_type column missing)
                elif plan_classifications and plan_name_col:
                    plan_name = row.get(plan_name_col)
                    if plan_name and str(plan_name) in plan_classifications:
                        classification = plan_classifications.get(str(plan_name))
                        is_medical = (classification == "Medical")
                
                # Priority 3: Fallback to expanded keyword matching
                else:
                    plan_name = row.get(plan_name_col) if plan_name_col else None
                    if plan_name:
                        plan_str = str(plan_name).lower()
                        is_medical = any(kw in plan_str for kw in [
                            'medical', 'health', 'hmo', 'ppo', 'epo', 'pos', 'hdhp', 
                            'choice plus', 'choice', 'select', 'advantage', 'open access',
                            'kaiser', 'anthem', 'aetna', 'cigna', 'united', 'uhc',
                            'blue cross', 'blue shield', 'bcbs', 'bc-', 'hs-',
                            'hsa', 'fsa health'
                        ])
                
                # ═══════════════════════════════════════════════════════════════
                # BUSINESS RULE: Medical Premium Validation
                # Medical plans are ALWAYS >= $200 in our business data.
                # Any plan with premium < $200 is NOT Medical (it's Dental/Vision/Life
                # or misclassified data). This prevents false positives.
                # ═══════════════════════════════════════════════════════════════
                if is_medical and premium < 200:
                    logger.debug(f"Plan classified as Medical but premium ${premium:.2f} < $200 → treating as non-Medical (likely Dental/Vision/Life)")
                    is_medical = False

                # If we already have an entry for this person, keep the one with highest premium
                # OR keep Medical over non-Medical (Dental/Vision/Life)
                if name_key in self.source_lookup:
                    existing_premium = self.source_lookup[name_key]['premium']
                    existing_plan = str(self.source_lookup[name_key]['plan']).lower() if self.source_lookup[name_key]['plan'] else ''
                    
                    # Check if existing entry is medical (may have been classified earlier)
                    existing_is_medical = self.source_lookup[name_key].get('is_medical', False)
                    
                    # Keep current entry if: (1) it's Medical and existing is not, OR (2) both same type but current has higher premium
                    if (is_medical and not existing_is_medical) or (is_medical == existing_is_medical and premium > existing_premium):
                        pass  # Will overwrite below
                    else:
                        continue  # Keep existing entry

                # ═══════════════════════════════════════════════════════════════
                # FINAL FILTER: Only keep Medical plans (>= $200) OR high-value non-Medical (>= $50)
                # ═══════════════════════════════════════════════════════════════
                # Rule 1: If it's Medical, premium must be >= $200 (already validated above, but double-check)
                # Rule 2: If it's NOT Medical, only keep if premium >= $200 (to avoid low-cost misclassified plans)
                # Rule 3: Exception - keep non-Medical Dental/Vision if premium >= $50 AND they're not misclassified Medical
                
                # Simple rule: Only keep if premium >= $200 (whether Medical or not)
                # This ensures we only get high-value plans that are definitely correct
                if premium < 200:
                    # Skip low-premium plans (whether Medical or not)
                    # These are likely misclassified or ancillary plans
                    continue

                self.source_lookup[name_key] = {
                    'plan':       row.get(plan_name_col) if plan_name_col else None,
                    'premium':    premium,
                    'raw_name':   raw_full_name,
                    'coverage':   row.get(col_map.get('coverage')) if col_map.get('coverage') else None,
                    'is_medical': is_medical,  # Store for priority comparison
                }

            logger.info(f"Source lookup built: {len(self.source_lookup)} entries (Medical prioritized, premium >= $50 filter).")
            return True

        except Exception as e:
            logger.error(f"Failed to load source: {e}")
            return False

    # ------------------------------------------------------------------
    # Template filling
    # ------------------------------------------------------------------
    def _write_header(self, ws, row, col, label):
        cell = ws.cell(row=row, column=col)
        cell.value     = label
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER

    def _detect_census_structure(self, ws):
        """Scan the sheet for its header row and map all relevant columns."""
        COLUMN_KW = ('first', 'last', 'plan', 'premium', 'coverage', 'discrep', 'birth', 'gender', 'name')
        df_full = pd.DataFrame([[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, 30)])

        for r in range(1, 30):
            row_vals = [str(ws.cell(row=r, column=c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
            joined = " ".join(row_vals)
            hit_count = sum(1 for k in COLUMN_KW if k in joined)
            if hit_count < 2: continue

            self.header_row     = r
            self.data_start_row = r + 1

            df_headers = pd.DataFrame([row_vals], columns=[f"col_{i}" for i in range(1, len(row_vals)+1)])
            # Simulate real headers for LLM
            df_headers.columns = [str(ws.cell(r, c).value or f"Unnamed_{c}") for c in range(1, ws.max_column+1)]

            col_map = self._detect_columns_llm(df_headers, "census")

            # Map back to 1-based indices
            header_to_idx = {str(ws.cell(r, c).value).strip().lower(): c for c in range(1, ws.max_column+1) if ws.cell(r,c).value}

            def get_idx(field):
                h = col_map.get(field)
                return header_to_idx.get(str(h).strip().lower()) if h else None

            name_col = get_idx('full_name')
            f_col = get_idx('first_name')
            l_col = get_idx('last_name')

            if name_col:
                self.census_name_mode, self.census_name_col = 'full', name_col
            elif f_col and l_col:
                self.census_name_mode, self.census_first_col, self.census_last_col = 'split', f_col, l_col
            else:
                self.census_name_mode, self.census_name_col = 'full', 1

            self.census_coverage_col = get_idx('coverage')
            self.plan_col            = get_idx('plan_name')
            self.premium_col         = get_idx('premium')
            self.discrepancy_col     = get_idx('discrepancy')

            # Append missing columns
            max_col = ws.max_column
            if not self.plan_col:
                self.plan_col = max_col + 1
                self._write_header(ws, r, self.plan_col, 'Plan Name')
                max_col += 1
            if not self.premium_col:
                self.premium_col = max_col + 1
                self._write_header(ws, r, self.premium_col, 'Monthly Premium')
                max_col += 1
            if not self.discrepancy_col:
                self.discrepancy_col = max_col + 1
                self._write_header(ws, r, self.discrepancy_col, 'Discrepancies')

            # Column widths
            from openpyxl.utils import get_column_letter
            for col_idx, width in [(self.plan_col, 22), (self.premium_col, 18), (self.discrepancy_col, 15)]:
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            logger.info(f"Census structure — mode: {self.census_name_mode}, header_row: {r}")
            return True

        return False

    def fill_template(self, template_path, output_path):
        """Fill plan, premium and discrepancy columns in the census template."""
        try:
            wb = load_workbook(template_path)
            ws = next((wb[s] for s in wb.sheetnames if any(k in s.lower() for k in ('census', 'table', 'employee'))), wb.active)
            if not self._detect_census_structure(ws): return False

            filled_count = not_found_count = 0
            seen_census_names = set()
            last_data_row = self.data_start_row - 1

            for row_idx in range(self.data_start_row, ws.max_row + 1):
                if self.census_name_mode == 'full':
                    raw_name = ws.cell(row=row_idx, column=self.census_name_col).value
                    if not raw_name: break
                    emp_display = str(raw_name).strip()
                else:
                    f, l = ws.cell(row_idx, self.census_first_col).value, ws.cell(row_idx, self.census_last_col).value
                    if not f and not l: break
                    emp_display = f"{f or ''} {l or ''}".strip()

                norm_name = self.normalize_name(emp_display)
                seen_census_names.add(norm_name)
                last_data_row = row_idx

                plan_cell, prem_cell, disc_cell = [ws.cell(row=row_idx, column=c) for c in (self.plan_col, self.premium_col, self.discrepancy_col)]
                for cell in (plan_cell, prem_cell, disc_cell):
                    cell.font, cell.border = _CELL_FONT, _BORDER
                plan_cell.alignment, prem_cell.alignment, disc_cell.alignment = _LEFT, _CENTER, _CENTER

                # ── WAIVER ONLY (WO) SKIP ───────────────────────────────────
                # If coverage is 'WO' (Waiver Only), no need to fill.
                # Leave Plan, Premium, and Discrepancy/Notes columns empty/blank.
                if self.census_coverage_col is not None:
                    cov_val = ws.cell(row=row_idx, column=self.census_coverage_col).value
                    if cov_val is not None and str(cov_val).strip().upper() == 'WO':
                        logger.info(f"  Skipping waiver row {row_idx}: {emp_display} (coverage='WO')")
                        plan_cell.value = ''
                        prem_cell.value = ''
                        disc_cell.value = ''
                        continue

                if norm_name in self.source_lookup:
                    data = self.source_lookup[norm_name]
                    coverage = ws.cell(row_idx, self.census_coverage_col).value if self.census_coverage_col else None
                    status = discrepancy_status(emp_display, data['raw_name'], coverage, data['coverage'], True)
                    plan_cell.value = data['plan']
                    if data['premium'] is not None:
                        prem_cell.value, prem_cell.number_format = data['premium'], '$#,##0.00'
                    disc_cell.value, disc_cell.fill = status, (_FILL_CORRECT if status == 'Correct' else _FILL_NOT_MATCH)
                    filled_count += 1
                else:
                    disc_cell.value, disc_cell.fill = 'not available on invoice', _FILL_NOT_INVOICE
                    not_found_count += 1

            # Append source-only
            app_row = last_data_row
            for key, data in self.source_lookup.items():
                if key in seen_census_names: continue
                app_row += 1
                if self.census_name_mode == 'full': ws.cell(app_row, self.census_name_col).value = data['raw_name']
                else:
                    pts = str(data['raw_name']).split()
                    ws.cell(app_row, self.census_first_col).value, ws.cell(app_row, self.census_last_col).value = pts[0], " ".join(pts[1:])
                if self.census_coverage_col: ws.cell(app_row, self.census_coverage_col).value = data['coverage']
                ws.cell(app_row, self.plan_col).value = data['plan']
                if data['premium']:
                    c = ws.cell(app_row, self.premium_col)
                    c.value, c.number_format = data['premium'], '$#,##0.00'
                c = ws.cell(app_row, self.discrepancy_col)
                c.value, c.fill = NOT_ON_CENSUS_STATUS, _FILL_NOT_MATCH

            wb.save(output_path)
            logger.info(f"Done! Filled: {filled_count} | Not in Source: {not_found_count}")
            return True
        except Exception as e:
            logger.error(f"Failed to fill template: {e}")
            return False

def main():
    parser = argparse.ArgumentParser(description='Dynamic Insurance Census Filler')
    parser.add_argument('source',   help='Path to source/invoice Excel file')
    parser.add_argument('template', help='Path to census template Excel file')
    parser.add_argument('output', nargs='?', default='filled_output.xlsx')
    args = parser.parse_args()
    filler = DynamicCensusFiller()
    if filler.load_source(args.source):
        if not filler.fill_template(args.template, args.output): sys.exit(1)
    else: sys.exit(1)

if __name__ == "__main__":
    main()