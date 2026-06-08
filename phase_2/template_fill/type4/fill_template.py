"""
fill_template.py  —  Technology 2: openpyxl (+ pandas for census reading)
==========================================================================
Fill the RAPT Census xlsx template using data extracted from a CA Choice
PDF invoice (via pdf_extractor.py).

Usage:
    python fill_template.py <invoice.pdf> <rapt_template.xlsx> [output.xlsx]

Inputs
------
1. invoice.pdf       – California Choice PDF invoice
2. rapt_template.xlsx – Empty RAPT Census template to fill

Output
------
Filled RAPT Census xlsx with:
  - Demographic data from the existing rows in the template (kept as-is)
  - Plan Name and Monthly Total Premium columns filled from the PDF invoice
  - Discrepancies column filled with match results
  - Rows on the invoice but NOT in the template appended at the bottom

Technologies used
-----------------
- pdf_extractor.py  (Technology 1: pdfplumber) — called to get invoice data
- openpyxl           (Technology 2)             — reads/writes the xlsx template
- pandas             (Technology 2 support)     — reads census if needed
"""

import re
import sys
import logging
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Technology 1: PDF extractor (separate file — not connected, standalone)
from pdf_extractor import extract_employees_from_pdf

from validation import (
    discrepancy_status,
    NOT_ON_CENSUS_STATUS,
    canonical_coverage_tier,
    normalize_text,
)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants  (same palette as type2/type3 for visual consistency)
# ---------------------------------------------------------------------------
_HEADER_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL    = PatternFill("solid", start_color="4472C4")
_CELL_FONT      = Font(name="Arial", size=10)
_CENTER         = Alignment(horizontal="center", vertical="center")
_LEFT           = Alignment(horizontal="left",   vertical="center")
_THIN           = Side(style="thin", color="D9D9D9")
_BORDER         = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_CORRECT   = PatternFill("solid", start_color="C6EFCE")   # green
_FILL_MISMATCH  = PatternFill("solid", start_color="FFC7CE")   # red
_FILL_MISSING   = PatternFill("solid", start_color="FFEB9C")   # yellow

# ---------------------------------------------------------------------------
# Name helpers (kept self-contained — Technology 2 does NOT call Technology 1
# for name logic; they share only validation.py)
# ---------------------------------------------------------------------------
_SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v", "esq"}


def _clean_name(name) -> str:
    if not name or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).lower().strip()
    if "," in s:
        parts = [p.strip() for p in s.split(",")]
        s = f"{parts[1]} {parts[0]}" if len(parts) >= 2 else s
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _tokens(name) -> list:
    raw = _clean_name(name)
    parts = [p for p in raw.split() if p not in _SUFFIXES]
    if not parts:
        return []
    if len(parts) >= 3:
        result = [parts[0]]
        for p in parts[1:-1]:
            if len(p) == 1:
                continue  # drop middle initial
            result.append(p)
        result.append(parts[-1])
        return result
    return parts


def _make_key(name) -> str:
    return " ".join(_tokens(name))


def _lookup_keys(name) -> list:
    """Return canonical key + reversed key for 2-token names."""
    toks = _tokens(name)
    if not toks:
        return []
    key = " ".join(toks)
    keys = [key]
    if len(toks) == 2 and toks[0] != toks[1]:
        keys.append(f"{toks[1]} {toks[0]}")
    return keys


# ---------------------------------------------------------------------------
# RAPT column detection  (Technology 2: openpyxl)
# ---------------------------------------------------------------------------
def _col_key(val) -> str:
    return re.sub(r"\s+", " ", str(val or "").replace("*", "").strip().lower())


def detect_rapt_columns(ws) -> dict:
    """
    Scan the sheet for its header row and map relevant column indices.
    Returns a dict with keys: header_row, data_start, first, last, gender,
    dob, zip, relation, dep_of, coverage, cobra, plan, premium, disc.
    """
    for r in range(1, 30):
        row_keys = {
            c: _col_key(ws.cell(row=r, column=c).value)
            for c in range(1, ws.max_column + 1)
        }
        if "first name" not in row_keys.values():
            continue

        cols = {"header_row": r, "data_start": r + 1}
        for c, key in row_keys.items():
            if   key == "data row":                         cols["data_row"] = c
            elif key == "first name":                       cols["first"]    = c
            elif key == "last name":                        cols["last"]     = c
            elif "gender" in key:                           cols["gender"]   = c
            elif "date of birth" in key or key == "birth":  cols["dob"]      = c
            elif "zip" in key:                              cols["zip"]      = c
            elif "relationship" in key:                     cols["relation"] = c
            elif "dependent of" in key:                     cols["dep_of"]   = c
            elif "coverage" in key and "cobra" not in key:  cols["coverage"] = c
            elif "cobra" in key:                            cols["cobra"]    = c
            elif "monthly" in key and "premium" in key:    cols["premium"]  = c
            elif "plan" in key:                             cols["plan"]     = c
            elif "discrepanc" in key:                       cols["disc"]     = c

        # Append Discrepancies column if absent
        if "disc" not in cols:
            dc = (cols.get("premium") or ws.max_column) + 1
            cell = ws.cell(row=r, column=dc)
            cell.value     = "Discrepancies"
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border    = _BORDER
            ws.column_dimensions[get_column_letter(dc)].width = 30
            cols["disc"] = dc

        # Append Monthly Total Premium column if absent
        if "premium" not in cols:
            pc = (cols.get("plan") or ws.max_column) + 1
            cell = ws.cell(row=r, column=pc)
            cell.value     = "Monthly Total Premium"
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border    = _BORDER
            ws.column_dimensions[get_column_letter(pc)].width = 22
            cols["premium"] = pc

            # Re-insert disc after premium
            dc2 = pc + 1
            cell2 = ws.cell(row=r, column=dc2)
            cell2.value     = "Discrepancies"
            cell2.font      = _HEADER_FONT
            cell2.fill      = _HEADER_FILL
            cell2.alignment = _CENTER
            cell2.border    = _BORDER
            ws.column_dimensions[get_column_letter(dc2)].width = 30
            cols["disc"] = dc2

        # Column widths for key columns
        for key_name, width in [("plan", 35), ("premium", 22), ("first", 15), ("last", 18)]:
            if key_name in cols:
                ws.column_dimensions[get_column_letter(cols[key_name])].width = width

        logger.info(f"RAPT columns detected: {cols}")
        return cols

    raise ValueError("Could not find 'First Name' header row in template (checked first 30 rows).")


# ---------------------------------------------------------------------------
# Cell writing helpers
# ---------------------------------------------------------------------------
def _wcell(ws, row, col, value, align=None, fmt=None, fill=None):
    if col is None:
        return
    cell           = ws.cell(row=row, column=col)
    cell.value     = value
    cell.font      = _CELL_FONT
    cell.border    = _BORDER
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt
    if fill:  cell.fill          = fill


# ---------------------------------------------------------------------------
# Invoice lookup builder
# ---------------------------------------------------------------------------
def build_invoice_lookup(employees: list) -> dict:
    """
    Convert extracted employee list → dict keyed by all candidate name keys.
    Each value is the employee dict (raw_name, first, last, coverage, plan, premium).
    """
    lookup = {}
    for emp in employees:
        full_name = f"{emp['first']} {emp['last']}"
        for key in _lookup_keys(full_name):
            if key:
                lookup[key] = emp
        # Also index by "last first" as-is from the raw PDF name
        raw_key = _make_key(emp["raw_name"])
        if raw_key and raw_key not in lookup:
            lookup[raw_key] = emp
    logger.info(f"Invoice lookup: {len({v['raw_name'] for v in lookup.values()})} unique employees")
    return lookup


# ---------------------------------------------------------------------------
# Template filling  (Technology 2: openpyxl)
# ---------------------------------------------------------------------------
def fill_template(invoice_lookup: dict, template_path: str, output_path: str) -> bool:
    """
    Fill the RAPT Census xlsx template.

    For each employee row already in the template:
      - Look up by name in invoice_lookup
      - Fill Plan Name, Monthly Total Premium, Discrepancies
    For each invoice employee NOT in the template:
      - Append a new row marked NOT_ON_CENSUS_STATUS
    """
    try:
        wb = load_workbook(template_path)
        ws = next(
            (wb[s] for s in wb.sheetnames
             if any(k in s.lower() for k in ("sheet", "census", "table", "employee"))),
            wb.active,
        )

        cols = detect_rapt_columns(ws)
        data_start = cols["data_start"]

        filled          = 0
        not_on_invoice  = 0
        appended        = 0
        seen_keys       = set()   # canonical keys already written
        last_data_row   = data_start - 1

        # ── Iterate existing template rows ──────────────────────────────────
        for row_idx in range(data_start, ws.max_row + 1):
            first_val = ws.cell(row=row_idx, column=cols.get("first", 2)).value
            last_val  = ws.cell(row=row_idx, column=cols.get("last",  3)).value

            if not first_val and not last_val:
                # Check if truly empty (end of data)
                if ws.cell(row=row_idx, column=1).value is None:
                    break
                continue

            first_str = str(first_val or "").strip()
            last_str  = str(last_val  or "").strip()
            emp_display = f"{first_str} {last_str}".strip()
            last_data_row = row_idx

            # Check relationship — skip plan/premium/disc for dependents
            relation_col = cols.get("relation")
            if relation_col:
                rel = str(ws.cell(row=row_idx, column=relation_col).value or "").strip().lower()
                if rel in ("ch", "sp", "child", "spouse", "dependent", "dep"):
                    # Dependents: clear out plan/premium/disc so no junk remains
                    _wcell(ws, row_idx, cols.get("plan"),    None, _LEFT)
                    _wcell(ws, row_idx, cols.get("premium"), None, _CENTER)
                    _wcell(ws, row_idx, cols.get("disc"),    None, _CENTER)
                    logger.debug(f"  Skipping dependent: {first_str} {last_str}")
                    continue

            # ── WAIVER ONLY (WO) SKIP ───────────────────────────────────────
            # If coverage is 'WO' (Waiver Only), no need to fill.
            # Leave Plan, Premium, and Discrepancy/Notes columns empty/blank.
            coverage_col = cols.get("coverage")
            if coverage_col:
                cov_val = ws.cell(row=row_idx, column=coverage_col).value
                if cov_val is not None and str(cov_val).strip().upper() == 'WO':
                    logger.info(f"  Skipping waiver row {row_idx}: {first_str} {last_str} (coverage='WO')")
                    _wcell(ws, row_idx, cols.get("plan"),    None, _LEFT)
                    _wcell(ws, row_idx, cols.get("premium"), None, _CENTER)
                    _wcell(ws, row_idx, cols.get("disc"),    None, _CENTER)
                    continue

            # Look up in invoice
            inv = None
            for key in _lookup_keys(emp_display):
                if key in invoice_lookup:
                    inv = invoice_lookup[key]
                    seen_keys.add(key)
                    # Mark reversed key as seen too
                    toks = key.split()
                    if len(toks) == 2:
                        seen_keys.add(f"{toks[1]} {toks[0]}")
                    break

            # Census coverage tier (from template row)
            coverage_col = cols.get("coverage")
            census_coverage = (
                ws.cell(row=row_idx, column=coverage_col).value if coverage_col else None
            )

            if inv:
                # Employee is on the invoice — check if they waived coverage (no plan/premium)
                is_waiver = (
                    inv.get("coverage", "").upper() in ("WO", "WP", "NC", "RC", "NE")
                    and not inv.get("plan")
                    and inv.get("premium") is None
                )
                if is_waiver:
                    # Waived: no plan/premium to fill — mark as not available on invoice
                    _wcell(ws, row_idx, cols.get("plan"),    None, _LEFT)
                    _wcell(ws, row_idx, cols.get("premium"), None, _CENTER)
                    _wcell(ws, row_idx, cols.get("disc"), "not available on invoice",
                           _CENTER, fill=_FILL_MISSING)
                    filled += 1
                    logger.info(f"  Matched (waived): {first_str} {last_str}")
                else:
                    status = discrepancy_status(
                        extracted_name          = emp_display,
                        invoice_name            = inv["raw_name"],
                        extracted_coverage_tier = census_coverage,
                        invoice_coverage_tier   = inv["coverage"],
                        name_is_matched         = True,
                    )
                    _wcell(ws, row_idx, cols.get("plan"),    inv["plan"],    _LEFT)
                    if inv["premium"] is not None:
                        _wcell(ws, row_idx, cols.get("premium"), inv["premium"], _CENTER, "$#,##0.00")
                    _wcell(ws, row_idx, cols.get("disc"), status, _CENTER,
                           fill=_FILL_CORRECT if status == "Correct" else _FILL_MISMATCH)
                    filled += 1
                    logger.info(f"  Matched: {first_str} {last_str} → {status}")
            else:
                _wcell(ws, row_idx, cols.get("disc"), "not available on invoice",
                       _CENTER, fill=_FILL_MISSING)
                not_on_invoice += 1
                logger.debug(f"  No match: {emp_display}")

        # ── Append invoice-only rows ────────────────────────────────────────
        # Figure out the next available data-row number
        last_row_num_col = cols.get("data_row")
        next_data_row_num = 1
        if last_row_num_col:
            for r in range(data_start, last_data_row + 1):
                v = ws.cell(row=r, column=last_row_num_col).value
                try:
                    n = int(v)
                    if n > next_data_row_num:
                        next_data_row_num = n
                except (TypeError, ValueError):
                    pass
            next_data_row_num += 1

        append_ws_row = last_data_row + 1

        for inv_key, inv in invoice_lookup.items():
            if inv_key in seen_keys:
                continue
            toks = inv_key.split()
            if len(toks) == 2 and f"{toks[1]} {toks[0]}" in seen_keys:
                continue

            # Mark as seen so we don't write same person twice
            seen_keys.add(inv_key)
            if len(toks) == 2:
                seen_keys.add(f"{toks[1]} {toks[0]}")

            # Build a minimal row for the invoice-only employee
            first_str = inv.get("first", "")
            last_str  = inv.get("last",  "")

            _wcell(ws, append_ws_row, cols.get("data_row"), next_data_row_num, _CENTER)
            _wcell(ws, append_ws_row, cols.get("first"),    first_str,          _LEFT)
            _wcell(ws, append_ws_row, cols.get("last"),     last_str,           _LEFT)
            _wcell(ws, append_ws_row, cols.get("coverage"), inv.get("coverage"), _CENTER)
            _wcell(ws, append_ws_row, cols.get("cobra"),    "N",                _CENTER)
            _wcell(ws, append_ws_row, cols.get("plan"),     inv.get("plan"),    _LEFT)
            if inv.get("premium") is not None:
                _wcell(ws, append_ws_row, cols.get("premium"), inv["premium"], _CENTER, "$#,##0.00")
            _wcell(ws, append_ws_row, cols.get("disc"), NOT_ON_CENSUS_STATUS,
                   _CENTER, fill=_FILL_MISMATCH)

            logger.info(f"  Appended invoice-only: {first_str} {last_str}")
            append_ws_row    += 1
            next_data_row_num += 1
            appended          += 1

        wb.save(output_path)
        logger.info(
            f"\nSaved '{output_path}'.\n"
            f"  Matched: {filled}  |  Not on invoice: {not_on_invoice}  "
            f"|  Appended (invoice-only): {appended}"
        )
        return True

    except Exception as exc:
        logger.error(f"fill_template failed: {exc}", exc_info=True)
        return False


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------
def run(pdf_path: str, template_path: str, output_path: str) -> bool:
    """
    Orchestrate the two separate technologies:
      1. pdf_extractor.py  → extracts employees from PDF
      2. openpyxl          → fills the xlsx template
    The two technologies are NOT connected internally — they communicate only
    through the plain Python list returned by extract_employees_from_pdf().
    """
    # --- Technology 1: pdfplumber (via pdf_extractor.py) ---
    employees = extract_employees_from_pdf(pdf_path)
    if not employees:
        logger.error("No employees extracted from PDF. Aborting.")
        return False

    # Build lookup (Technology 2 helper — pure dict, no pdfplumber here)
    invoice_lookup = build_invoice_lookup(employees)

    # --- Technology 2: openpyxl ---
    return fill_template(invoice_lookup, template_path, output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description=(
            "Type4 RAPT Census Filler — California Choice PDF Invoice\n\n"
            "Technology 1 (pdf_extractor.py): pdfplumber  — parses the CA Choice PDF\n"
            "Technology 2 (fill_template.py): openpyxl    — fills the RAPT Census xlsx"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("invoice",  help="California Choice PDF invoice path")
    parser.add_argument("template", help="RAPT Census xlsx template path")
    parser.add_argument(
        "output",
        nargs="?",
        default="filled_rapt_output_type4.xlsx",
        help="Output xlsx path (default: filled_rapt_output_type4.xlsx)",
    )
    args = parser.parse_args()

    ok = run(args.invoice, args.template, args.output)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
