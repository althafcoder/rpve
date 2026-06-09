"""
fill_template.py  –  3-input RAPT census filler
================================================
Usage:
    python fill_template.py <invoice> <ref_census> <rapt_template> [output]

Inputs
------
1. invoice        – BENEFITS_BILLING xlsx   (plan name, premium, coverage)
2. ref_census     – TEPCensus xlsx          (employee demographics + health plan coverage tier)
3. rapt_template  – RAPT_Census xlsx        (empty RAPT output template to fill)

Output
------
Filled RAPT xlsx with all RAPT columns including Discrepancies.
"""

import re
import logging
import argparse
import os
import json
from collections import defaultdict
from openai import OpenAI
from dotenv import load_dotenv

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from validation import discrepancy_status, NOT_ON_CENSUS_STATUS
import census_normalizer

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
_HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
_HEADER_FILL   = PatternFill('solid', start_color='4472C4')
_CELL_FONT     = Font(name='Arial', size=10)
_CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=False)
_LEFT          = Alignment(horizontal='left',   vertical='center')
_THIN          = Side(style='thin', color='D9D9D9')
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_CORRECT  = PatternFill('solid', start_color='C6EFCE')
_FILL_MISMATCH = PatternFill('solid', start_color='FFC7CE')
_FILL_MISSING  = PatternFill('solid', start_color='FFEB9C')

# Name suffixes that should be ignored when building lookup keys
_SUFFIXES = {'jr', 'sr', 'ii', 'iii', 'iv', 'v', 'esq'}


# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------
def _clean_name(name) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    if not name or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).lower().strip()
    if ',' in s:                          # "Last, First" → "first last"
        parts = [p.strip() for p in s.split(',')]
        s = f"{parts[1]} {parts[0]}" if len(parts) >= 2 else s
    s = re.sub(r'[^a-z0-9\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def _tokens(name) -> list:
    """Return significant name tokens, removing:
    - known suffixes (jr, sr, ii, iii, etc.) from ANY position
    - middle initials (length 1 tokens between first and last tokens in a 3+ token name)
    """
    raw = _clean_name(name)
    parts = [p for p in raw.split() if p not in _SUFFIXES]
    if not parts:
        return []
        
    cleaned_parts = []
    if len(parts) >= 3:
        cleaned_parts.append(parts[0])
        for p in parts[1:-1]:
            if len(p) > 1:
                cleaned_parts.append(p)
        cleaned_parts.append(parts[-1])
    else:
        cleaned_parts = parts
        
    return cleaned_parts


def _make_key(name) -> str:
    """
    Canonical lookup key: tokens sorted alphabetically.
    This ensures 'First Last' and 'Last First' always produce the same key.
    """
    return " ".join(sorted(_tokens(name)))


def _lookup_keys(name) -> list:
    """
    Return candidate lookup keys. Since _make_key is now order-independent,
    we only need one canonical key.
    """
    key = _make_key(name)
    return [key] if key else []


def _is_valid_name(name) -> bool:
    text = str(name or '').strip().lower()
    if not text or text == 'nan':
        return False
    blocked = ('total', 'summary', 'record', 'employee details')
    return not any(b in text for b in blocked)


# ---------------------------------------------------------------------------
# Invoice Loading Helpers
# ---------------------------------------------------------------------------
def _detect_invoice_columns_llm(df: pd.DataFrame) -> dict:
    """Uses LLM to dynamically map unpredictable invoice column headers."""
    load_dotenv()
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        logger.warning("OPENAI_API_KEY missing. Skipping LLM invoice column detection.")
        return {}

    client = OpenAI(api_key=api_key)
    headers = list(df.columns)
    sample_data = df.head(3).to_dict(orient='records')
    
    system_prompt = (
        "You are a highly accurate data mapping expert. Your task is to map extracted invoice column headers "
        "to our standard internal fields. Use the provided sample data to understand the content of each column.\n\n"
        "Standard Fields:\n"
        "- full_name: The employee or subscriber's full name.\n"
        "- coverage: Coverage tier or level (e.g., EE, ES, FAM, EC).\n"
        "- plan_name: The name of the medical or dental plan being billed.\n"
        "- premium: The current billed premium amount (usually a dollar value).\n"
        "- zip: Zip code or postal code.\n\n"
        "CRITICAL RULES:\n"
        "1. Return ONLY a valid JSON object mapping our Standard Fields to the EXACT column headers from the file.\n"
        "2. Do NOT guess blindly. If a field truly does not exist in the file, omit it from the JSON."
    )
    
    user_prompt = (
        f"Column Headers:\n{headers}\n\n"
        f"Sample Data (First 3 rows):\n{json.dumps(sample_data, indent=2, default=str)}"
    )

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.0
        )
        mapping = json.loads(response.choices[0].message.content)
        valid_mapping = {k: v for k, v in mapping.items() if v in headers}
        logger.info(f"LLM Dynamic Invoice Mapping: {valid_mapping}")
        return valid_mapping
    except Exception as e:
        logger.error(f"LLM invoice column detection failed: {e}")
        return {}


def load_invoice(path: str) -> dict:
    """
    Returns dict keyed by ALL candidate lookup keys for each employee name.
    Dynamically maps columns using an LLM, falling back to static indices.
    """
    xl = pd.ExcelFile(path)
    sheet = next(
        (s for s in xl.sheet_names if any(k in s.lower() for k in ('employee', 'detail', 'data'))),
        xl.sheet_names[0]
    )

    # Detect header row
    probe = pd.read_excel(path, sheet_name=sheet, nrows=10, header=None)
    hrow = 0
    for i, row in probe.iterrows():
        if any(str(v).lower() in ('full name', 'first name', 'plan name', 'employee name', 'subscriber') for v in row if pd.notna(v)):
            hrow = i
            break

    df = pd.read_excel(path, sheet_name=sheet, skiprows=hrow)
    # Convert all columns to strings to avoid type issues during dict lookup
    df.columns = df.columns.astype(str)
    
    col_map = _detect_invoice_columns_llm(df)
    
    # Static Fallback mapping if LLM fails
    if not col_map:
        logger.warning("Falling back to static index-based invoice column mapping.")
        cols = list(df.columns)
        col_map = {
            'full_name': cols[0] if len(cols) > 0 else None,
            'coverage':  cols[4] if len(cols) > 4 else None,
            'plan_name': cols[5] if len(cols) > 5 else None,
            'premium':   cols[7] if len(cols) > 7 else None,
            'zip':       cols[11] if len(cols) > 11 else None
        }

    lookup = {}

    # Medical vs Ancillary filtering
    ANCILLARY_KEYWORDS = {'dental', 'vision', 'life', 'ltd', 'std', 'disability', 'ad&d', 'voluntary', 'accident', 'critical', 'hospital'}

    for _, row in df.iterrows():
        name_col = col_map.get('full_name')
        if not name_col:
            continue
            
        raw_full = str(row.get(name_col, '')).strip()
        if not raw_full or not _is_valid_name(raw_full) or raw_full == 'nan':
            continue

        # Canonical lookup key
        canon_keys = _lookup_keys(raw_full)
        if not canon_keys:
            continue
        primary_key = canon_keys[0]

        # If we already have a medical row for this person, skip
        if primary_key in lookup:
            continue

        cov_col  = col_map.get('coverage')
        plan_col = col_map.get('plan_name')
        prem_col = col_map.get('premium')
        zip_col  = col_map.get('zip')

        coverage_raw = row.get(cov_col) if cov_col else None
        plan_raw     = row.get(plan_col) if plan_col else None
        premium_raw  = row.get(prem_col) if prem_col else 0.0
        zip_raw      = row.get(zip_col) if zip_col else None

        if isinstance(premium_raw, str):
            premium_raw = re.sub(r'[^\d.]', '', premium_raw)
            try:    premium_raw = float(premium_raw)
            except: premium_raw = 0.0
        elif pd.isna(premium_raw):
            premium_raw = 0.0

        # Note: The $250 premium filter has been removed here.
        # For ADP/PEO invoices, Phase 1's PEO collapse already selected the single
        # best plan row per employee. Filtering by $250 at this stage would drop
        # employees whose best extracted row was Dental/Vision, making them appear
        # as "not available on invoice" even though they ARE on the invoice.

        zip_val = str(zip_raw).strip() if zip_raw and pd.notna(zip_raw) else ''

        entry = {
            'plan':     plan_raw,
            'premium':  premium_raw,
            'raw_name': raw_full,
            'coverage': coverage_raw,
            'zip':      zip_val,
        }

        # Store under canonical key
        lookup[primary_key] = entry
        # Store under reversed key if applicable to allow flexible lookup
        if len(canon_keys) > 1:
            lookup[canon_keys[1]] = entry

    unique = len({v['raw_name'] for v in lookup.values()})
    logger.info(f"Invoice loaded: {unique} employees (filtered for premium >= $250)")
    return lookup



# ---------------------------------------------------------------------------
# Step 2: Universal reference census loader (delegates to normalizer)
# ---------------------------------------------------------------------------
def load_ref_census(path: str) -> dict:
    """
    Universal census loader — delegates to census_normalizer.
    Returns dict keyed by canonical 'firstname lastname' lookup key.
    """
    normalized_data = census_normalizer.normalize_census_to_list(path)
    
    # Convert list of objects to the dictionary format expected by the filler
    result = {}
    for _, data in normalized_data.items():
        # Ensure the key is the canonical version used by tokens
        canon_key = _make_key(data.get('first', '') + ' ' + data.get('last', ''))
        result[canon_key] = data
        
    logger.info(f"Ref census loaded via normalizer: {len(result)} employees")
    return result


# ---------------------------------------------------------------------------
# Step 3: Detect RAPT template column layout
# ---------------------------------------------------------------------------
def _col_key(val) -> str:
    return re.sub(r'\s+', ' ', str(val or '').replace('*', '').strip().lower())


def detect_rapt_columns(ws) -> dict:
    for r in range(1, 25):
        row_keys = {c: _col_key(ws.cell(row=r, column=c).value)
                    for c in range(1, ws.max_column + 1)}
        if 'first name' not in row_keys.values():
            continue

        cols = {'header_row': r, 'data_start': r + 1}
        for c, key in row_keys.items():
            if   key == 'data row':                          cols['data_row'] = c
            elif key == 'first name':                        cols['first']    = c
            elif key == 'last name':                         cols['last']     = c
            elif 'gender' in key:                            cols['gender']   = c
            elif 'date of birth' in key or key == 'birth':  cols['dob']      = c
            elif 'zip' in key:                               cols['zip']      = c
            elif 'relationship' in key:                      cols['relation'] = c
            elif 'dependent of' in key:                      cols['dep_of']   = c
            elif 'coverage' in key and 'cobra' not in key:  cols['coverage'] = c
            elif 'cobra' in key:                             cols['cobra']    = c
            elif 'monthly' in key and 'premium' in key:     cols['premium']  = c
            elif 'plan' in key:                              cols['plan']     = c
            elif 'discrepanc' in key:                        cols['disc']     = c

        # Append Discrepancies column if absent
        if 'disc' not in cols:
            dc   = (cols.get('premium') or ws.max_column) + 1
            from openpyxl.utils import get_column_letter
            cell = ws.cell(row=r, column=dc)
            cell.value     = 'Discrepancies'
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border    = _BORDER
            ws.column_dimensions[get_column_letter(dc)].width = 22
            cols['disc'] = dc

        # Column widths
        from openpyxl.utils import get_column_letter
        for k, w in [('plan', 32), ('premium', 18), ('first', 15), ('last', 18), ('dob', 14)]:
            if k in cols:
                ws.column_dimensions[get_column_letter(cols[k])].width = w

        logger.info(f"RAPT columns: {cols}")
        return cols

    raise ValueError("RAPT header row not found (no 'First Name' column in first 25 rows)")


# ---------------------------------------------------------------------------
# Step 4: Write rows into RAPT template
# ---------------------------------------------------------------------------
def _wcell(ws, row, col, value, align=None, fmt=None, fill=None):
    if col is None:
        return
    cell         = ws.cell(row=row, column=col)
    cell.value   = value
    cell.font    = _CELL_FONT
    cell.border  = _BORDER
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt
    if fill:  cell.fill          = fill


def _write_dob(ws, row, col, dob):
    if col is None or dob is None:
        return
    try:
        ts = pd.Timestamp(dob)
        if pd.isna(ts):
            return
        cell               = ws.cell(row=row, column=col)
        cell.value         = ts.to_pydatetime()
        cell.font          = _CELL_FONT
        cell.border        = _BORDER
        cell.alignment     = _CENTER
        cell.number_format = 'MM/DD/YYYY'
    except Exception:
        pass


def write_employee_row(ws, row_idx, data_row_num, emp, inv, cols):
    zip_val = (inv.get('zip') or emp.get('zip') or '') if inv else emp.get('zip', '')

    _wcell(ws, row_idx, cols.get('data_row'), data_row_num,    _CENTER)
    _wcell(ws, row_idx, cols.get('first'),    emp['first'],    _LEFT)
    _wcell(ws, row_idx, cols.get('last'),     emp['last'],     _LEFT)
    _wcell(ws, row_idx, cols.get('gender'),   emp['gender'],   _CENTER)
    _write_dob(ws, row_idx, cols.get('dob'),  emp.get('dob'))
    _wcell(ws, row_idx, cols.get('zip'),      zip_val,         _CENTER)
    _wcell(ws, row_idx, cols.get('relation'), 'EE',            _CENTER)
    _wcell(ws, row_idx, cols.get('dep_of'),   '',              _CENTER)
    _wcell(ws, row_idx, cols.get('coverage'), emp['coverage'], _CENTER)
    _wcell(ws, row_idx, cols.get('cobra'),    'N',             _CENTER)

    # ── WAIVER ONLY (WO) SKIP ───────────────────────────────────────
    # If coverage is 'WO' (Waiver Only), no need to fill.
    # Leave Plan, Premium, and Discrepancy/Notes columns empty/blank.
    if emp.get('coverage') is not None and str(emp['coverage']).strip().upper() == 'WO':
        logger.info(f"  Skipping waiver row {row_idx}: {emp['first']} {emp['last']} (coverage='WO')")
        _wcell(ws, row_idx, cols.get('plan'),    None, _LEFT)
        _wcell(ws, row_idx, cols.get('premium'), None, _CENTER)
        _wcell(ws, row_idx, cols.get('disc'),    None, _CENTER)
        return

    if inv:
        # Check if the census plan description is just a generic election category
        cen_plan = emp.get('plan_desc')
        inv_plan = inv.get('plan')
        
        is_generic_census = False
        if cen_plan:
            cen_lower = str(cen_plan).strip().lower()
            if cen_lower in {'base', 'core', 'buy up', 'buy-up', 'waived', 'not eligible', 'high', 'low', 'standard', 'premium', 'basic'}:
                is_generic_census = True

        # Prefer the census plan name UNLESS it is a generic election category,
        # in which case the invoice plan name (if available) is much better.
        if cen_plan and not is_generic_census:
            plan_name = cen_plan
        else:
            plan_name = inv_plan or cen_plan or ''
            
        _wcell(ws, row_idx, cols.get('plan'), plan_name, _LEFT)
        if inv.get('premium') is not None:
            _wcell(ws, row_idx, cols.get('premium'), inv['premium'], _CENTER, '$#,##0.00')

        status = discrepancy_status(
            extracted_name          = f"{emp['first']} {emp['last']}",
            invoice_name            = inv['raw_name'],
            extracted_coverage_tier = emp['coverage'],
            invoice_coverage_tier   = inv['coverage'],
            name_is_matched         = True,
        )
        _wcell(ws, row_idx, cols.get('disc'), status, _CENTER,
               fill=_FILL_CORRECT if status == 'Correct' else _FILL_MISMATCH)
    else:
        _wcell(ws, row_idx, cols.get('disc'), 'not available on invoice',
               _CENTER, fill=_FILL_MISSING)


def write_dependent_row(ws, row_idx, data_row_num, dep, emp_row_num, cols):
    zip_val = dep.get('zip', '')
    _wcell(ws, row_idx, cols.get('data_row'), data_row_num,    _CENTER)
    _wcell(ws, row_idx, cols.get('first'),    dep['first'],    _LEFT)
    _wcell(ws, row_idx, cols.get('last'),     dep['last'],     _LEFT)
    _wcell(ws, row_idx, cols.get('gender'),   dep['gender'],   _CENTER)
    _write_dob(ws, row_idx, cols.get('dob'),  dep.get('dob'))
    _wcell(ws, row_idx, cols.get('zip'),      zip_val,         _CENTER)
    _wcell(ws, row_idx, cols.get('relation'), dep['relation'], _CENTER)
    _wcell(ws, row_idx, cols.get('dep_of'),   emp_row_num,     _CENTER)
    # Explicitly clear out Coverage, Cobra, Plan, Premium, and Disc for dependents
    # so junk data from the original template doesn't bleed through
    _wcell(ws, row_idx, cols.get('coverage'), None, _CENTER)
    _wcell(ws, row_idx, cols.get('cobra'),    None, _CENTER)
    _wcell(ws, row_idx, cols.get('plan'),     None, _LEFT)
    _wcell(ws, row_idx, cols.get('premium'),  None, _CENTER)
    _wcell(ws, row_idx, cols.get('disc'),     None, _CENTER)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------
def fill_rapt_template(invoice_path, ref_census_path, template_path, output_path):
    invoice_lookup = load_invoice(invoice_path)
    ref_lookup     = load_ref_census(ref_census_path)
    
    # Save standardized census for reference
    import os
    out_dir = os.path.dirname(output_path)
    std_census_path = os.path.join(out_dir, "STANDARDIZED_COMMON_CENSUS.xlsx")
    _save_standardized_census(ref_lookup, std_census_path)

    wb = load_workbook(template_path)
    ws = next(
        (wb[s] for s in wb.sheetnames
         if any(k in s.lower() for k in ('sheet', 'census', 'table', 'employee'))),
        wb.active
    )

    cols         = detect_rapt_columns(ws)
    write_row    = cols['data_start']
    data_row_num = 1

    matched = not_on_invoice = not_on_ref = 0
    seen_invoice_keys = set()   # canonical keys already written

    # --- Rows from reference census ---
    for canon_key, emp in ref_lookup.items():
        # Try all lookup-key variants to find invoice match
        inv = None
        for k in _lookup_keys(f"{emp['first']} {emp['last']}"):
            if k in invoice_lookup:
                inv = invoice_lookup[k]
                seen_invoice_keys.add(k)
                # Also mark the reversed key as seen to prevent duplicate append
                toks = k.split()
                if len(toks) == 2:
                    seen_invoice_keys.add(f"{toks[1]} {toks[0]}")
                break

        if inv:
            matched += 1
        else:
            not_on_invoice += 1

        emp_row_num = data_row_num
        write_employee_row(ws, write_row, data_row_num, emp, inv, cols)
        write_row += 1; data_row_num += 1

        for dep in emp.get('dependents', []):
            write_dependent_row(ws, write_row, data_row_num, dep, emp_row_num, cols)
            write_row += 1; data_row_num += 1

            # Mark this dependent's name as seen in the invoice so it is NOT
            # appended again as an invoice-only "Not on census" row.
            # This handles the case where the census lists someone as a dependent
            # (CH) but the invoice bills them as a standalone EE subscriber.
            for dep_k in _lookup_keys(f"{dep['first']} {dep['last']}"):
                if dep_k in invoice_lookup:
                    seen_invoice_keys.add(dep_k)
                    dep_toks = dep_k.split()
                    if len(dep_toks) == 2:
                        seen_invoice_keys.add(f"{dep_toks[1]} {dep_toks[0]}")
                    logger.info(
                        f"  Dependent '{dep['first']} {dep['last']}' found on invoice "
                        f"(key='{dep_k}') — marked as seen, will not be re-appended."
                    )


    # --- Invoice-only rows (on invoice but NOT in ref census) ---
    for inv_key, inv in invoice_lookup.items():
        if inv_key in seen_invoice_keys:
            continue
        # Skip if the reversed key was already written
        toks = inv_key.split()
        if len(toks) == 2 and f"{toks[1]} {toks[0]}" in seen_invoice_keys:
            continue

        not_on_ref += 1
        seen_invoice_keys.add(inv_key)   # prevent processing same person twice

        raw    = str(inv['raw_name']).strip()
        c_name = _clean_name(raw)
        t      = [x for x in c_name.split() if x not in _SUFFIXES]
        # Determine order and split:
        # If raw contains a comma, the format is 'Last, First Middle' which _clean_name already swaps to 'First Middle Last'.
        # Otherwise, the format is 'Last First Middle'.
        if ',' in raw:
            if len(t) >= 2:
                first = " ".join(t[:-1]).title()
                last = t[-1].title()
            else:
                first, last = (t[0].title() if t else raw), ''
        else:
            if len(t) >= 2:
                first = " ".join(t[1:]).title()
                last = t[0].title()
            else:
                first, last = (t[0].title() if t else raw), ''

        fake_emp = {
            'first': first, 'last': last, 'gender': '',
            'dob': None, 'zip': inv.get('zip', ''),
            'coverage': inv.get('coverage', ''),
            'dependents': [],
        }
        write_employee_row(ws, write_row, data_row_num, fake_emp, inv, cols)

        # Override discrepancy to NOT_ON_CENSUS_STATUS
        dc = ws.cell(row=write_row, column=cols['disc'])
        dc.value = NOT_ON_CENSUS_STATUS
        dc.fill  = _FILL_MISMATCH
        dc.alignment = _CENTER
        dc.font  = _CELL_FONT
        dc.border = _BORDER

        write_row += 1; data_row_num += 1

    wb.save(output_path)
    logger.info(
        f"Saved '{output_path}'.  "
        f"Matched={matched} | Not on invoice={not_on_invoice} | "
        f"Invoice-only (appended)={not_on_ref}"
    )


def _save_standardized_census(ref_lookup, output_path):
    """Saves the normalized census data to a flat Excel file for reference."""
    rows = []
    for canon_key, emp in ref_lookup.items():
        # Employee Row
        rows.append({
            'Employee/Dependent': 'Employee',
            'First Name': emp.get('first', ''),
            'Last Name': emp.get('last', ''),
            'Gender': emp.get('gender', ''),
            'DOB': emp.get('dob', ''),
            'Zip Code': emp.get('zip', ''),
            'Coverage Tier': emp.get('coverage', ''),
            'Relation': 'EE',
            'Dependent Of': ''
        })
        # Dependent Rows
        for dep in emp.get('dependents', []):
            rows.append({
                'Employee/Dependent': 'Dependent',
                'First Name': dep.get('first', ''),
                'Last Name': dep.get('last', ''),
                'Gender': dep.get('gender', ''),
                'DOB': dep.get('dob', ''),
                'Zip Code': dep.get('zip', ''),
                'Coverage Tier': '',
                'Relation': dep.get('relation', ''),
                'Dependent Of': f"{emp['first']} {emp['last']}"
            })
    
    df = pd.DataFrame(rows)
    try:
        df.to_excel(output_path, index=False)
        logger.info(f"Saved standardized common census to '{output_path}'")
    except Exception as e:
        logger.error(f"Failed to save standardized census: {e}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description='Fill RAPT census template from invoice + reference census'
    )
    parser.add_argument('invoice',    help='BENEFITS_BILLING xlsx')
    parser.add_argument('ref_census', help='TEPCensus xlsx')
    parser.add_argument('template',   help='RAPT_Census xlsx (empty template)')
    parser.add_argument('output', nargs='?', default='filled_rapt_output.xlsx')
    args = parser.parse_args()
    fill_rapt_template(args.invoice, args.ref_census, args.template, args.output)


if __name__ == '__main__':
    main()