
import openpyxl
from datetime import datetime

# =====================
# Load RAPT Census (the output template to verify)
# =====================
wb_rapt = openpyxl.load_workbook(r'census\RAPT Census - Texas Environmental Plastics 1.xlsx', data_only=True)
ws_rapt = wb_rapt.active

header_row = None
rapt_data = []
for i, row in enumerate(ws_rapt.iter_rows(values_only=True), start=1):
    if row[0] == 'Data Row *':
        header_row = i
    elif header_row and any(v is not None for v in row):
        rapt_data.append({'row_num': i, 'data': list(row)})

print(f'RAPT Census header at row: {header_row}')
print(f'RAPT Census data rows: {len(rapt_data)}')

# =====================
# Load filled_tep_v2 (the filled output)
# =====================
wb_filled = openpyxl.load_workbook(r'output\filled_tep_v2.xlsx', data_only=True)
ws_filled = wb_filled.active
filled_data = []
header_filled = None
for i, row in enumerate(ws_filled.iter_rows(values_only=True), start=1):
    if row[0] == 'Data Row *':
        header_filled = i
    elif header_filled and any(v is not None for v in row):
        filled_data.append({'row_num': i, 'data': list(row)})

print(f'Filled TEP header at row: {header_filled}')
print(f'Filled TEP data rows: {len(filled_data)}')
print()

# Columns: [0]=DataRow, [1]=FirstName, [2]=LastName, [3]=Gender, 
#           [4]=DOB, [5]=HomeZip, [6]=Relationship, [7]=DependentOf,
#           [8]=Coverage, [9]=Cobra, [10]=PlanName, [11]=Premium, [12]=Discrepancies

# =====================
# Load TEPCensus (reference)
# =====================
wb_tep = openpyxl.load_workbook(r'reference_census\TEPCensus.xlsx', data_only=True)
ws_tep = wb_tep.active
tep_rows = list(ws_tep.iter_rows(values_only=True))
# Col indices: 7=LastName, 8=FirstName, 2=Role, 10=DOB, 11=Gender, 15=PlanDesc, 16=PlanType, 17=Coverage
tep_people = {}
for row in tep_rows[1:]:
    last = str(row[7]).strip() if row[7] else ''
    first = str(row[8]).strip() if row[8] else ''
    role = row[2]
    dob = row[10]
    gender = row[11]
    plan_desc = row[15]
    key = (last.upper(), first.upper())
    if key not in tep_people:
        tep_people[key] = {
            'last': last, 'first': first, 'role': role,
            'dob': dob, 'gender': gender, 'plans': []
        }
    tep_people[key]['plans'].append(plan_desc)

print(f'TEP Census unique people: {len(tep_people)}')

# =====================
# Load Billing (source)
# =====================
wb_bill = openpyxl.load_workbook(r'sources\4-10-26_BENEFITS.BILLING_2621_202610_RPVE_20260422_111345.xlsx', data_only=True)
ws_bill = wb_bill.active
bill_rows = list(ws_bill.iter_rows(values_only=True))
bill_data = {}
for row in bill_rows[2:-1]:  # skip title, header and total
    fname = str(row[1]).strip().upper() if row[1] else ''
    lname = str(row[3]).strip().upper() if row[3] else ''
    plan = row[5]
    premium_str = str(row[7]).replace(',', '').strip() if row[7] else '0'
    try:
        premium = float(premium_str)
    except:
        premium = 0.0
    bill_data[(lname, fname)] = {'plan': plan, 'premium': premium}

print(f'Billing records: {len(bill_data)}')
print()

# =====================
# Verify filled_tep_v2 data
# =====================
print('=' * 80)
print('VERIFICATION OF filled_tep_v2.xlsx DATA')
print('=' * 80)

issues = []
ok_count = 0
not_avail_count = 0

for entry in filled_data:
    d = entry['data']
    row_num = entry['row_num']
    data_row = d[0]
    first = str(d[1]).strip() if d[1] else ''
    last = str(d[2]).strip() if d[2] else ''
    gender = d[3]
    dob = d[4]
    home_zip = d[5]
    relationship = d[6]
    dep_of = d[7]
    coverage = d[8]
    cobra = d[9]
    plan_name = d[10]
    premium = d[11]
    discrepancy = d[12]

    key_tep = (last.upper(), first.upper())
    
    row_issues = []
    
    # Check 1: Person exists in TEPCensus
    tep_match = tep_people.get(key_tep)
    if not tep_match:
        # Try partial match (last name only)
        last_matches = [v for k, v in tep_people.items() if k[0] == last.upper()]
        if last_matches:
            row_issues.append(f'Name mismatch: {first} {last} not exact match in TEPCensus (last name found)')
        else:
            row_issues.append(f'Person NOT found in TEPCensus: {first} {last}')
    else:
        # Check gender
        tep_gender = tep_match['gender']
        if tep_gender:
            expected_g = 'M' if tep_gender == 'Male' else ('F' if tep_gender == 'Female' else None)
            if expected_g and gender != expected_g:
                row_issues.append(f'Gender mismatch: filled={gender}, TEP={tep_gender}')
        elif gender is not None:
            # TEP has no gender, check if filled has one
            pass  # ok to have gender filled even if TEP doesn't

        # Check DOB
        tep_dob = tep_match['dob']
        if tep_dob and dob:
            if isinstance(dob, datetime) and isinstance(tep_dob, datetime):
                if dob.date() != tep_dob.date():
                    row_issues.append(f'DOB mismatch: filled={dob.date()}, TEP={tep_dob.date()}')

    # Check 2: Billing verification (employees only, not dependents)
    if relationship == 'EE' and discrepancy != 'not available on invoice':
        bill_key = (last.upper(), first.upper())
        # Try exact match
        bill_match = bill_data.get(bill_key)
        
        if not bill_match:
            # Try with Jr. handling
            for bk, bv in bill_data.items():
                if bk[0].replace(' JR.', '').replace(' JR', '') == last.upper().replace(' JR.', '').replace(' JR', ''):
                    if bk[1] == first.upper():
                        bill_match = bv
                        break
        
        if not bill_match:
            row_issues.append(f'NOT FOUND IN BILLING: {first} {last}')
        else:
            # Check plan name
            if plan_name and bill_match['plan']:
                if str(plan_name).strip().upper() != str(bill_match['plan']).strip().upper():
                    row_issues.append(f'Plan mismatch: filled="{plan_name}", billing="{bill_match["plan"]}"')
            # Check premium
            if premium is not None and bill_match['premium']:
                if abs(float(premium) - bill_match['premium']) > 0.01:
                    row_issues.append(f'Premium mismatch: filled={premium}, billing={bill_match["premium"]}')
    
    if relationship == 'EE' and discrepancy == 'not available on invoice':
        not_avail_count += 1
        # Verify truly not in billing
        bill_key = (last.upper(), first.upper())
        bill_match = bill_data.get(bill_key)
        if bill_match:
            row_issues.append(f'MARKED AS NOT ON INVOICE BUT FOUND IN BILLING: {first} {last}')

    if row_issues:
        issues.append({'row': row_num, 'data_row': data_row, 'name': first + ' ' + last, 'issues': row_issues})
    else:
        ok_count += 1

print(f'Total rows checked: {len(filled_data)}')
print(f'Rows with no issues: {ok_count}')
print(f'Rows "not available on invoice": {not_avail_count}')
print(f'Rows with issues: {len(issues)}')
print()

if issues:
    print('ISSUES FOUND:')
    print('-' * 60)
    for iss in issues:
        print(f'  Row {iss["row"]} (DataRow={iss["data_row"]}) - {iss["name"]}:')
        for msg in iss['issues']:
            print(f'    !! {msg}')
else:
    print('All rows verified successfully - no issues found.')

print()
# =====================
# Check if all billing employees are in filled output
# =====================
print('=' * 80)
print('BILLING COVERAGE CHECK: Are all billed employees in the filled output?')
print('=' * 80)
filled_employees = set()
for entry in filled_data:
    d = entry['data']
    if d[6] == 'EE':
        first = str(d[1]).strip().upper() if d[1] else ''
        last = str(d[2]).strip().upper() if d[2] else ''
        filled_employees.add((last, first))

for bill_key, bill_val in bill_data.items():
    lname, fname = bill_key
    found = (lname, fname) in filled_employees
    if not found:
        # Try Jr. case
        for fe in filled_employees:
            fe_last = fe[0].replace(' JR.', '').replace(' JR', '')
            bk_last = lname.replace(' JR.', '').replace(' JR', '')
            if fe_last == bk_last and fe[1] == fname:
                found = True
                break
    status = 'OK' if found else '!! MISSING FROM OUTPUT !!'
    print(f'  {lname}, {fname}: {status}')

print()
# =====================
# Column structure check
# =====================
print('=' * 80)
print('COLUMN STRUCTURE CHECK: filled_tep_v2 vs RAPT Census template')
print('=' * 80)
print(f'filled_tep_v2 columns: {ws_filled.max_column} (expected 13 with Discrepancies)')
print(f'RAPT Census columns: {ws_rapt.max_column} (expected 12, no Discrepancies)')
print()

# Check header row of both
for i, row in enumerate(ws_filled.iter_rows(values_only=True), start=1):
    if row[0] == 'Data Row *':
        print('filled_tep_v2 headers:')
        for j, col in enumerate(row):
            clean = str(col).split('\n')[0] if col else 'None'
            print(f'  Col {j+1}: {clean}')
        break

for i, row in enumerate(ws_rapt.iter_rows(values_only=True), start=1):
    if row[0] == 'Data Row *':
        print()
        print('RAPT Census headers:')
        for j, col in enumerate(row):
            clean = str(col).split('\n')[0] if col else 'None'
            print(f'  Col {j+1}: {clean}')
        break

print()
print('Home Zip Code check (should be populated):')
for entry in filled_data:
    d = entry['data']
    if d[6] == 'EE' and d[5] is None:
        print(f'  Row {entry["row_num"]} ({d[1]} {d[2]}): Home Zip Code is EMPTY')
