import os
import sys
import json
import argparse
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openai import OpenAI
from dotenv import load_dotenv

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Styling
_FONT = Font(name='Arial', size=10)
_FILL_LLM = PatternFill('solid', start_color='DDEBF7')  # Light blue
_CENTER = Alignment(horizontal='center', vertical='center')
_LEFT = Alignment(horizontal='left', vertical='center')

def run_llm_resolution(validated_excel: Path, audit_json: Path, output_excel: Path = None) -> dict:
    if not validated_excel.exists() or not audit_json.exists():
        logger.error("Missing input files for LLM resolution.")
        return {}

    with open(audit_json, 'r', encoding='utf-8') as f:
        audit_data = json.load(f)

    entries = audit_data.get('entries', [])
    unclaimed = audit_data.get('unclaimed_invoices', [])

    # Find unresolved census rows
    unresolved_census = []
    for entry in entries:
        if entry.get('action') in ['unresolved', 'flagged_possible']:
            # We skip 'deleted_duplicate' and 'kept_unresolved_appended' which are appended rows
            unresolved_census.append({
                'row': entry['row'],
                'raw_name': entry['raw_name']
            })

    if not unresolved_census or not unclaimed:
        logger.info("No unresolved census names or unclaimed invoices to process via LLM.")
        return {'status': 'skipped', 'matches': 0, 'output_path': str(validated_excel)}

    logger.info(f"LLM Resolution: {len(unresolved_census)} unresolved census names vs {len(unclaimed)} unclaimed invoices.")

    load_dotenv()
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        logger.error("OPENAI_API_KEY missing. Skipping LLM resolution.")
        return {}

    client = OpenAI(api_key=api_key)

    # Prepare Prompt
    system_prompt = (
        "You are an expert data matching assistant. Your job is to match unresolved names from a Census "
        "to available Unclaimed names from an Invoice.\n"
        "Be logical. Look for nicknames (e.g. Robert = Bob), severe typos, or swapped names.\n"
        "DO NOT guess wildly. The first and last names must reasonably correspond. DO NOT match completely different names just because they are left over.\n"
        "Return ONLY valid JSON in this exact format:\n"
        "{\n"
        "  \"matches\": [\n"
        "    {\"census_name\": \"...\", \"invoice_name\": \"...\"}\n"
        "  ]\n"
        "}\n"
        "If a census name cannot be confidently matched to any invoice name, do not include it in the output array."
    )

    user_prompt = (
        f"Unresolved Census Names:\n{json.dumps([c['raw_name'] for c in unresolved_census], indent=2)}\n\n"
        f"Unclaimed Invoice Names:\n{json.dumps([u['raw_name'] for u in unclaimed], indent=2)}"
    )

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1
        )
        content = response.choices[0].message.content
        match_data = json.loads(content)
        matches = match_data.get('matches', [])
    except Exception as e:
        logger.error(f"OpenAI API call failed: {e}")
        return {}

    if not matches:
        logger.info("LLM found no confident matches.")
        return {'status': 'completed', 'matches': 0, 'output_path': str(validated_excel)}

    logger.info(f"LLM proposed {len(matches)} matches.")

    # Build lookup maps
    invoice_map = {u['raw_name'].lower(): u for u in unclaimed}
    
    # Update Excel
    wb = load_workbook(str(validated_excel))
    ws = next(
        (wb[s] for s in wb.sheetnames
         if any(k in s.lower() for k in ('census', 'employee', 'table', 'sheet'))),
        wb.active
    )

    # Find columns
    plan_col, prem_col, disc_col = None, None, None
    for r in range(1, 40):
        for c in range(1, min(ws.max_column + 1, 60)):
            val = str(ws.cell(row=r, column=c).value or '').strip().lower()
            if 'plan' in val: plan_col = c
            if 'premium' in val: prem_col = c
            if 'discrep' in val: disc_col = c
        if plan_col and prem_col and disc_col:
            break

    match_count = 0
    target_invoice_raw_names = set()
    for match in matches:
        c_name = match.get('census_name')
        i_name = match.get('invoice_name')
        if not c_name or not i_name: continue

        c_name_lower = str(c_name).strip().lower()
        target_row = next((c['row'] for c in unresolved_census if str(c['raw_name']).strip().lower() == c_name_lower), None)
        target_invoice = invoice_map.get(str(i_name).strip().lower())

        if target_row and target_invoice and disc_col:
            logger.info(f"Applying match: Row {target_row} ({c_name}) -> Invoice ({target_invoice['raw_name']})")
            if plan_col and target_invoice.get('plan'):
                cell = ws.cell(row=target_row, column=plan_col)
                if not cell.value:
                    cell.value = target_invoice['plan']
                    cell.font = _FONT
                    cell.alignment = _LEFT

            if prem_col and target_invoice.get('premium') is not None:
                cell = ws.cell(row=target_row, column=prem_col)
                val = cell.value
                is_empty = val is None or str(val).strip() == '' or str(val).strip().upper() in ('#N/A', 'N/A', 'NA', '0', '0.0', '0.00')
                if is_empty:
                    cell.value = target_invoice['premium']
                    cell.font = _FONT
                    cell.alignment = _CENTER
                    cell.number_format = '$#,##0.00'

            cell = ws.cell(row=target_row, column=disc_col)
            cell.value = f"LLM Matched -> {target_invoice['raw_name']}"[:40]
            cell.fill = _FILL_LLM
            cell.font = _FONT
            cell.alignment = _CENTER

            cell.alignment = _CENTER

            match_count += 1
            # Keep track of successfully matched invoice names to delete their appended rows
            target_invoice_raw_names.add(target_invoice['raw_name'].strip().lower())
        else:
            logger.warning(f"Failed to apply match: {c_name} -> {i_name}. target_row={target_row}, target_invoice={bool(target_invoice)}, disc_col={disc_col}")

    # Delete the appended "Not on census" rows for the invoices we just matched
    rows_to_delete = set()
    if disc_col and target_invoice_raw_names:
        for row_idx in range(1, ws.max_row + 1):
            disc_val = str(ws.cell(row=row_idx, column=disc_col).value or "").strip().lower()
            if "not on census" in disc_val:
                first = str(ws.cell(row=row_idx, column=2).value or "").strip()
                last = str(ws.cell(row=row_idx, column=3).value or "").strip()
                appended_name = f"{first} {last}".strip().lower()
                if appended_name in target_invoice_raw_names:
                    rows_to_delete.add(row_idx)

    for r in sorted(list(rows_to_delete), reverse=True):
        ws.delete_rows(r)

    if rows_to_delete:
        logger.info(f"Deleted {len(rows_to_delete)} appended 'Not on census' rows.")

    if output_excel is None:
        output_excel = validated_excel.with_name(validated_excel.name.replace("VALIDATED_", "LLM_RESOLVED_"))
    
    wb.save(str(output_excel))
    logger.info(f"Saved LLM resolved Excel to {output_excel}")

    return {
        'status': 'completed',
        'matches': match_count,
        'output_path': str(output_excel)
    }

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Phase 4: LLM Fallback Resolver")
    parser.add_argument("validated_excel")
    parser.add_argument("audit_json")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    result = run_llm_resolution(
        Path(args.validated_excel),
        Path(args.audit_json),
        Path(args.output) if args.output else None
    )
    if result.get('output_path'):
        print(f"LLM Output: {result['output_path']}")
