"""
data_validation.py  —  Unified Post-Fill Name Normalisation & Discrepancy Resolver
====================================================================================
Last step in the pipeline — runs AFTER fill_template.py (type1 / type2 / type3).

Pipeline:
  Phase 1  →  Invoice PDF → extraction.xlsx   (raw invoice data)
  Phase 2  →  fill_template.py → filled_census.xlsx   (filled with strict name match)
  Phase 3  →  data_validation.py → validated_census.xlsx  (THIS FILE — fuzzy resolve)

What it does:
  1. Reads the filled Excel (output of fill_template).
  2. Also reads the original Phase 1 invoice Excel to get all invoice names.
  3. For every row still flagged as "Not on census" or "not available on invoice":
       a. Normalises both names (strips middle init, handles LAST FIRST / First Last)
       b. Pass 1 — exact canonical match
       c. Pass 2 — token-swap match (LAST FIRST  ↔  First Last)
       d. Pass 3 — fuzzy similarity match (threshold configurable)
  4. If a match is found: fills in plan/premium and updates Discrepancy cell.
  5. Saves a new validated Excel + JSON audit log.

Usage:
    python data_validation.py <filled_excel> <invoice_excel> [output] [--threshold 85]

    filled_excel   — output Excel from fill_template.py
    invoice_excel  — Phase 1 extraction Excel (raw invoice data)
    output         — (optional) output path, default: <filled>_validated.xlsx
    --threshold    — fuzzy match minimum score 0-100 (default 85)
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openai import OpenAI
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Optional fuzzy library — falls back to SequenceMatcher if not installed
# ---------------------------------------------------------------------------
try:
    from rapidfuzz import fuzz as _fuzz
    def _similarity(a: str, b: str) -> float:
        return _fuzz.token_sort_ratio(a, b)
except ImportError:
    try:
        from fuzzywuzzy import fuzz as _fuzz  # type: ignore
        def _similarity(a: str, b: str) -> float:
            return _fuzz.token_sort_ratio(a, b)
    except ImportError:
        from difflib import SequenceMatcher
        def _similarity(a: str, b: str) -> float:  # type: ignore
            return SequenceMatcher(None, a, b).ratio() * 100

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
_FONT         = Font(name='Arial', size=10)
_FONT_BOLD    = Font(name='Arial', size=10, bold=True)
_CENTER       = Alignment(horizontal='center', vertical='center')
_LEFT         = Alignment(horizontal='left',   vertical='center')
_FILL_GREEN   = PatternFill('solid', start_color='C6EFCE')   # correct
_FILL_YELLOW  = PatternFill('solid', start_color='FFEB9C')   # fuzzy / uncertain
_FILL_RED     = PatternFill('solid', start_color='FFC7CE')   # still unresolved
_FILL_ORANGE  = PatternFill('solid', start_color='FFD966')   # possible match

# Discrepancy status strings (matching what fill_template uses)
_NOT_ON_CENSUS   = "Not on census"
_NOT_ON_INVOICE  = "not available on invoice"
_CORRECT         = "Correct"

# Known suffixes / titles to strip from names
_STRIP_TOKENS = {
    'jr', 'sr', 'ii', 'iii', 'iv', 'v', 'esq', 'phd', 'md', 'dds',
    'mr', 'mrs', 'ms', 'dr', 'prof',
}


# ===========================================================================
# NAME NORMALISATION ENGINE
# ===========================================================================

def _clean(raw) -> str:
    """Lowercase, remove punctuation except spaces, collapse whitespace."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip().lower()
    # Handle "Last, First" comma format → "first last"
    if ',' in s:
        parts = [p.strip() for p in s.split(',')]
        if len(parts) >= 2:
            s = f"{parts[1]} {parts[0]}"
    s = re.sub(r"[^a-z\s]", " ", s)   # keep only letters and spaces
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _tokens(raw) -> list[str]:
    """
    Return significant name tokens after:
      - Lowercasing and punctuation removal
      - Dropping known suffixes/titles
      - Dropping middle initials (length 1 tokens between first and last tokens in a 3+ token name)
    """
    cleaned = _clean(raw)
    parts = [p for p in cleaned.split() if p not in _STRIP_TOKENS]
    if not parts:
        return []
        
    if len(parts) >= 3:
        result = [parts[0]]
        for p in parts[1:-1]:
            if len(p) == 1:
                continue  # drop middle initial
            result.append(p)
        result.append(parts[-1])
        return result
    return parts


def canonical(raw) -> str:
    """
    Return canonical form: always 'firstname lastname' order, stripped of
    middle initials and suffixes.

    Examples:
        'BAKER HOLLI M'   → 'holli baker'  (token-swap because invoice is LAST FIRST)
        'Holli Baker'     → 'holli baker'
        'Baker, Holli M'  → 'holli baker'
        'LORENZ LUCAS'    → 'lucas lorenz'  (swap)
    """
    toks = _tokens(raw)
    if not toks:
        return ""
    # We store canonical as 'first last' — but since we don't know
    # which is first, we return ALL permutations via lookup_keys()
    return " ".join(toks)


def lookup_keys(raw) -> list[str]:
    """
    Return all candidate lookup keys for a name:
      - Original token order
      - Reversed (for LAST FIRST ↔ First Last resolution)
    For 3+ token names after stripping, all permutations of first/last pair.
    """
    toks = _tokens(raw)
    if not toks:
        return []
    key = " ".join(toks)
    keys = [key]
    if len(toks) == 2 and toks[0] != toks[1]:
        keys.append(f"{toks[1]} {toks[0]}")   # reversed swap
    elif len(toks) >= 3:
        # For 3 tokens: try first+last pair only (middle stripped)
        swapped = f"{toks[-1]} {toks[0]}"
        keys.append(swapped)
    return list(dict.fromkeys(keys))  # unique, preserve order


# ===========================================================================
# INVOICE LOADER — reads Phase 1 extraction Excel
# ===========================================================================

def _detect_invoice_columns_llm(df: pd.DataFrame) -> dict:
    """Uses LLM to dynamically map unpredictable invoice column headers."""
    load_dotenv()
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        logger.warning("OPENAI_API_KEY missing. Skipping LLM invoice column detection.")
        return {}

    client = OpenAI(api_key=api_key)
    headers = list(df.columns)
    sample_data = df.head(3).to_dict(orient='records')
    
    system_prompt = (
        "You are a highly accurate data mapping expert. Your task is to map extracted invoice column headers "
        "to our standard internal fields. Use the provided sample data to understand the content of each column.\n\n"
        "Standard Fields:\n"
        "- full_name: The employee or subscriber's full name (or last name if split).\n"
        "- first_name: First name (if separate).\n"
        "- last_name: Last name (if separate).\n"
        "- coverage: Coverage tier or level (e.g., EE, ES, FAM, EC, Employee Only, Family, WO).\n"
        "- plan_name: The detailed plan/product name (e.g., 'UnitedHealthcare Choice Plus HDHP 5000', 'Delta Dental PPO 500').\n"
        "- plan_type: The plan CATEGORY/TYPE (e.g., 'Medical', 'Dental', 'Vision', 'Life', 'Choice Plus'). Look for columns like 'Coverage Type', 'Benefit Type', 'Plan Type'.\n"
        "- premium: The billed premium amount - look for columns like 'Premium', 'Monthly Premium', 'GHP EE Monthly Contribution', 'Total Cost', 'Amount', 'Monthly Cost'.\n\n"
        "CRITICAL RULES:\n"
        "1. Return ONLY a valid JSON object mapping our Standard Fields to the EXACT column headers from the file.\n"
        "2. Do NOT guess blindly. If a field truly does not exist in the file, omit it from the JSON.\n"
        "3. If names are split across columns (e.g. 'NAME' and 'Unnamed: 1'), map 'NAME' to 'full_name' or 'last_name' and 'Unnamed: 1' to 'first_name'.\n"
        "4. IMPORTANT: 'GHP EE Monthly Contribution' is a premium column - map it to 'premium'.\n"
        "5. IMPORTANT: 'Coverage Type' is a plan category column - map it to 'plan_type'.\n"
        "6. IMPORTANT: 'Coverage Level' is a coverage tier column - map it to 'coverage'."
    )
    
    user_prompt = (
        f"Column Headers:\n{headers}\n\n"
        f"Sample Data (First 3 rows):\n{json.dumps(sample_data, indent=2, default=str)}"
    )

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.0
        )
        mapping = json.loads(response.choices[0].message.content)
        valid_mapping = {k: v for k, v in mapping.items() if v in headers}
        logger.info(f"LLM Dynamic Invoice Mapping (Validation): {valid_mapping}")
        return valid_mapping
    except Exception as e:
        logger.error(f"LLM invoice column detection failed: {e}")
        return {}


def load_invoice_data(invoice_path: str | Path) -> dict[str, dict]:
    """
    Load Phase 1 extraction Excel. Returns dict keyed by ALL candidate
    lookup keys for each employee.

    Supports dynamic column detection — works with any Phase 1 output.
    """
    path = Path(invoice_path)
    if not path.exists():
        logger.error(f"Invoice file not found: {path}")
        return {}

    xl = pd.ExcelFile(str(path))
    # Prefer sheet named 'employee', 'detail', 'data', else first sheet
    sheet = next(
        (s for s in xl.sheet_names
         if any(k in s.lower() for k in ('employee', 'detail', 'data'))),
        xl.sheet_names[0]
    )

    # Auto-detect header row
    probe = pd.read_excel(str(path), sheet_name=sheet, nrows=15, header=None)
    hrow = 0
    for i, row in probe.iterrows():
        row_str = " ".join(str(v).lower() for v in row if pd.notna(v))
        if any(kw in row_str for kw in ('name', 'plan', 'premium', 'coverage', 'total cost')):
            hrow = i
            break

    df = pd.read_excel(str(path), sheet_name=sheet, skiprows=hrow)
    df.columns = [str(c).strip() for c in df.columns]

    # Try LLM Mapping
    col_map_llm = _detect_invoice_columns_llm(df)
    
    # Static detection (fallback/supplement)
    col_map: dict[str, str | None] = {
        'full': col_map_llm.get('full_name'),
        'first': col_map_llm.get('first_name'),
        'last': col_map_llm.get('last_name'),
        'plan': col_map_llm.get('plan_name'),
        'premium': col_map_llm.get('premium'),
        'coverage': col_map_llm.get('coverage'),
    }
    
    # Fill in gaps with static rules
    for col in df.columns:
        cl = col.lower()
        if not col_map['full'] and ('full' in cl and 'name' in cl): col_map['full'] = col
        if not col_map['full'] and ('employee' in cl and 'name' in cl): col_map['full'] = col
        if not col_map['first'] and (('first' in cl or 'fname' in cl) and 'name' in cl): col_map['first'] = col
        if not col_map['last'] and (('last' in cl or 'lname' in cl) and 'name' in cl): col_map['last'] = col
        if not col_map['plan'] and ('plan' in cl and ('name' in cl or 'desc' in cl)): col_map['plan'] = col
        if not col_map['premium'] and ('premium' in cl or 'current' in cl or 'total cost' in cl): col_map['premium'] = col
        if not col_map['coverage'] and ('coverage' in cl or 'tier' in cl): col_map['coverage'] = col

    # Ultimate fallback: treat column 0 as full name
    if not col_map['full'] and not col_map['first']:
        col_map['full'] = df.columns[0]

    # Clean whitespace and replace empty/whitespace strings with NaN so ffill works (non-destructive add-on)
    for col_key in ['full', 'first', 'last']:
        col_name = col_map.get(col_key)
        if col_name and col_name in df.columns:
            cleaned_series = df[col_name].astype(str).str.strip().replace({'': None, 'nan': None, 'None': None, '<NA>': None})
            df[col_name] = cleaned_series.ffill()

    lookup: dict[str, dict] = {}
    blocked = ('total', 'subtotal', 'grand total', 'summary', 'record')

    for _, row in df.iterrows():
        # Build raw name intelligently from all available components
        name_parts = []
        if col_map['first'] and pd.notna(row.get(col_map['first'])):
            name_parts.append(str(row[col_map['first']]).strip())
        if col_map['last'] and pd.notna(row.get(col_map['last'])):
            name_parts.append(str(row[col_map['last']]).strip())
        if col_map['full'] and pd.notna(row.get(col_map['full'])):
            val = str(row[col_map['full']]).strip()
            if val not in name_parts:
                name_parts.append(val)
        
        raw_name = " ".join(name_parts).strip()

        if not raw_name or any(b in raw_name.lower() for b in blocked):
            continue

        # Clean premium
        prem_raw = row.get(col_map['premium']) if col_map['premium'] else 0.0
        if isinstance(prem_raw, str):
            prem_raw = re.sub(r'[^\d.]', '', prem_raw)
            try:    prem_raw = float(prem_raw)
            except: prem_raw = 0.0
        elif pd.isna(prem_raw):
            prem_raw = 0.0

        # Business Rule Filter: Only keep plans with premium >= $250
        # Aligned with Phase 2 (fill_template_v2.py) which uses the same $250 threshold.
        if prem_raw < 200:
            continue

        # If we already have a medical row for this person, skip
        has_existing = False
        for key in lookup_keys(raw_name):
            if key in lookup:
                has_existing = True
                break
        if has_existing:
            continue

        entry = {
            'raw_name': raw_name,
            'plan':     row.get(col_map['plan'])     if col_map['plan']     else None,
            'premium':  prem_raw,
            'coverage': row.get(col_map['coverage']) if col_map['coverage'] else None,
            'tokens':   set(_tokens(raw_name)) # Store for Pass 2.5
        }

        # Store under all lookup keys
        for key in lookup_keys(raw_name):
            if key:
                lookup[key] = entry

    unique = len({v['raw_name'] for v in lookup.values()})
    logger.info(f"Invoice lookup built: {unique} unique employees (filtered for premium >= $200)")
    return lookup


# ===========================================================================
# FILLED EXCEL SCANNER — finds all discrepancy rows and column positions
# ===========================================================================

def _find_columns(ws) -> dict[str, int | None]:
    """
    Auto-detect key column positions from the header row of the filled Excel.
    Searches for the row with the most keyword matches.
    """
    best_cols = {}
    best_score = -1
    header_row = 1

    for r in range(1, 40):
        row_vals = {
            c: str(ws.cell(row=r, column=c).value or '').strip().lower()
            for c in range(1, min(ws.max_column + 1, 50))
        }
        joined = " ".join(row_vals.values())
        
        # Scoring this row as a potential header
        score = 0
        current_cols = {
            'name': None, 'first': None, 'last': None,
            'plan': None, 'premium': None, 'disc': None,
            'relation': None,   # NEW — Relationship / Relation column (EE / CH / SP)
            'coverage': None,   # NEW — Coverage / Tier column
        }
        
        for c, v in row_vals.items():
            if   ('employee' in v and 'name' in v) or ('full' in v and 'name' in v):
                current_cols['name'] = c; score += 2
            elif 'first' in v and 'name' in v:
                current_cols['first'] = c; score += 2
            elif 'last' in v and 'name' in v:
                current_cols['last'] = c; score += 2
            elif 'premium' in v:
                current_cols['premium'] = c; score += 1
            elif 'plan' in v:
                current_cols['plan'] = c; score += 1
            elif 'discrep' in v or v == 'notes':
                current_cols['disc'] = c; score += 3  # High weight for validation column
            elif 'relation' in v and 'discrep' not in v:  # catches 'Relationship', 'Relation', 'Relationship to Employee', etc.
                current_cols['relation'] = c; score += 1
            elif 'coverage' in v or 'tier' in v:
                current_cols['coverage'] = c; score += 1
        
        if score > best_score and (current_cols['first'] or current_cols['name']):
            best_score = score
            best_cols = current_cols
            header_row = r

    # Fallback: if plan is 11 (Type 1 default) and premium is None, assume column 12 is premium
    if best_cols.get('plan') == 11 and best_cols.get('premium') is None:
        best_cols['premium'] = 12
        logger.info("Fallback: set premium column to 12 based on plan column at 11")

    best_cols['header_row'] = header_row
    return best_cols


def _get_name_from_row(ws, row_idx: int, cols: dict) -> str:
    """Extract employee name from a worksheet row based on detected column layout."""
    if cols['name']:
        val = ws.cell(row=row_idx, column=cols['name']).value
        return str(val).strip() if val else ""
    if cols['first'] and cols['last']:
        f = str(ws.cell(row=row_idx, column=cols['first']).value or '').strip()
        l = str(ws.cell(row=row_idx, column=cols['last']).value  or '').strip()
        return f"{f} {l}".strip()
    return ""


# ===========================================================================
# MATCHING ENGINE
# ===========================================================================

def _has_initials(tokens: list[str]) -> bool:
    """Return True if at least one token looks like an initial (single letter)."""
    return any(len(t) == 1 for t in tokens)


def _initials_match(initial_toks: list[str], full_toks: list[str]) -> bool:
    """
    Check if initialed name tokens match full name tokens.
    Strategy:
      1. Last names must match exactly.
      2. Each initial (single-letter token) from the invoice must match
         the starting letter of at least one token in the census full name.
    Example: ['r', 'l', 'brown'] matches ['ruth', 'clark', 'brown']
             because last_name 'brown'=='brown', 'r' matches 'ruth', 'l' matches 'clark' (no — but let's be lenient)
    Actually: we match last name, then check each initial against the remaining full tokens.
    """
    if not initial_toks or not full_toks:
        return False

    # Extract last name from each (assume last token)
    inv_last = initial_toks[-1]
    cen_last = full_toks[-1]

    # Last name must match (exact or one contains the other for hyphenated)
    if inv_last != cen_last:
        # Handle hyphenated names: "clark-brown" contains "brown"
        if inv_last not in cen_last and cen_last not in inv_last:
            return False

    # Get the initials (all single-letter tokens) from the invoice name
    inv_initials = [t for t in initial_toks[:-1] if len(t) == 1]
    if not inv_initials:
        return False  # No initials to match — not an initial-based name

    # Get the non-last-name tokens from the census (first/middle names)
    cen_first_tokens = full_toks[:-1]
    if not cen_first_tokens:
        return False  # Census has no first name tokens

    # Check that the FIRST initial matches the start of at least one census token.
    # Additional initials (middle initials) are allowed but not required to match,
    # since census often lacks middle names (e.g., "J.A. Miller" vs "Jeffrey Miller").
    available = list(cen_first_tokens)
    first_matched = False
    total_matched = 0
    for i, ini in enumerate(inv_initials):
        matched = False
        for idx, ft in enumerate(available):
            if ft.startswith(ini):
                available.pop(idx)
                matched = True
                total_matched += 1
                break
        if i == 0 and not matched:
            return False  # First initial MUST match
        if i == 0:
            first_matched = True

    # Require at least the first initial matched
    return first_matched


def match_name(
    raw_name: str,
    invoice_lookup: dict[str, dict],
    threshold: float = 85.0,
) -> tuple[dict | None, str, float]:
    """
    Try to find invoice_lookup entry for raw_name using multi-pass strategy.

    Returns:
        (entry, match_type, confidence)
          entry      — the matched invoice dict, or None
          match_type — 'exact' | 'token_swap' | 'initial' | 'fuzzy' | 'none'
          confidence — 0-100
    """
    if not raw_name:
        return None, 'none', 0.0

    # --- Pass 1 & 2: exact canonical + token-swap ---
    raw_toks = _tokens(raw_name)
    raw_tok_set = set(raw_toks)
    
    for key in lookup_keys(raw_name):
        if key in invoice_lookup:
            return invoice_lookup[key], 'canonical', 100.0

    # --- Pass 2.5: Order-independent Token Set Match ---
    # This catches "Garcia Eileen" vs "Eileen Garcia" perfectly.
    for inv_key, entry in invoice_lookup.items():
        if entry.get('tokens') == raw_tok_set:
            return entry, 'token_swap', 100.0

    # --- Pass 2.75: Initial-to-Full-Name Match ---
    # Handles invoices that use initials (e.g. "R.L. Brown") vs census
    # full names (e.g. "Ruth Clark-Brown"). We try both directions:
    # census might have initials OR invoice might have initials.
    # We use _all_tokens_dedup which preserves single-letter middle tokens
    # (unlike _tokens which drops them as "middle initials") and deduplicates
    # doubled tokens from the concatenated raw_name format ("first last full_name").
    def _all_tokens_dedup(raw) -> list[str]:
        """Like _tokens but keeps ALL single-letter tokens (initials).
        Also deduplicates tokens that appear when raw_name concatenates
        first+last and full_name (e.g., 'R.L. Brown Brown, R.L.' -> ['r','l','brown'])."""
        cleaned = _clean(raw)
        all_toks = [p for p in cleaned.split() if p not in _STRIP_TOKENS]
        # Deduplicate: take shortest non-repeating prefix
        # e.g., ['r','l','r','l','brown','brown'] -> ['r','l','brown']
        # The raw_name is "first last full_name" so the first half is the canonical form
        if len(all_toks) >= 4:
            half = len(all_toks) // 2
            first_half = all_toks[:half]
            second_half = all_toks[half:]
            # Check if the second half is a permutation of the first half
            if sorted(first_half) == sorted(second_half):
                return first_half
        # For non-duplicated names, just deduplicate while preserving order
        seen = set()
        result = []
        for t in all_toks:
            if t not in seen:
                seen.add(t)
                result.append(t)
        return result

    raw_all_toks = _all_tokens_dedup(raw_name)
    if len(raw_all_toks) >= 2:
        # Collect ALL initial-match candidates first, then decide.
        # This prevents false positives when multiple people share
        # the same last name + first initial (e.g., "R. Smith" could
        # match "Robert Smith" or "Rachel Smith").
        initial_candidates = []
        seen_raw_names = set()
        for inv_key, entry in invoice_lookup.items():
            entry_raw = entry.get('raw_name', '')
            if entry_raw in seen_raw_names:
                continue
            seen_raw_names.add(entry_raw)

            inv_all_toks = _all_tokens_dedup(entry_raw)
            if not inv_all_toks or len(inv_all_toks) < 2:
                continue

            matched = False
            conf = 0.0

            # Case A: Invoice has initials, census has full names
            if _has_initials(inv_all_toks) and not _has_initials(raw_all_toks):
                if _initials_match(inv_all_toks, raw_all_toks):
                    matched, conf = True, 95.0

            # Case B: Census has initials, invoice has full names
            if not matched and _has_initials(raw_all_toks) and not _has_initials(inv_all_toks):
                if _initials_match(raw_all_toks, inv_all_toks):
                    matched, conf = True, 95.0

            # Case C: BOTH sides have initials (e.g., "Beatrice M Eley" vs "B.M. Eley")
            if not matched and _has_initials(inv_all_toks) and _has_initials(raw_all_toks):
                inv_ic = sum(1 for t in inv_all_toks if len(t) == 1)
                raw_ic = sum(1 for t in raw_all_toks if len(t) == 1)
                if inv_ic > raw_ic and _initials_match(inv_all_toks, raw_all_toks):
                    matched, conf = True, 90.0
                elif raw_ic > inv_ic and _initials_match(raw_all_toks, inv_all_toks):
                    matched, conf = True, 90.0

            if matched:
                initial_candidates.append((entry, conf))

        # Decision: only auto-resolve if exactly 1 unique candidate
        if len(initial_candidates) == 1:
            return initial_candidates[0][0], 'initial', initial_candidates[0][1]
        elif len(initial_candidates) > 1:
            # Multiple candidates — ambiguous, flag as 'possible' for human review
            # Pick the one with highest confidence, but cap at 75% to force review
            best = max(initial_candidates, key=lambda x: x[1])
            logger.warning(
                f"Initial match ambiguous for '{raw_name}': "
                f"{len(initial_candidates)} candidates found — flagging as possible"
            )
            return best[0], 'possible', 75.0

    # --- Pass 3: fuzzy across all known keys ---
    best_score = 0.0
    best_entry = None
    raw_canon = canonical(raw_name)

    for inv_key, entry in invoice_lookup.items():
        score = _similarity(raw_canon, inv_key)
        if score > best_score:
            best_score = score
            best_entry = entry

    if best_entry and best_score >= threshold:
        return best_entry, 'fuzzy', best_score

    if best_entry and best_score >= 50:
        return best_entry, 'possible', best_score

    return None, 'none', 0.0


# ===========================================================================
# COVERAGE NORMALISATION
# ===========================================================================

def canonical_coverage_tier(value) -> str:
    """
    Normalize common coverage-tier aliases across census and invoice files.
    Returns empty string if value is null or represents an empty/NA placeholder.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    
    # Normalize: lower, strip punctuation, remove spaces, upper
    raw = str(value).strip().lower()
    if not raw or raw in ('n/a', 'na', 'none', 'null', 'nan', '-', ''):
        return ""

    token = re.sub(r"[^a-z0-9]", "", raw).upper()

    tier_map = {
        "E": "EE", "EE": "EE", "EMPLOYEE": "EE", "EMPLOYER": "EE", "EMPLOYEEONLY": "EE",
        "S": "ES", "ES": "ES", "SPOUSE": "ES", "SS": "ES", "EMPLOYEESPOUSE": "ES", "EMPLOYEEANDSPOUSE": "ES",
        "C": "EC", "EC": "EC", "CH": "EC", "CHILD": "EC", "CHILDREN": "EC", "EMPLOYEECHILDREN": "EC", "EMPLOYEEANDCHILDREN": "EC",
        "F": "FAM", "FAM": "FAM", "FAMILY": "FAM", "EF": "FAM", "EMPLOYEEFAMILY": "FAM",
    }

    result = tier_map.get(token)
    if result:
        return result

    # Keyword fallback for unrecognised long-form strings
    if 'spouse' in raw and ('child' in raw or 'fam' in raw or '1+' in raw):
        return "FAM"
    if 'spouse' in raw or 'partner' in raw:
        return "ES"
    if 'child' in raw or '1+' in raw or 'dep' in raw:
        return "EC"
    if 'only' in raw or 'employee' in raw:
        return "EE"

    return token


# ===========================================================================
# MAIN VALIDATOR
# ===========================================================================

def run_validation(
    filled_path: str | Path,
    invoice_path: str | Path,
    output_path: str | Path | None = None,
    threshold: float = 85.0,
    template_type: str = 'type1',  # CH/SP skip applies ONLY for type1 (Engage)
) -> dict:
    """
    Core validation logic — works for type1, type2, and type3 filled Excels.

    Args:
        filled_path    — Excel produced by fill_template.py
        invoice_path   — Phase 1 extraction Excel
        output_path    — destination for validated Excel (default: _validated.xlsx)
        threshold      — fuzzy match confidence threshold (0-100)
        template_type  — 'type1' | 'type2' | 'type3'  (CH/SP skip only for type1)

    Returns:
        dict with validation statistics and audit log entries
    """
    filled_path  = Path(filled_path)
    invoice_path = Path(invoice_path)

    if output_path is None:
        output_path = filled_path.with_stem(filled_path.stem + "_validated")
    output_path = Path(output_path)

    if not filled_path.exists():
        logger.error(f"Filled Excel not found: {filled_path}")
        return {}
    if not invoice_path.exists():
        logger.error(f"Invoice Excel not found: {invoice_path}")
        return {}

    # Load invoice lookup
    invoice_lookup = load_invoice_data(invoice_path)
    if not invoice_lookup:
        logger.error("Invoice lookup is empty — cannot validate.")
        try:
            import shutil
            shutil.copyfile(str(filled_path), str(output_path))
            logger.info(f"Copied filled Excel to output path -> {output_path}")
            
            # Save a blank audit log JSON
            audit_path = output_path.with_suffix('.audit.json')
            empty_stats = {
                'total_rows':         0,
                'already_correct':    0,
                'resolved_canonical': 0,
                'resolved_fuzzy':     0,
                'still_unresolved':   0,
                'possible_matches':   0,
                'appended_deleted':   0,
            }
            with open(str(audit_path), 'w', encoding='utf-8') as f:
                json.dump({
                    'stats': empty_stats,
                    'entries': [],
                    'unclaimed_invoices': []
                }, f, indent=2, default=str)
            logger.info(f"Audit log saved -> {audit_path}")
            
            return {
                'stats':       empty_stats,
                'output_path': str(output_path),
                'audit_path':  str(audit_path),
                'audit_log':   [],
            }
        except Exception as e:
            logger.error(f"Failed to copy filled Excel or write audit log: {e}")
            return {}

    # Open filled workbook
    wb = load_workbook(str(filled_path))
    ws = next(
        (wb[s] for s in wb.sheetnames
         if any(k in s.lower() for k in ('census', 'employee', 'table', 'sheet'))),
        wb.active
    )

    col_positions = _find_columns(ws)
    disc_col = col_positions.get('disc')
    data_start = (col_positions.get('header_row') or 1) + 1

    # Fallback: if disc_col not found by header scan, search all cells for 'discrepan'
    if disc_col is None:
        for r in range(1, min(ws.max_row + 1, 50)):
            for c in range(1, min(ws.max_column + 1, 60)):
                val = str(ws.cell(row=r, column=c).value or '').strip().lower()
                if 'discrepan' in val:
                    disc_col = c
                    col_positions['disc'] = c
                    logger.info(f"Fallback: found Discrepancies column at col {c} (row {r})")
                    break
            if disc_col:
                break

    if disc_col is None:
        logger.warning(
            "Discrepancies column not found in the filled Excel — "
            "Phase 3 will skip row-level validation but will still save the workbook."
        )

    # ------------------------------------------------------------------
    # Scan all rows for discrepancy flags
    # ------------------------------------------------------------------
    audit_log: list[dict] = []
    stats = {
        'total_rows':         0,
        'already_correct':    0,
        'resolved_canonical': 0,
        'resolved_fuzzy':     0,
        'still_unresolved':   0,
        'possible_matches':   0,
        'appended_deleted':   0,
    }

    claimed_invoices = set()
    rows_to_delete = []

    # Relation values that identify a dependent (Child or Spouse).
    # If the template has a Relationship column and a row is CH/SP → skip.
    # Applies to ANY template type. If no Relationship column exists,
    # rel_col stays None and the guard below never fires.
    _DEPENDENT_RELATIONS = {'ch', 'sp', 'child', 'spouse', 'dependent', 'dep'}

    rel_col = col_positions.get('relation')  # None if template has no Relationship column

    for row_idx in range(data_start, ws.max_row + 1):
        # Stop at first completely empty row
        row_is_empty = all(
            ws.cell(row=row_idx, column=c).value is None
            for c in range(1, min(ws.max_column + 1, 6))
        )
        if row_is_empty:
            break

        # ── DEPENDENT SKIP (CH / SP) ─────────────────────────────────────
        # If the Relationship column exists and this row is a dependent,
        # skip it entirely.  Phase 2 never fills Plan / Premium / Discrepancy
        # for dependents, so attempting to validate them would be pointless
        # and could inflate "unresolved" counts incorrectly.
        if rel_col is not None:
            rel_val = str(ws.cell(row=row_idx, column=rel_col).value or '').strip().lower()
            if rel_val in _DEPENDENT_RELATIONS:
                logger.debug(f"  Row {row_idx}: skipping dependent row (relation='{rel_val.upper()}')")
                continue   # <-- nothing to do for CH / SP rows

        # Skip discrepancy logic if column was not found
        if disc_col is None:
            stats['total_rows'] += 1
            stats['still_unresolved'] += 1
            continue

        disc_cell = ws.cell(row=row_idx, column=disc_col)
        disc_val  = str(disc_cell.value or '').strip()
        raw_name  = _get_name_from_row(ws, row_idx, col_positions)

        if not raw_name and not disc_val:
            continue

        # ── WAIVER ONLY (WO) SKIP ─────────────────────────────────────────
        # If the coverage tier is 'WO', do not match or fill.
        # Clear out Plan, Premium, and Discrepancy/Notes columns.
        cov_col = col_positions.get('coverage')
        if cov_col is not None:
            cov_val = ws.cell(row=row_idx, column=cov_col).value
            if cov_val is not None and str(cov_val).strip().upper() == 'WO':
                logger.info(f"  Row {row_idx}: skipping waiver row (coverage='WO')")
                plan_col = col_positions.get('plan')
                prem_col = col_positions.get('premium')
                if plan_col:
                    ws.cell(row=row_idx, column=plan_col).value = None
                if prem_col:
                    ws.cell(row=row_idx, column=prem_col).value = None
                if disc_col:
                    ws.cell(row=row_idx, column=disc_col).value = None
                
                audit_entry = {
                    'row':            row_idx,
                    'raw_name':       raw_name,
                    'original_status': disc_val,
                    'match_type':     'waiver',
                    'confidence':     0.0,
                    'matched_to':     None,
                    'action':         'skipped_waiver'
                }
                audit_log.append(audit_entry)
                continue

        stats['total_rows'] += 1

        is_not_census  = _NOT_ON_CENSUS.lower()  in disc_val.lower()
        is_not_invoice = _NOT_ON_INVOICE.lower() in disc_val.lower()
        is_correct     = not (is_not_census or is_not_invoice)

        if is_not_census:
            # handle appended row
            if match_name(raw_name, invoice_lookup, threshold)[0] and match_name(raw_name, invoice_lookup, threshold)[0]['raw_name'] in claimed_invoices:
                rows_to_delete.append(row_idx)
                stats['appended_deleted'] += 1
            else:
                disc_cell.value      = _NOT_ON_CENSUS
                disc_cell.fill       = _FILL_RED
                disc_cell.font       = _FONT
                disc_cell.alignment  = _CENTER
                stats['still_unresolved'] += 1
            continue

        # Run name matching for everything else (Correct or not_on_invoice)
        match_entry, match_type, confidence = match_name(raw_name, invoice_lookup, threshold)

        audit_entry = {
            'row':            row_idx,
            'raw_name':       raw_name,
            'original_status': disc_val,
            'match_type':     match_type,
            'confidence':     round(confidence, 1),
            'matched_to':     match_entry['raw_name'] if match_entry else None,
            'action':         'unresolved'
        }

        # ── Determine Employee & Coverage Status ──────────────────────────
        emp_status = None
        matched_suffix = ""
        if match_type in ('canonical', 'token_swap'):
            emp_status = "Matched"
        elif match_type == 'initial':
            emp_status = f"Initial Match ({confidence:.0f}%)"
            matched_suffix = f" -> {match_entry['raw_name']}"
        elif match_type == 'fuzzy' and confidence >= threshold:
            emp_status = f"Fuzzy Match ({confidence:.0f}%)"
            matched_suffix = f" -> {match_entry['raw_name']}"
        elif match_type == 'possible':
            emp_status = f"Possible Match ({confidence:.0f}%)"
            matched_suffix = f" -> {match_entry['raw_name']}"

        cov_status = "not found on invoice"
        if match_entry and emp_status:
            inv_tier = canonical_coverage_tier(match_entry.get('coverage'))
            cen_tier = canonical_coverage_tier(ws.cell(row=row_idx, column=col_positions['coverage']).value if col_positions.get('coverage') else None)
            
            if not cen_tier:
                cov_status = "not found on census"
            elif not inv_tier:
                cov_status = "not found on invoice"
            elif inv_tier == cen_tier:
                cov_status = "Matched"
            else:
                cov_status = "Mismatched"

        # ── Apply Updates ──────────────────────────────────────────────
        if emp_status:
            label = f"Employee Verified: {emp_status}{matched_suffix} | Coverage Verified: {cov_status}"
            
            fill = _FILL_RED
            if "Matched" in emp_status and cov_status == "Matched":
                fill = _FILL_GREEN
            elif "Possible" in emp_status:
                fill = _FILL_ORANGE
            else:
                fill = _FILL_YELLOW

            # User hint: if Possible match (lower confidence), don't fill plan/premium
            fill_data = (match_type != 'possible')
            
            _apply_match(ws, row_idx, match_entry, col_positions, label[:100], fill, fill_data=fill_data)
            
            if is_correct:
                stats['already_correct'] += 1
                audit_entry['action'] = 'updated_correct'
                claimed_invoices.add(match_entry['raw_name']) # Already matched, keep claimed
            elif match_type == 'possible':
                stats['possible_matches'] += 1
                audit_entry['action'] = 'flagged_possible'
                # DO NOT add to claimed_invoices. This keeps the appended row at bottom 
                # so the user doesn't lose data, and Phase 4 (LLM) can try to match it.
            elif match_type in ('fuzzy', 'initial'):
                stats['resolved_fuzzy'] += 1
                audit_entry['action'] = 'resolved_fuzzy'
                claimed_invoices.add(match_entry['raw_name'])
            else:
                stats['resolved_canonical'] += 1
                audit_entry['action'] = 'resolved_canonical'
                claimed_invoices.add(match_entry['raw_name'])
        else:
            if is_not_invoice:
                disc_cell.value     = _NOT_ON_INVOICE
                disc_cell.fill      = _FILL_RED
                disc_cell.font      = _FONT
                disc_cell.alignment = _CENTER
                stats['still_unresolved'] += 1
                audit_entry['action'] = 'unresolved'
            elif is_correct:
                stats['already_correct'] += 1
                audit_entry['action'] = 'kept_correct_no_match'

        audit_log.append(audit_entry)

    # ------------------------------------------------------------------
    # Delete duplicate appended rows (safely backwards)
    # ------------------------------------------------------------------
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, amount=1)

    # ------------------------------------------------------------------
    # Save validated workbook
    # ------------------------------------------------------------------
    wb.save(str(output_path))
    logger.info(f"Validated Excel saved -> {output_path}")

    # ------------------------------------------------------------------
    # Save audit log JSON
    # ------------------------------------------------------------------
    audit_path = output_path.with_suffix('.audit.json')
    
    # Calculate unclaimed invoices
    unique_invoices = {entry['raw_name']: entry for entry in invoice_lookup.values()}
    unclaimed_invoices = [
        # Remove the 'tokens' set since it is not JSON serializable natively
        {k: v for k, v in entry.items() if k != 'tokens'} 
        for name, entry in unique_invoices.items() 
        if name not in claimed_invoices
    ]
    
    with open(str(audit_path), 'w', encoding='utf-8') as f:
        json.dump({
            'stats': stats, 
            'entries': audit_log,
            'unclaimed_invoices': unclaimed_invoices
        }, f, indent=2, default=str)
    logger.info(f"Audit log saved -> {audit_path}")

    # Summary
    logger.info(
        "\n" + "="*55 + "\n"
        "  VALIDATION SUMMARY\n"
        f"  Total rows scanned   : {stats['total_rows']}\n"
        f"  Already correct      : {stats['already_correct']}\n"
        f"  Resolved (canonical) : {stats['resolved_canonical']}\n"
        f"  Resolved (fuzzy)     : {stats['resolved_fuzzy']}\n"
        f"  Possible matches     : {stats['possible_matches']} (review needed)\n"
        f"  Still unresolved     : {stats['still_unresolved']}\n"
        f"  Appended deleted     : {stats['appended_deleted']} (removed duplicates)\n"
        + "="*55
    )

    return {
        'stats':       stats,
        'output_path': str(output_path),
        'audit_path':  str(audit_path),
        'audit_log':   audit_log,
    }


def _apply_match(ws, row_idx: int, entry: dict, cols: dict,
                 label: str, fill: PatternFill, fill_data: bool = True) -> None:
    """Write matched invoice data into the worksheet row and update Discrepancy cell."""
    disc_col = cols.get('disc')
    plan_col = cols.get('plan')
    prem_col = cols.get('premium')

    # Update Discrepancy cell
    if disc_col:
        cell            = ws.cell(row=row_idx, column=disc_col)
        cell.value      = label
        cell.fill       = fill
        cell.font       = _FONT
        cell.alignment  = _CENTER

    if not fill_data:
        return

    # Fill plan name if cell is empty
    if plan_col and entry.get('plan'):
        plan_cell = ws.cell(row=row_idx, column=plan_col)
        if not plan_cell.value:
            plan_cell.value     = entry['plan']
            plan_cell.font      = _FONT
            plan_cell.alignment = _LEFT

    # Fill premium if cell is empty or has placeholder value
    if prem_col and entry.get('premium') is not None:
        prem_cell = ws.cell(row=row_idx, column=prem_col)
        # Check if empty or contains #N/A or 0 placeholder
        val = prem_cell.value
        is_empty = val is None or str(val).strip() == '' or str(val).strip().upper() in ('#N/A', 'N/A', 'NA', '0', '0.0', '0.00')
        if is_empty:
            prem_cell.value         = entry['premium']
            prem_cell.font          = _FONT
            prem_cell.alignment     = _CENTER
            prem_cell.number_format = '$#,##0.00'


# ===========================================================================
# CLI ENTRY POINT
# ===========================================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Post-fill Data Validator — resolves name mismatches between "
            "census and invoice using fuzzy name normalisation. "
            "Works for type1, type2, and type3 filled Excels."
        )
    )
    parser.add_argument(
        'filled_excel',
        help='Path to the filled census Excel (output of fill_template.py)'
    )
    parser.add_argument(
        'invoice_excel',
        help='Path to the Phase 1 invoice extraction Excel'
    )
    parser.add_argument(
        'output', nargs='?', default=None,
        help='Output path for validated Excel (default: <filled>_validated.xlsx)'
    )
    parser.add_argument(
        '--threshold', type=float, default=85.0,
        help='Fuzzy match minimum confidence 0-100 (default: 85)'
    )
    parser.add_argument(
        '--template-type', dest='template_type', default='type1',
        choices=['type1', 'type2', 'type3'],
        help='Template type — CH/SP dependent-skip only applies to type1 (Engage). Default: type1'
    )
    args = parser.parse_args()

    result = run_validation(
        filled_path   = args.filled_excel,
        invoice_path  = args.invoice_excel,
        output_path   = args.output,
        threshold     = args.threshold,
        template_type = args.template_type,
    )

    if not result:
        return 1

    try:
        print(f"\n[OK] Validated Excel : {result['output_path']}")
        print(f"[OK] Audit log       : {result['audit_path']}")
        print(f"[OK] Stats           : {result['stats']}")
    except Exception:
        pass  # Suppress any encoding errors on Windows cp1252 terminals
    return 0


if __name__ == "__main__":
    sys.exit(main())
